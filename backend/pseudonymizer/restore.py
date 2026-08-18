"""仮名化の復元（再識別）。

仮名化 → ローカルで加工・統合 → **復元** という流れを支える。復元の入力は
「加工されたあとのファイル」なので、仮名化と違って列定義（ホワイトリスト）が使えない。
列は増減・改名され、複数種別が 1 ファイルに結合されている前提で処理する。

したがって復元は **テキストとしての置換** で行う。

  * 識別子: :data:`pseudonymizer.transforms` の採番結果（``.pseudonym_map.json``）から
    「仮名 → 元の値」の逆引き表を作り、長い仮名から順に単語境界つきで置換する。
  * 時刻: ソルトのオフセットを打ち消す方向にずらす。**仮名化が出力する形式に厳密に
    一致する文字列だけ** を対象にする（それらしい日付に手を出さない）。

復元できないもの:

  * マッピングに無い値（加工で生まれた集計値・新しいラベル）は **そのまま通す**。
  * ``vlan_id`` は仮名が裸の整数（``format_pseudonym`` が ``str(idx)``）であり、
    実データの数値と区別できない。テキスト置換では戻せないので対象外にする。
    VLAN を残したい場合は仮名化時に ``--keep-vlan`` を使う。
  * MAC は仮名化時にコロンなし小文字へ正規化される。元がコロン区切りだった場合、
    戻るのは正規化後の表記になる（プロジェクト規約に合わせた形）。

このモジュールは仮名化エンジン（``transforms.py`` / ``leakcheck.py`` / ``salt.py``）を
一切変更せず、その出力である仮名とソルトだけを読む。
"""
from __future__ import annotations

import csv
import io
import os
import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, Sequence

from .salt import DEFAULT_MAP_FILENAME, DEFAULT_SALT_FILENAME, SaltMaterial, load_salt
from .schemas import TransformType as T
from .transforms import MappingStore, format_pseudonym, load_mapping

#: 復元の対象にする変換型。``VLAN`` は仮名が裸の整数なので入れない（モジュール docstring 参照）。
RESTORABLE_TYPES: tuple[T, ...] = (
    T.SITE_ID,
    T.AP_ID,
    T.MAP_ID,
    T.SITE_NAME,
    T.AP_NAME,
    T.HOSTNAME,
    T.SSID,
    T.MAP_NAME,
    T.AP_MAC,
    T.CLIENT_MAC,
    T.IP,
)

#: 復元後であることを示す印（ファイル名の末尾）
RESTORED_SUFFIX = "_restored"

#: 仮名化が付ける印。復元時は外す（``_pseudonymized_restored`` は紛らわしい）
PSEUDONYMIZED_SUFFIX = "_pseudonymized"

#: テキストとして扱う拡張子
TEXT_EXTENSIONS: frozenset[str] = frozenset({".csv", ".tsv", ".json", ".txt", ".md", ".log"})
#: openpyxl でセル単位に扱う拡張子
XLSX_EXTENSIONS: frozenset[str] = frozenset({".xlsx"})
SUPPORTED_EXTENSIONS: frozenset[str] = TEXT_EXTENSIONS | XLSX_EXTENSIONS

#: 種別ごとの区切り文字（残存検出で列名を出すために使う）
_CSV_DELIMITERS = {".csv": ",", ".tsv": "\t"}

#: 置換件数のキー（識別子は :class:`TransformType` の値をそのまま使う）
COUNT_TIMESTAMP = "TIMESTAMP"
COUNT_TIMESTAMP_COMPACT = "TIMESTAMP_COMPACT"
COUNT_FILENAME = "FILENAME"

#: 仮名化が出力する時刻の形式だけを対象にする。それ以外の日付らしき文字列は触らない。
_TS_PATTERN = r"(?<!\d)\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}(?!\d)"
_TS_FORMAT = "%Y-%m-%d %H:%M:%S"
#: ファイル名に使われる ``YYYYMMDD_HHMM`` / ``YYYYMMDD_HHMMSS``
_TS_COMPACT_PATTERN = r"(?<!\d)\d{8}_(?:\d{6}|\d{4})(?!\d)"
_NAME_TIMESTAMP = re.compile(_TS_COMPACT_PATTERN)

#: 単語境界。仮名がより長いトークンの一部に食い込まないようにする。
_WORD = "[0-9A-Za-z_]"

_BOM = "﻿"

#: マッピングに無い「仮名らしき文字列」。残っていたら警告する（値そのものは出さない）。
RESIDUAL_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    ("AP_NAME", re.compile(rf"(?<!{_WORD})AP_\d{{3,}}(?!{_WORD})")),
    ("SITE_NAME", re.compile(rf"(?<!{_WORD})SITE_\d{{2,}}(?!{_WORD})")),
    ("MAP_NAME", re.compile(rf"(?<!{_WORD})FLOOR_\d{{2,}}(?!{_WORD})")),
    ("HOSTNAME", re.compile(rf"(?<!{_WORD})HOST_\d{{3,}}(?!{_WORD})")),
    ("SSID", re.compile(rf"(?<!{_WORD})SSID_\d{{2,}}(?!{_WORD})")),
    ("MAC", re.compile(rf"(?<!{_WORD})02[0-9a-f]{{10}}(?!{_WORD})")),
    ("UUID", re.compile(rf"(?<!{_WORD})[123]0000000-0000-4000-8000-\d{{12}}(?!{_WORD})")),
)

#: 残存の報告に載せる行番号の最大数（1 グループあたり）
_MAX_REPORTED_ROWS = 5


class RestoreError(RuntimeError):
    """復元処理を続行できない。"""


class UnsupportedFormatError(RestoreError):
    """対応していないファイル形式。"""


class MissingMaterialError(RestoreError):
    """ソルト／マッピングが無く、そもそも復元できない（呼び出し側で 400 にする）。"""


# ---------------------------------------------------------------------------
# レポート
# ---------------------------------------------------------------------------


@dataclass
class ResidualGroup:
    """マッピングに無い仮名らしき文字列。**値そのものは持たない。**"""

    kind: str
    #: 列名（分からなければ空文字）
    column: str
    count: int
    #: 該当行（先頭のいくつかだけ）
    rows: list[int] = field(default_factory=list)
    #: xlsx のシート名（それ以外は空文字）
    sheet: str = ""

    def describe(self) -> str:
        where = []
        if self.sheet:
            where.append(f"シート {self.sheet}")
        if self.column:
            where.append(f"列 {self.column}")
        rows = ", ".join(str(r) for r in self.rows)
        if rows:
            more = " ..." if self.count > len(self.rows) else ""
            where.append(f"行 {rows}{more}")
        return f"{self.kind}: {self.count} 件" + (f"（{' / '.join(where)}）" if where else "")


@dataclass
class FileReport:
    """1 ファイルの復元結果。"""

    source_name: str
    filename: str
    counts: dict[str, int] = field(default_factory=dict)
    residuals: list[ResidualGroup] = field(default_factory=list)

    @property
    def total_replacements(self) -> int:
        return sum(self.counts.values())

    @property
    def residual_total(self) -> int:
        return sum(g.count for g in self.residuals)

    def to_json(self) -> dict:
        return {
            "source_name": self.source_name,
            "filename": self.filename,
            "counts": dict(sorted(self.counts.items())),
            "total_replacements": self.total_replacements,
            "residuals": [
                {
                    "kind": g.kind,
                    "column": g.column,
                    "sheet": g.sheet,
                    "count": g.count,
                    "rows": list(g.rows),
                }
                for g in self.residuals
            ],
            "residual_total": self.residual_total,
        }


@dataclass
class RestoreReport:
    """複数ファイルをまとめた復元結果。"""

    files: list[FileReport] = field(default_factory=list)

    @property
    def counts(self) -> dict[str, int]:
        total: dict[str, int] = {}
        for f in self.files:
            for key, value in f.counts.items():
                total[key] = total.get(key, 0) + value
        return total

    @property
    def residual_total(self) -> int:
        return sum(f.residual_total for f in self.files)

    def to_json(self) -> dict:
        return {
            "files": [f.to_json() for f in self.files],
            "counts": dict(sorted(self.counts.items())),
            "residual_total": self.residual_total,
        }


#: 置換件数を人が読む形にするラベル
COUNT_LABELS: dict[str, str] = {
    T.AP_NAME.value: "AP名",
    T.AP_MAC.value: "AP MAC",
    T.AP_ID.value: "AP ID",
    T.SITE_NAME.value: "サイト名",
    T.SITE_ID.value: "サイト ID",
    T.MAP_NAME.value: "フロア名",
    T.MAP_ID.value: "フロア ID",
    T.CLIENT_MAC.value: "クライアント MAC",
    T.HOSTNAME.value: "ホスト名",
    T.SSID.value: "SSID",
    T.IP.value: "IP",
    COUNT_TIMESTAMP: "時刻",
    COUNT_TIMESTAMP_COMPACT: "時刻（YYYYMMDD_HHMM）",
    COUNT_FILENAME: "ファイル名の日付",
}


# ---------------------------------------------------------------------------
# 逆引き表
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class _Entry:
    ttype: T
    original: str


def build_reverse_index(mapping: MappingStore) -> dict[str, _Entry]:
    """「仮名 → (変換型, 元の値)」の逆引き表を作る。"""
    reverse: dict[str, _Entry] = {}
    for ttype in RESTORABLE_TYPES:
        for value, idx in (mapping.assignments.get(ttype) or {}).items():
            pseudonym = format_pseudonym(ttype, idx)
            existing = reverse.get(pseudonym)
            if existing is not None and existing.original != value:
                # 同じ仮名に別の値が割り当たっている＝マッピングが壊れている。
                # 戻し先を勝手に選ぶと静かに嘘のデータができるので止める。
                raise RestoreError(
                    f"mapping is inconsistent: pseudonym for {ttype.value} maps to "
                    "more than one original value"
                )
            reverse[pseudonym] = _Entry(ttype, value)
    return reverse


def _identifier_pattern(pseudonyms: Iterable[str]) -> str | None:
    """仮名の選択パターン。**長い順** に並べて部分一致の誤爆を防ぐ。

    Python の ``|`` は最長一致ではなく「先に書いた枝が勝つ」ので、
    ``SITE_0012`` を ``SITE_001`` より前に置く必要がある。
    """
    ordered = sorted(pseudonyms, key=lambda s: (-len(s), s))
    if not ordered:
        return None
    return "|".join(re.escape(p) for p in ordered)


# ---------------------------------------------------------------------------
# 復元エンジン
# ---------------------------------------------------------------------------


class RestoreEngine:
    """仮名化されたテキスト・xlsx を元の値に戻す。"""

    def __init__(
        self,
        material: SaltMaterial,
        mapping: MappingStore,
        *,
        time_restore: bool = True,
    ) -> None:
        self._offset = material.time_offset_seconds
        self._time_restore = time_restore
        self._reverse = build_reverse_index(mapping)
        self._pattern = self._compile()

    # -- パターン ---------------------------------------------------------

    def _compile(self) -> re.Pattern[str] | None:
        """識別子と時刻を **1 パスで** 置換するためのパターン。

        2 回に分けると、戻したあとの実データ（AP 名に日付が入っている等）を
        2 回目の走査が拾ってしまう。1 パスなら置換結果は再走査されない。
        """
        branches: list[str] = []
        ident = _identifier_pattern(self._reverse)
        if ident is not None:
            branches.append(rf"(?P<ident>(?<!{_WORD})(?:{ident})(?!{_WORD}))")
        if self._time_restore and self._offset:
            branches.append(rf"(?P<ts>{_TS_PATTERN})")
            branches.append(rf"(?P<tsc>{_TS_COMPACT_PATTERN})")
        if not branches:
            return None
        return re.compile("|".join(branches))

    # -- 時刻 -------------------------------------------------------------

    def _unshift(self, dt: datetime) -> datetime:
        """仮名化時のシフトを打ち消す（仮名化は ``dt + offset``）。"""
        return dt - timedelta(seconds=self._offset)

    def _restore_ts(self, text: str) -> str | None:
        try:
            dt = datetime.strptime(text, _TS_FORMAT)
        except ValueError:
            return None
        return self._unshift(dt).strftime(_TS_FORMAT)

    def _restore_ts_compact(self, text: str) -> str | None:
        _, _, time_part = text.partition("_")
        fmt = "%Y%m%d_%H%M%S" if len(time_part) == 6 else "%Y%m%d_%H%M"
        try:
            dt = datetime.strptime(text, fmt)
        except ValueError:
            return None  # 8 桁の連番など、日付として読めないものは触らない
        return self._unshift(dt).strftime(fmt)

    # -- テキスト ---------------------------------------------------------

    def restore_text(self, text: str) -> tuple[str, dict[str, int]]:
        """文字列を復元して (復元後, 種類ごとの置換件数) を返す。"""
        counts: dict[str, int] = {}
        if self._pattern is None:
            return text, counts

        def bump(key: str) -> None:
            counts[key] = counts.get(key, 0) + 1

        def replace(m: re.Match[str]) -> str:
            raw = m.group(0)
            groups = m.groupdict()
            if groups.get("ident") is not None:
                entry = self._reverse.get(raw)
                if entry is None:
                    return raw
                bump(entry.ttype.value)
                return entry.original
            if groups.get("ts") is not None:
                restored = self._restore_ts(raw)
                if restored is None:
                    return raw
                bump(COUNT_TIMESTAMP)
                return restored
            restored = self._restore_ts_compact(raw)
            if restored is None:
                return raw
            bump(COUNT_TIMESTAMP_COMPACT)
            return restored

        return self._pattern.sub(replace, text), counts

    # -- ファイル名 -------------------------------------------------------

    def restore_name(self, name: str) -> tuple[str, int]:
        """ファイル名の日付を中身と同じだけ戻し、``_restored`` を付ける。

        日付が見つからないファイル名（加工後に付け替えられた名前など）は
        エラーにせず、印だけ付ける。
        """
        stem, ext = os.path.splitext(name)
        shifted = 0
        if self._time_restore and self._offset:
            match = _NAME_TIMESTAMP.search(stem)
            if match is not None:
                restored = self._restore_ts_compact(match.group(0))
                if restored is not None:
                    stem = stem[: match.start()] + restored + stem[match.end() :]
                    shifted = 1
        if stem.endswith(PSEUDONYMIZED_SUFFIX):
            stem = stem[: -len(PSEUDONYMIZED_SUFFIX)]
        if not stem.endswith(RESTORED_SUFFIX):
            stem += RESTORED_SUFFIX
        return f"{stem}{ext}", shifted

    # -- ファイル ---------------------------------------------------------

    def restore_file(self, src: Path, out_dir: Path) -> FileReport:
        """1 ファイルを復元して ``out_dir`` に書き出す。入力は上書きしない。"""
        ext = src.suffix.lower()
        if ext not in SUPPORTED_EXTENSIONS:
            raise UnsupportedFormatError(
                f"対応していない形式です: {src.name}"
                f"（対応: {', '.join(sorted(SUPPORTED_EXTENSIONS))}）"
            )
        filename, name_count = self.restore_name(src.name)
        dst = out_dir / filename
        if os.path.realpath(dst) == os.path.realpath(src):
            raise RestoreError(f"出力が入力を上書きします: {src.name}")
        out_dir.mkdir(parents=True, exist_ok=True)

        if ext in XLSX_EXTENSIONS:
            counts, residuals = self._restore_xlsx(src, dst)
        else:
            counts, residuals = self._restore_text_file(src, dst, ext)
        if name_count:
            counts[COUNT_FILENAME] = name_count
        return FileReport(
            source_name=src.name, filename=filename, counts=counts, residuals=residuals
        )

    def _restore_text_file(
        self, src: Path, dst: Path, ext: str
    ) -> tuple[dict[str, int], list[ResidualGroup]]:
        data = src.read_bytes()
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError as e:
            raise RestoreError(
                f"UTF-8 として読めませんでした: {src.name}（{e.reason}）"
            ) from None
        had_bom = data.startswith(_BOM.encode("utf-8"))

        restored, counts = self.restore_text(text)
        residuals = scan_residuals(restored, delimiter=_CSV_DELIMITERS.get(ext))

        out = restored.encode("utf-8")
        if had_bom:
            out = _BOM.encode("utf-8") + out
        dst.write_bytes(out)
        return counts, residuals

    def _restore_xlsx(self, src: Path, dst: Path) -> tuple[dict[str, int], list[ResidualGroup]]:
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # pragma: no cover - requirements.txt に含まれている
            raise RestoreError("xlsx の復元には openpyxl が必要です") from e

        counts: dict[str, int] = {}
        residuals: list[ResidualGroup] = []
        wb = load_workbook(src)
        try:
            for ws in wb.worksheets:
                header: dict[int, str] = {}
                cells: list[tuple[int, int, str]] = []
                for row in ws.iter_rows():
                    for cell in row:
                        value = cell.value
                        if isinstance(value, str):
                            new_value, cell_counts = self.restore_text(value)
                            if cell_counts:
                                cell.value = new_value
                                for key, n in cell_counts.items():
                                    counts[key] = counts.get(key, 0) + n
                            if cell.row == 1:
                                header[cell.column] = new_value
                            else:
                                cells.append((cell.row, cell.column, new_value))
                        elif isinstance(value, datetime) and self._time_restore and self._offset:
                            # Excel で開いた時点で日時セルになっているものも戻す
                            cell.value = self._unshift(value)
                            counts[COUNT_TIMESTAMP] = counts.get(COUNT_TIMESTAMP, 0) + 1
                        elif cell.row == 1 and value is not None:
                            header[cell.column] = str(value)
                residuals.extend(_group_residuals(_scan_cells(cells, header), sheet=ws.title))
            wb.save(dst)
        finally:
            wb.close()
        return counts, residuals


# ---------------------------------------------------------------------------
# 残存の検出（値そのものは絶対に持ち出さない）
# ---------------------------------------------------------------------------


def _find_kinds(text: str) -> list[str]:
    found: list[str] = []
    for kind, pattern in RESIDUAL_PATTERNS:
        found.extend(kind for _ in pattern.finditer(text))
    return found


def _scan_cells(
    cells: Sequence[tuple[int, int, str]], header: dict[int, str]
) -> list[tuple[str, str, int]]:
    """(行, 列番号, 値) の列から (種類, 列名, 行) を拾う。列番号は 1 始まり。"""
    hits: list[tuple[str, str, int]] = []
    for row_no, col_no, value in cells:
        column = header.get(col_no) or f"col{col_no}"
        for kind in _find_kinds(value):
            hits.append((kind, column, row_no))
    return hits


def _group_residuals(
    hits: Sequence[tuple[str, str, int]], *, sheet: str = ""
) -> list[ResidualGroup]:
    groups: dict[tuple[str, str], ResidualGroup] = {}
    for kind, column, row_no in hits:
        key = (kind, column)
        group = groups.get(key)
        if group is None:
            group = ResidualGroup(kind=kind, column=column, count=0, sheet=sheet)
            groups[key] = group
        group.count += 1
        if len(group.rows) < _MAX_REPORTED_ROWS and row_no not in group.rows:
            group.rows.append(row_no)
    return sorted(groups.values(), key=lambda g: (g.kind, g.column))


def scan_residuals(text: str, *, delimiter: str | None = None) -> list[ResidualGroup]:
    """復元後のテキストに残った「仮名らしき文字列」を探す。

    区切り文字が分かる形式（csv / tsv）は列名まで出す。それ以外は行番号だけ。
    **戻り値に値そのものは含めない**（警告としてそのまま表示されるため）。
    """
    if delimiter is None:
        hits = [
            (kind, "", line_no)
            for line_no, line in enumerate(text.splitlines(), start=1)
            for kind in _find_kinds(line)
        ]
        return _group_residuals(hits)

    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)
    try:
        rows = list(reader)
    except csv.Error:
        # 壊れた csv は行ベースにフォールバックする（復元自体は済んでいる）
        return scan_residuals(text, delimiter=None)
    if not rows:
        return []
    header = {i: name for i, name in enumerate(rows[0], start=1)}
    cells = [
        (row_no, col_no, value)
        for row_no, row in enumerate(rows[1:], start=2)
        for col_no, value in enumerate(row, start=1)
    ]
    return _group_residuals(_scan_cells(cells, header))


# ---------------------------------------------------------------------------
# エンジンの組み立て
# ---------------------------------------------------------------------------


def default_paths() -> tuple[str, str]:
    """サーバ側と同じソルト・マッピングの置き場所。"""
    from . import service  # 循環 import を避けるため遅延（service は cli を読む）

    return service.SALT_PATH, service.MAP_PATH


def load_engine(
    salt_path: str | None = None,
    map_path: str | None = None,
    *,
    time_restore: bool = True,
) -> RestoreEngine:
    """ソルトとマッピングを読んで復元エンジンを作る。

    どちらも機密ファイルであり、**HTTP から取得できる場所に置いてはならない**
    （:mod:`pseudonymizer.service` の docstring 参照）。
    """
    if salt_path is None or map_path is None:
        default_salt, default_map = default_paths()
        salt_path = salt_path or default_salt
        map_path = map_path or default_map
    if not os.path.exists(salt_path):
        raise MissingMaterialError(
            f"ソルトファイルが見つかりません: {salt_path}\n"
            f"  仮名化に使ったソルト（{DEFAULT_SALT_FILENAME}）が無いと復元できません。"
        )
    material = load_salt(salt_path)
    if not os.path.exists(map_path):
        raise MissingMaterialError(
            f"マッピングファイルが見つかりません: {map_path}\n"
            f"  仮名化に使ったマッピング（{DEFAULT_MAP_FILENAME}）が無いと復元できません。"
        )
    mapping = load_mapping(map_path, material)
    return RestoreEngine(material, mapping, time_restore=time_restore)
