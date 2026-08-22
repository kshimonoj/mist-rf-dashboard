"""CLI と API が共用する分析パイプラインと出力の書き出し。

``rrm/cli.py`` と ``backend/routers/rrm.py`` は、どちらもこのモジュールだけを
呼ぶ。読み込み（:mod:`rrm.loader`）・分類（:mod:`rrm.events`）・前後サンプルの
突合（:mod:`rrm.metrics`）のロジックはここでも再実装しない。
**CLI と UI で結果が食い違わないこと** がこのモジュールの存在理由である。

出力の列名は英字にしてある。仮名化（``pseudonymizer``）へ後から乗せるとき、
日本語の列名は leak check（非 ASCII の検出）と相性が悪いため。

ネットワークアクセス・LLM 呼び出しは行わない。
"""
from __future__ import annotations

import math
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import events as ev
from . import loader
from . import metrics as met

#: 結果 CSV の列（**この順で出す**。CLI / API / 保存済み結果で同一）
RESULT_COLUMNS: tuple[str, ...] = (
    "event_timestamp", "classification", "reason", "site_name", "ap_name", "ap_mac", "band",
    "pre_channel", "post_channel", "channel_changed",
    "before_timestamp", "after_timestamp", "match_status", "contaminated",
    "clients_before", "clients_after", "clients_delta",
    "util_24_before", "util_24_after", "util_24_delta",
    "util_5_before", "util_5_after", "util_5_delta",
    "util_6_before", "util_6_after", "util_6_delta",
    "impact_clients",
)

#: 出力の列名の接頭辞 → ap_metrics の列名。**3 バンドすべてを必ず出す**
#: （イベントの band に関わらず。サンプルで全件ゼロでも列は落とさない）
VALUE_FIELDS: tuple[tuple[str, str], ...] = (
    ("clients", "num_clients"),
    ("util_24", "radio_24_utilization"),
    ("util_5", "radio_5_utilization"),
    ("util_6", "radio_6_utilization"),
)

#: 端末数として整数で扱う列の接頭辞
_INT_PREFIXES: frozenset[str] = frozenset({"clients"})

#: 時間帯別集計のバケット幅（秒）。1 時間固定
BUCKET_SECONDS: int = 3600

#: 連続したバケットを埋める上限。超えたら **データのあるバケットだけ** にする
MAX_BUCKETS: int = 24 * 31

#: xlsx のグラフに出すバケット数の上限（直近から）。**超過分は必ず明記する**
CHART_MAX_BUCKETS: int = 48

#: AP 別集計に出す上限（多い順）
TOP_AP_COUNT: int = 30

#: 受け付ける時刻表記（これで解釈できなければ pandas に委ねる）
TIME_FORMATS: tuple[str, ...] = ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S")

#: 進捗フェーズ（「まだ動いている」と伝えるためのもの。進捗率ではない）
PHASE_LOADING = "loading"
PHASE_EVENTS = "events"
PHASE_METRICS = "metrics"
PHASE_AGGREGATE = "aggregate"
PHASE_WRITING = "writing"

#: json の書式バージョン
META_VERSION = 1

_TITLE_FONT = Font(bold=True, size=14)
_HEADER_FONT = Font(bold=True)

TITLE = "RRM / RADAR チャネル変更分析"


class AnalysisError(RuntimeError):
    """分析を続行できない状態（CLI は終了コード 1、API は 400 / failed）。"""


class ParamError(AnalysisError):
    """パラメータが不正。``field_name`` にどの項目が不正かを持つ。"""

    def __init__(self, message: str, field_name: str | None = None) -> None:
        super().__init__(message)
        self.field_name = field_name


# loader 側の失敗も呼び出し側からは同じ扱いにする（CLI は 1、API は failed）
LoadError = loader.LoadError
NoMetricsError = loader.NoMetricsError
NoEventsError = loader.NoEventsError
SiteNotFoundError = loader.SiteNotFoundError
UnclassifiedInputError = loader.UnclassifiedInputError


# ---------------------------------------------------------------------------
# パラメータ
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class AnalysisParams:
    """分析条件。

    :param sites: 対象サイト（site_id または site_name）。**複数指定できる**。
        空なら全サイト（floorpeak と違い「サイト全体のピーク」のような単一サイト
        前提の定義が無く、サイト別比較を出すため）。
    """

    sites: tuple[str, ...] = ()
    window_start: pd.Timestamp | None = None
    window_end: pd.Timestamp | None = None


def parse_time(text: str, label: str, field_name: str | None = None) -> pd.Timestamp:
    """``YYYY-MM-DD HH:MM`` / ISO8601 を naive な Timestamp にする（TZ 付きは拒否）。"""
    text = str(text).strip()
    for fmt in TIME_FORMATS:
        try:
            return pd.Timestamp(datetime.strptime(text, fmt))
        except ValueError:
            continue
    try:
        ts = pd.Timestamp(text)
    except Exception as exc:  # 多様な例外を投げうる外部入力の境界
        raise ParamError(f"{label} を解釈できません: {text!r}", field_name) from exc
    if pd.isna(ts):
        raise ParamError(f"{label} を解釈できません: {text!r}", field_name)
    if ts.tzinfo is not None:
        raise ParamError(
            f"{label} にタイムゾーンは付けられません（ログが naive のため）: {text!r}",
            field_name,
        )
    return ts


# ---------------------------------------------------------------------------
# 整形ヘルパ
# ---------------------------------------------------------------------------


def fmt_dt(value: object) -> str:
    if value is None:
        return "-"
    try:
        if pd.isna(value):
            return "-"
    except (TypeError, ValueError):
        pass
    return pd.Timestamp(value).strftime("%Y-%m-%d %H:%M:%S")


def _dt_or_none(value: object) -> str | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return pd.Timestamp(value).strftime("%Y-%m-%d %H:%M:%S")


def _dt_or_blank(value: object) -> str:
    return _dt_or_none(value) or ""


def _num(value: object) -> float | None:
    """JSON / CSV に載せられる数値へ落とす（NaN・NA は None）。"""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    return None if math.isnan(out) else out


def _int(value: object) -> int | None:
    out = _num(value)
    return None if out is None else int(round(out))


def _cell_argb(rgb: str) -> str:
    """セルの塗りつぶし用に ARGB（不透明）へ変換する。

    ``PatternFill`` に 6 桁の RGB をそのまま渡すと openpyxl が alpha を ``00``
    （完全透明）で補って塗りが見えなくなる。セルの塗りは常にこれを通すこと。
    """
    return f"FF{rgb}"


def class_color(classification: object) -> str:
    """分類 → 色（RRGGBB）。未知の分類でも落ちないこと（灰色に落とす）。"""
    return ev.CLASS_COLORS.get(str(classification), "9E9E9E")


def condition_text(
    params: AnalysisParams,
    loaded: loader.RrmLoad,
    n_files: int,
) -> str:
    """分析条件のテキスト。

    **保存済み結果を後から見たときに「いつの・どこの・何の値か」が分かること。**
    """
    sites = ", ".join(loaded.site_labels) if loaded.site_labels else "すべて"
    left = fmt_dt(params.window_start) if params.window_start is not None else "(指定なし)"
    right = fmt_dt(params.window_end) if params.window_end is not None else "(指定なし)"
    return (
        f"分析条件: 対象サイト={sites} / "
        f"期間 {left} 〜 {right}（半開区間 [start, end)） / "
        f"バケット幅={BUCKET_SECONDS} 秒 / "
        f"サンプリング間隔={loaded.interval_seconds:g} 秒"
        f"（{'推定' if loaded.interval_estimated else '既定値'}） / "
        f"照合しきい値={met.MAX_GAP_FACTOR:g}倍 / "
        f"レーダー突合={ev.RADAR_MATCH_SECONDS:g} 秒 / "
        f"走査ファイル数={n_files} / "
        f"使用ap_metricsファイル数={loader.used_files(loaded.report, loader.METRICS_FILE_TYPES)} / "
        f"使用ap_eventsファイル数={loader.used_files(loaded.report, [loader.EVENTS_FILE_TYPE])}"
    )


# ---------------------------------------------------------------------------
# 明細行
# ---------------------------------------------------------------------------


def _empty_rows() -> pd.DataFrame:
    df = pd.DataFrame(columns=list(RESULT_COLUMNS))
    for column in RESULT_COLUMNS:
        if column.startswith("clients") or column in ("pre_channel", "post_channel", "impact_clients"):
            df[column] = df[column].astype("Int64")
        elif column.startswith("util_"):
            df[column] = df[column].astype("float64")
        elif column in ("channel_changed", "contaminated"):
            df[column] = df[column].astype(bool)
    return df


def build_rows(
    actions: pd.DataFrame,
    index: met.MetricIndex,
    change_index: met.ChangeEventIndex,
    interval_seconds: float,
) -> pd.DataFrame:
    """``AP_RRM_ACTION`` の 1 行ごとに前後サンプルを突き合わせ、明細行を作る。

    - **no-op（``pre_channel == channel``）の行も残す。** ``channel_changed`` 列で
      区別できるようにし、集計側で「チャネル変更回数」と「no-op 回数」を分ける。
    - **汚染した行も残す。** ``contaminated`` に印を付けるだけで、除外しない。
    - ``match_status != ok`` の行は **差分（``*_delta``）を空にする**。前後の生値と
      その時刻は残す（なぜ照合できなかったかを読み手が確かめられるようにするため）。
    - ``impact_clients`` は ``clients_before``（変更前に接続していた端末数）。
    """
    if actions.empty:
        return _empty_rows()

    records: list[dict[str, Any]] = []
    for row in actions.itertuples(index=False):
        mac = str(getattr(row, "ap_mac", "") or "")
        band = str(getattr(row, "band", "") or "")
        match = index.match(mac, row.event_timestamp, interval_seconds=interval_seconds)
        contaminated = change_index.is_contaminated(mac, row.event_timestamp, band, match)

        record: dict[str, Any] = {
            "event_timestamp": _dt_or_blank(row.event_timestamp),
            "classification": row.classification,
            "reason": str(getattr(row, "reason", "") or ""),
            "site_name": str(getattr(row, "site_name", "") or ""),
            "ap_name": str(getattr(row, "ap_name", "") or ""),
            "ap_mac": mac,
            "band": str(getattr(row, "band", "") or ""),
            "pre_channel": _int(row.pre_channel_num),
            "post_channel": _int(row.post_channel_num),
            "channel_changed": bool(row.channel_changed),
            "before_timestamp": _dt_or_blank(match.before_timestamp),
            "after_timestamp": _dt_or_blank(match.after_timestamp),
            "match_status": match.status,
            "contaminated": bool(contaminated),
        }
        for prefix, column in VALUE_FIELDS:
            before = _num(match.before.get(column))
            after = _num(match.after.get(column))
            delta = (after - before) if (match.ok and before is not None and after is not None) else None
            if prefix in _INT_PREFIXES:
                record[f"{prefix}_before"] = None if before is None else int(round(before))
                record[f"{prefix}_after"] = None if after is None else int(round(after))
                record[f"{prefix}_delta"] = None if delta is None else int(round(delta))
            else:
                record[f"{prefix}_before"] = before
                record[f"{prefix}_after"] = after
                record[f"{prefix}_delta"] = None if delta is None else round(delta, 3)
        # インパクトの定義: 変更前に接続していた端末数
        record["impact_clients"] = record["clients_before"]
        records.append(record)

    df = pd.DataFrame.from_records(records, columns=list(RESULT_COLUMNS))
    for column in ("pre_channel", "post_channel", "clients_before", "clients_after",
                   "clients_delta", "impact_clients"):
        df[column] = pd.array(df[column], dtype="Int64")
    for column in RESULT_COLUMNS:
        if column.startswith("util_"):
            df[column] = pd.to_numeric(df[column], errors="coerce")
    for column in ("channel_changed", "contaminated"):
        df[column] = df[column].astype(bool)
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 集計
# ---------------------------------------------------------------------------


def _mean(series: pd.Series) -> float | None:
    values = pd.to_numeric(series, errors="coerce").dropna()
    if values.empty:
        return None
    return round(float(values.mean()), 3)


def _impact_sum(frame: pd.DataFrame) -> int:
    return int(pd.to_numeric(frame["impact_clients"], errors="coerce").fillna(0).sum())


def _bucket_series(rows: pd.DataFrame) -> pd.Series:
    return pd.to_datetime(rows["event_timestamp"], errors="coerce").dt.floor("h")


def hourly_summary(
    rows: pd.DataFrame,
    window_start: pd.Timestamp | None,
    window_end: pd.Timestamp | None,
) -> tuple[list[dict[str, Any]], list[str]]:
    """1 時間バケットの時系列（分類別のチャネル変更回数とインパクト合計）。

    期間が複数日なら日付込みの時系列になる（**時刻 0〜23 時での集約はしない**）。
    """
    notes: list[str] = []
    if rows.empty:
        return [], notes

    buckets = _bucket_series(rows)
    changed = rows["channel_changed"].astype(bool)

    counts: dict[pd.Timestamp, dict[str, int]] = {}
    for bucket, classification, impact, is_changed in zip(
        buckets, rows["classification"],
        pd.to_numeric(rows["impact_clients"], errors="coerce").fillna(0).astype(int),
        changed,
    ):
        if pd.isna(bucket) or not is_changed:
            continue
        slot = counts.setdefault(bucket, {})
        key = str(classification)
        slot[f"changes_{key}"] = slot.get(f"changes_{key}", 0) + 1
        slot[f"impact_{key}"] = slot.get(f"impact_{key}", 0) + int(impact)

    present = sorted(b for b in buckets.dropna().unique())
    if not present:
        return [], notes

    first = pd.Timestamp(window_start).floor("h") if window_start is not None else present[0]
    if window_end is not None:
        last = (pd.Timestamp(window_end) - pd.Timedelta(seconds=1)).floor("h")
    else:
        last = present[-1]
    if last < first:
        first, last = present[0], present[-1]

    span = int((last - first).total_seconds() // BUCKET_SECONDS) + 1
    if span > MAX_BUCKETS:
        notes.append(
            f"時間帯別の集計が {span} バケットになるため、連続した時間の穴埋めをやめて"
            f"データのあるバケット {len(present)} 件だけを出しました（上限 {MAX_BUCKETS} バケット）"
        )
        timeline = list(present)
    else:
        timeline = list(pd.date_range(first, last, freq="h"))

    out: list[dict[str, Any]] = []
    for bucket in timeline:
        slot = counts.get(bucket, {})
        item: dict[str, Any] = {"bucket": pd.Timestamp(bucket).strftime("%Y-%m-%d %H:%M:%S")}
        total_changes = 0
        total_impact = 0
        for name in ev.CLASSIFICATIONS:
            changes = int(slot.get(f"changes_{name}", 0))
            impact = int(slot.get(f"impact_{name}", 0))
            item[f"changes_{name}"] = changes
            item[f"impact_{name}"] = impact
            total_changes += changes
            total_impact += impact
        item["changes_total"] = total_changes
        item["impact_total"] = total_impact
        out.append(item)
    return out, notes


def _group_stats(frame: pd.DataFrame) -> dict[str, Any]:
    """1 グループ分の統計（変更回数・no-op・インパクト・Δ の平均）。"""
    changed = frame[frame["channel_changed"].astype(bool)]
    known = frame["pre_channel"].notna() & frame["post_channel"].notna()
    noop = int((known & ~frame["channel_changed"].astype(bool)).sum())
    impact_total = _impact_sum(changed)
    stats: dict[str, Any] = {
        "events": int(len(frame)),
        "changes": int(len(changed)),
        "noop": noop,
        "unknown_channel": int((~known).sum()),
        "impact_total": impact_total,
        "impact_avg": round(impact_total / len(changed), 3) if len(changed) else None,
        "contaminated": int(changed["contaminated"].astype(bool).sum()),
        "unmatched": int((changed["match_status"] != met.MATCH_OK).sum()),
        "delta_clients_avg": _mean(changed["clients_delta"]),
    }
    for prefix, _ in VALUE_FIELDS:
        if prefix == "clients":
            continue
        stats[f"delta_{prefix}_avg"] = _mean(changed[f"{prefix}_delta"])
    return stats


def classification_summary(rows: pd.DataFrame) -> list[dict[str, Any]]:
    """分類別（**必ず 3 分類すべてを出す**。0 件でも行を落とさない）。"""
    out: list[dict[str, Any]] = []
    for name in ev.CLASSIFICATIONS:
        frame = rows[rows["classification"] == name] if not rows.empty else rows
        out.append({"classification": name, **_group_stats(frame)})
    return out


def site_summary(rows: pd.DataFrame) -> list[dict[str, Any]]:
    """サイト別（変更回数の多い順。同数はサイト名の昇順）。"""
    if rows.empty:
        return []
    out: list[dict[str, Any]] = []
    for site_name in sorted(set(rows["site_name"].astype(str))):
        frame = rows[rows["site_name"].astype(str) == site_name]
        changed = frame[frame["channel_changed"].astype(bool)]
        item: dict[str, Any] = {"site_name": site_name, **_group_stats(frame)}
        for name in ev.CLASSIFICATIONS:
            item[f"changes_{name}"] = int((changed["classification"] == name).sum())
        out.append(item)
    out.sort(key=lambda d: (-d["changes"], d["site_name"]))
    return out


def ap_summary(rows: pd.DataFrame, top_n: int = TOP_AP_COUNT) -> list[dict[str, Any]]:
    """AP 別のチャネル変更回数（多い順）。どの AP が暴れているかを見るため。"""
    if rows.empty:
        return []
    changed = rows[rows["channel_changed"].astype(bool)]
    if changed.empty:
        return []
    out: list[dict[str, Any]] = []
    grouped = changed.groupby(
        [changed["site_name"].astype(str), changed["ap_name"].astype(str), changed["ap_mac"].astype(str)],
        sort=False,
    )
    for (site_name, ap_name, ap_mac), frame in grouped:
        item = {
            "site_name": site_name,
            "ap_name": ap_name,
            "ap_mac": ap_mac,
            "changes": int(len(frame)),
            "impact_total": _impact_sum(frame),
        }
        for name in ev.CLASSIFICATIONS:
            item[f"changes_{name}"] = int((frame["classification"] == name).sum())
        out.append(item)
    out.sort(key=lambda d: (-d["changes"], -d["impact_total"], d["ap_name"]))
    return out[:top_n]


# ---------------------------------------------------------------------------
# 分析
# ---------------------------------------------------------------------------


@dataclass
class AnalysisResult:
    """:func:`run_analysis` の戻り値。"""

    params: AnalysisParams
    n_files: int
    rows: pd.DataFrame
    meta: dict[str, Any]
    warnings: list[str]


def run_analysis(
    files: Sequence[Path],
    params: AnalysisParams,
    *,
    on_phase: Callable[[str], None] | None = None,
) -> AnalysisResult:
    """ログを読み込み、RRM / RADAR のチャネル変更を分析する（CLI と API の共通経路）。

    :raises loader.LoadError: 読み込みを続行できない（種別不明 / サイト無し / 0 行）
    """
    def phase(name: str) -> None:
        if on_phase is not None:
            on_phase(name)

    phase(PHASE_LOADING)
    loaded = loader.load_logs(
        files,
        sites=params.sites,
        window_start=params.window_start,
        window_end=params.window_end,
    )
    warnings: list[str] = list(loaded.report.warnings) + list(loaded.warnings)

    phase(PHASE_EVENTS)
    actions = ev.action_frame(loaded.events)
    radar = ev.radar_summary(loaded.events, loaded.events_all)
    config_changed = ev.count_event_type(loaded.events, ev.CONFIG_RRM_EVENT_TYPE)

    phase(PHASE_METRICS)
    index = met.MetricIndex(loaded.metrics)
    change_index = met.ChangeEventIndex(loaded.events_all)
    rows = build_rows(actions, index, change_index, loaded.interval_seconds)

    phase(PHASE_AGGREGATE)
    hourly, notes = hourly_summary(rows, params.window_start, params.window_end)
    warnings.extend(notes)

    meta = build_meta(
        params=params,
        loaded=loaded,
        rows=rows,
        hourly=hourly,
        radar=radar,
        config_changed=config_changed,
        n_files=len(files),
        warnings=warnings,
    )
    return AnalysisResult(
        params=params, n_files=len(files), rows=rows, meta=meta, warnings=warnings,
    )


def build_meta(
    *,
    params: AnalysisParams,
    loaded: loader.RrmLoad,
    rows: pd.DataFrame,
    hourly: list[dict[str, Any]],
    radar: ev.RadarSummary,
    config_changed: int,
    n_files: int,
    warnings: Sequence[str],
) -> dict[str, Any]:
    """結果に添えるメタ情報（json / API / xlsx で同じものを使う）。"""
    changed = rows[rows["channel_changed"].astype(bool)] if not rows.empty else rows
    known = (
        (rows["pre_channel"].notna() & rows["post_channel"].notna())
        if not rows.empty else pd.Series(dtype=bool)
    )
    noop_mask = known & ~rows["channel_changed"].astype(bool) if not rows.empty else pd.Series(dtype=bool)

    changes_by_class = {
        name: int((changed["classification"] == name).sum()) if not changed.empty else 0
        for name in ev.CLASSIFICATIONS
    }
    noop_by_class = {
        name: int((rows.loc[noop_mask, "classification"] == name).sum()) if not rows.empty else 0
        for name in ev.CLASSIFICATIONS
    }
    match_counts = {
        status: int((rows["match_status"] == status).sum()) if not rows.empty else 0
        for status in met.MATCH_STATUSES
    }

    chart_shown = min(len(hourly), CHART_MAX_BUCKETS)
    return {
        "version": META_VERSION,
        "requested_sites": list(params.sites),
        "site_ids": list(loaded.site_ids),
        "site_names": list(loaded.site_names),
        "site_labels": list(loaded.site_labels),
        "window_start": _dt_or_none(params.window_start),
        "window_end": _dt_or_none(params.window_end),
        "bucket_seconds": BUCKET_SECONDS,
        "interval_seconds": float(loaded.interval_seconds),
        "interval_estimated": bool(loaded.interval_estimated),
        "gap_factor": float(met.MAX_GAP_FACTOR),
        "radar_match_seconds": float(ev.RADAR_MATCH_SECONDS),
        # 件数
        "event_count": int(len(rows)),
        "change_count": int(len(changed)),
        "noop_count": int(noop_mask.sum()) if not rows.empty else 0,
        "unknown_channel_count": int((~known).sum()) if not rows.empty else 0,
        "changes_by_class": changes_by_class,
        "noop_by_class": noop_by_class,
        "match_status_counts": match_counts,
        "unmatched_count": int(len(rows) - match_counts.get(met.MATCH_OK, 0)),
        "contaminated_count": int(rows["contaminated"].astype(bool).sum()) if not rows.empty else 0,
        "impact_total": _impact_sum(changed) if not changed.empty else 0,
        # レーダー（AP_RRM_ACTION とは独立に数える）
        "radar_detected": int(radar.detected),
        "radar_with_change": int(radar.with_channel_change),
        "radar_without_action": int(radar.without_action),
        # 参考（本分析では未使用）
        "config_changed_by_rrm_count": int(config_changed),
        # 集計
        "hourly": hourly,
        "by_classification": classification_summary(rows),
        "by_site": site_summary(rows),
        "by_ap": ap_summary(rows),
        "top_ap_count": TOP_AP_COUNT,
        "chart_bucket_limit": CHART_MAX_BUCKETS,
        "chart_buckets_shown": chart_shown,
        "chart_buckets_total": len(hourly),
        # 色はここが唯一の定義。フロントで色分けを定義し直さないこと
        "classifications": list(ev.CLASSIFICATIONS),
        "class_colors": dict(ev.CLASS_COLORS),
        "match_statuses": list(met.MATCH_STATUSES),
        # 入力
        "files_scanned": int(loaded.report.files_scanned),
        "metrics_files": loader.used_files(loaded.report, loader.METRICS_FILE_TYPES),
        "events_files": loader.used_files(loaded.report, [loader.EVENTS_FILE_TYPE]),
        #: 種別を判定できなかった入力（分析の出力を入力として拾っていないかの手がかり）
        "unclassified_count": int(len(loaded.report.unclassified)),
        "metrics_rows": int(loaded.report.metrics_rows),
        "events_rows_all": int(len(loaded.events_all)),
        "events_rows_in_window": int(len(loaded.events)),
        "condition_text": condition_text(params, loaded, n_files),
        "warning_count": len(warnings),
        "warnings": list(warnings),
    }


# ---------------------------------------------------------------------------
# 出力（CLI と API で同一のファイルを作るため、書き出しはここだけに置く）
# ---------------------------------------------------------------------------


def write_csv(path: Path, rows: pd.DataFrame) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def summary_text(meta: dict[str, Any]) -> str:
    """xlsx / txt の先頭に置く「前提」テキスト。表だけを見て誤読しないためのもの。"""
    lines = [meta["condition_text"]]
    changes = meta.get("changes_by_class", {})
    noop = meta.get("noop_by_class", {})
    lines.append(
        f"対象イベント（AP_RRM_ACTION）{meta.get('event_count', 0)} 件 / "
        f"チャネル変更 {meta.get('change_count', 0)} 件（"
        + " ".join(f"{name}={changes.get(name, 0)}" for name in ev.CLASSIFICATIONS)
        + f"） / 評価のみ no-op {meta.get('noop_count', 0)} 件（"
        + " ".join(f"{name}={noop.get(name, 0)}" for name in ev.CLASSIFICATIONS)
        + ")"
    )
    lines.append(
        f"照合不可 {meta.get('unmatched_count', 0)} 件（"
        + " ".join(
            f"{status}={meta.get('match_status_counts', {}).get(status, 0)}"
            for status in met.MATCH_STATUSES
        )
        + f"） / 汚染 {meta.get('contaminated_count', 0)} 件 / "
        f"インパクト合計 {meta.get('impact_total', 0)} 台"
    )
    lines.append(
        f"AP_RADAR_DETECTED: 検知 {meta.get('radar_detected', 0)} 回 / "
        f"うちチャネル変更あり {meta.get('radar_with_change', 0)} 回 / "
        f"対応する AP_RRM_ACTION が無いもの {meta.get('radar_without_action', 0)} 回"
        "（AP_RRM_ACTION だけを数えるとこの分を取りこぼす）"
    )
    lines.append(
        f"参考: AP_CONFIG_CHANGED_BY_RRM {meta.get('config_changed_by_rrm_count', 0)} 件"
        "（reason を持たないため本分析では未使用）"
    )
    if meta.get("warnings"):
        lines.append(f"警告 {len(meta['warnings'])} 件:")
        lines.extend(f"  ⚠ {w}" for w in meta["warnings"])
    else:
        lines.append("警告: なし")
    return "\n".join(lines)


def write_summary(path: Path, meta: dict[str, Any]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(f"{TITLE}\n\n" + summary_text(meta) + "\n", encoding="utf-8")
    return path


def _cell(value: object) -> object:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        try:
            return value.item()
        except (TypeError, ValueError):
            return value
    return value


#: 保存済み csv を読み戻すときの型。**書き出したときと同じ型に戻すこと**
_READ_DTYPES: dict[str, str] = {
    "event_timestamp": "string", "classification": "string", "reason": "string",
    "site_name": "string", "ap_name": "string", "ap_mac": "string", "band": "string",
    "before_timestamp": "string", "after_timestamp": "string", "match_status": "string",
}

_BOOL_TRUE: frozenset[str] = frozenset({"true", "True", "TRUE", "1"})


def read_result_csv(path: Path) -> pd.DataFrame:
    """保存済みの結果 csv を読み戻す（**再分析はしない**）。"""
    df = pd.read_csv(path, dtype=_READ_DTYPES, encoding="utf-8-sig")
    missing = [c for c in RESULT_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"結果 csv の列が足りません: {', '.join(missing)}")
    for column, dtype in _READ_DTYPES.items():
        df[column] = df[column].fillna("")
    for column in ("pre_channel", "post_channel", "clients_before", "clients_after",
                   "clients_delta", "impact_clients"):
        df[column] = pd.to_numeric(df[column], errors="coerce").astype("Int64")
    for column in RESULT_COLUMNS:
        if column.startswith("util_"):
            df[column] = pd.to_numeric(df[column], errors="coerce")
    for column in ("channel_changed", "contaminated"):
        df[column] = df[column].map(lambda v: str(v).strip() in _BOOL_TRUE).astype(bool)
    return df[list(RESULT_COLUMNS)]


# ---------------------------------------------------------------------------
# xlsx
# ---------------------------------------------------------------------------


def _style_axes(chart: BarChart) -> None:
    """27 番で確定した軸・ラベルの設定。**すべてのグラフにこれを通すこと。**"""
    # openpyxl は delete を明示しないと Excel 側で軸が描画されない
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    # 横棒（type="bar"）ではカテゴリ軸が x_axis、値軸が y_axis
    # 先頭を上に出す（既定は下から積むため）。値軸は左から右に増える向きに戻す
    chart.x_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.orientation = "minMax"
    chart.x_axis.tickLblPos = "low"
    # 軸タイトルは冗長かつ横棒では表示位置が直感と合わないため付けない
    chart.x_axis.title = None
    chart.y_axis.title = None
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showVal = True
    chart.dataLabels.showLegendKey = False
    chart.gapWidth = 40
    chart.width = 17.5


def _write_table(ws, top: int, left: int, header: Sequence[str], rows: Sequence[Sequence[Any]]) -> int:
    """表を書いて、次に使える行番号を返す。"""
    for offset, name in enumerate(header):
        ws.cell(row=top, column=left + offset, value=name).font = _HEADER_FONT
    for r, values in enumerate(rows, start=1):
        for offset, value in enumerate(values):
            ws.cell(row=top + r, column=left + offset, value=_cell(value))
    return top + len(rows) + 1


def _write_chart_sheet(ws, meta: dict[str, Any]) -> None:
    ws.title = "chart"
    ws.cell(row=1, column=1, value=TITLE).font = _TITLE_FONT
    ws.cell(row=2, column=1, value=summary_text(meta)).alignment = Alignment(
        wrap_text=True, vertical="top"
    )

    hourly: list[dict[str, Any]] = list(meta.get("hourly") or [])
    shown = hourly[-CHART_MAX_BUCKETS:] if len(hourly) > CHART_MAX_BUCKETS else hourly
    if len(hourly) > len(shown):
        note = (
            f"グラフは直近 {len(shown)} バケットのみ（全 {len(hourly)} バケット）。"
            "すべてのバケットは summary シートにあります"
        )
    else:
        note = f"時間帯別（1 時間バケット / 全 {len(hourly)} バケット）"
    ws.cell(row=4, column=1, value=note).font = _HEADER_FONT

    # 凡例（分類の色。セルの塗りは必ず 8 桁 ARGB を通す）
    ws.cell(row=5, column=1, value="凡例（分類）").font = _HEADER_FONT
    for i, name in enumerate(ev.CLASSIFICATIONS):
        color = _cell_argb(class_color(name))
        swatch = ws.cell(row=6 + i, column=1, value=None)
        swatch.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=6 + i, column=2, value=name)

    top = 6 + len(ev.CLASSIFICATIONS) + 1
    header = ["bucket", *ev.CLASSIFICATIONS]
    table = [[item["bucket"], *[item.get(f"changes_{n}", 0) for n in ev.CLASSIFICATIONS]] for item in shown]
    next_row = _write_table(ws, top, 1, header, table)

    impact_top = next_row + 1
    ws.cell(row=impact_top - 1, column=1, value="分類別のインパクト合計（変更前の接続端末数）").font = _HEADER_FONT
    impact_rows = [
        [item["classification"], int(item.get("impact_total") or 0)]
        for item in (meta.get("by_classification") or [])
    ]
    _write_table(ws, impact_top, 1, ["classification", "impact_total"], impact_rows)

    ws.column_dimensions["A"].width = 22
    for column in ("B", "C", "D"):
        ws.column_dimensions[column].width = 14

    if table:
        chart = BarChart()
        chart.type = "bar"  # 横棒（27 番で確定した描画設定をそのまま使う）
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.style = 10
        chart.title = "時間帯別のチャネル変更回数（分類別の積み上げ）"
        values = Reference(ws, min_col=2, max_col=1 + len(ev.CLASSIFICATIONS),
                           min_row=top, max_row=top + len(table))
        cats = Reference(ws, min_col=1, min_row=top + 1, max_row=top + len(table))
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        _style_axes(chart)
        for series, name in zip(chart.series, ev.CLASSIFICATIONS):
            series.graphicalProperties.solidFill = class_color(name)
            series.graphicalProperties.line.solidFill = class_color(name)
        chart.height = 4 + 0.55 * len(table)
        ws.add_chart(chart, "F2")

    if impact_rows:
        chart2 = BarChart()
        chart2.type = "bar"
        chart2.style = 10
        chart2.title = "分類別のインパクト合計"
        chart2.legend = None  # 単一系列なので Excel の凡例には意味が無い
        values = Reference(ws, min_col=2, min_row=impact_top, max_row=impact_top + len(impact_rows))
        cats = Reference(ws, min_col=1, min_row=impact_top + 1, max_row=impact_top + len(impact_rows))
        chart2.add_data(values, titles_from_data=True)
        chart2.set_categories(cats)
        _style_axes(chart2)
        chart2.height = 4 + 0.55 * len(impact_rows)
        ws.add_chart(chart2, "S2")


def _write_data_sheet(ws, rows: pd.DataFrame, meta: dict[str, Any]) -> None:
    ws.cell(row=1, column=1, value=f"{TITLE}（明細）").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=summary_text(meta)).alignment = Alignment(
        wrap_text=True, vertical="top"
    )
    header_row = 4
    for col, name in enumerate(RESULT_COLUMNS, start=1):
        ws.cell(row=header_row, column=col, value=name).font = _HEADER_FONT
    for r, row in enumerate(rows.itertuples(index=False, name=None), start=header_row + 1):
        for col, value in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=_cell(value))
    for col in range(1, len(RESULT_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


_CLASS_HEADER = (
    "classification", "events", "changes", "noop", "unknown_channel",
    "impact_total", "impact_avg", "contaminated", "unmatched",
    "delta_clients_avg", "delta_util_24_avg", "delta_util_5_avg", "delta_util_6_avg",
)
_SITE_HEADER = ("site_name", *_CLASS_HEADER[1:], *[f"changes_{n}" for n in ev.CLASSIFICATIONS])
_AP_HEADER = (
    "site_name", "ap_name", "ap_mac", "changes", "impact_total",
    *[f"changes_{n}" for n in ev.CLASSIFICATIONS],
)
_HOURLY_HEADER = (
    "bucket",
    *[f"changes_{n}" for n in ev.CLASSIFICATIONS], "changes_total",
    *[f"impact_{n}" for n in ev.CLASSIFICATIONS], "impact_total",
)


def _write_summary_sheet(ws, meta: dict[str, Any]) -> None:
    ws.cell(row=1, column=1, value=f"{TITLE}（集計）").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=summary_text(meta)).alignment = Alignment(
        wrap_text=True, vertical="top"
    )

    row = 4
    ws.cell(row=row, column=1, value="分類別").font = _HEADER_FONT
    row = _write_table(
        ws, row + 1, 1, _CLASS_HEADER,
        [[item.get(k) for k in _CLASS_HEADER] for item in (meta.get("by_classification") or [])],
    )

    row += 1
    ws.cell(row=row, column=1, value="サイト別").font = _HEADER_FONT
    row = _write_table(
        ws, row + 1, 1, _SITE_HEADER,
        [[item.get(k) for k in _SITE_HEADER] for item in (meta.get("by_site") or [])],
    )

    row += 1
    ws.cell(row=row, column=1, value=f"AP 別（変更回数の多い順・上位 {meta.get('top_ap_count', TOP_AP_COUNT)}）").font = _HEADER_FONT
    row = _write_table(
        ws, row + 1, 1, _AP_HEADER,
        [[item.get(k) for k in _AP_HEADER] for item in (meta.get("by_ap") or [])],
    )

    row += 1
    ws.cell(row=row, column=1, value="時間帯別（1 時間バケット）").font = _HEADER_FONT
    _write_table(
        ws, row + 1, 1, _HOURLY_HEADER,
        [[item.get(k) for k in _HOURLY_HEADER] for item in (meta.get("hourly") or [])],
    )

    ws.column_dimensions["A"].width = 26
    for col in range(2, len(_SITE_HEADER) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 16


def build_workbook(rows: pd.DataFrame, meta: dict[str, Any]) -> Workbook:
    """3 シート（chart / data / summary）の Workbook を組み立てる。"""
    wb = Workbook()
    _write_chart_sheet(wb.active, meta)
    _write_data_sheet(wb.create_sheet("data"), rows, meta)
    _write_summary_sheet(wb.create_sheet("summary"), meta)
    return wb


def write_xlsx(path: Path, rows: pd.DataFrame, meta: dict[str, Any]) -> Path:
    wb = build_workbook(rows, meta)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path

