"""CLI と API が共用する分析パイプラインと出力の書き出し。

``floorpeak/cli.py`` と ``backend/routers/floorpeak.py`` は、どちらもこのモジュール
だけを呼ぶ。読み込み（:mod:`floorpeak.loader`）・ピーク選定（:mod:`floorpeak.peak`）・
フロア解決（:mod:`floorpeak.floors`）のロジックはここでも再実装しない。
**CLI と UI で結果が食い違わないこと** がこのモジュールの存在理由である。

出力の列名は英字にしてある。仮名化（``pseudonymizer``）へ後から乗せるとき、
日本語の列名は leak check（非 ASCII の検出）と相性が悪いため。

ネットワークアクセス・LLM 呼び出しは行わない。
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from hangap import loader as hangap_loader

from . import floors, loader, peak as peak_mod
from .floors import UNASSIGNED

#: 結果 CSV の列（**この順で出す**。CLI / API / 保存済み結果で同一）
RESULT_COLUMNS: tuple[str, ...] = (
    "ap_name", "mac", "model", "num_clients", "status",
    "map_id", "map_name", "x_m", "y_m", "rank_in_floor",
)

#: グラフに出す上位台数。フロアの AP がこれ未満ならある分だけ出す
TOP_N: int = 20

#: 受け付ける時刻表記（これで解釈できなければ pandas に委ねる）
TIME_FORMATS: tuple[str, ...] = ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S")

#: 進捗フェーズ（「まだ動いている」と伝えるためのもの。進捗率ではない）
PHASE_LOADING = "loading"
PHASE_PEAK = "peak"
PHASE_FLOORS = "floors"
PHASE_WRITING = "writing"

#: モデルごとの棒の色（RRGGBB）。**実サイトのモデル構成は変わりうる**ので、
#: ここに無いモデルは :data:`DEFAULT_MODEL_COLOR` に落とす（落ちないこと）。
MODEL_COLORS: dict[str, str] = {
    "AP12": "8C564B",
    "AP21": "6B8E23",
    "AP24": "9467BD",
    "AP32": "2CA02C",
    "AP32E": "17BECF",
    "AP33": "BCBD22",
    "AP34": "E377C2",
    "AP41": "FF7F0E",
    "AP43": "D62728",
    "AP45": "1F77B4",
    "AP47": "C2185B",
    "AP63": "0F766E",
    "AP63E": "7C3AED",
    "BT11": "A0522D",
}

#: 辞書に無いモデルの色（灰色）
DEFAULT_MODEL_COLOR: str = "9E9E9E"

#: json の書式バージョン
META_VERSION = 1

_TITLE_FONT = Font(bold=True, size=14)
_HEADER_FONT = Font(bold=True)


class AnalysisError(RuntimeError):
    """分析を続行できない状態（CLI は終了コード 1、API は 400 / failed）。"""


class ParamError(AnalysisError):
    """パラメータが不正。``field_name`` にどの項目が不正かを持つ。"""

    def __init__(self, message: str, field_name: str | None = None) -> None:
        super().__init__(message)
        self.field_name = field_name


# loader 側の失敗も呼び出し側からは同じ扱いにしたい（CLI は 1、API は failed）
LoadError = loader.LoadError
NoMetricsError = loader.NoMetricsError
SiteNotFoundError = loader.SiteNotFoundError
UnclassifiedInputError = loader.UnclassifiedInputError


# ---------------------------------------------------------------------------
# パラメータ
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class AnalysisParams:
    """分析条件。

    :param site: 対象サイト（site_id または site_name）。**単一指定が必須**。
        「サイト全体のピーク」は複数サイトでは定義できない。
    :param at: 時点を手動で指定する。指定すると ``window_start`` /
        ``window_end`` は **無視する**（指定時点の最近傍バケットを全期間から選ぶ）。
    """

    site: str
    window_start: pd.Timestamp | None = None
    window_end: pd.Timestamp | None = None
    at: pd.Timestamp | None = None


def parse_time(text: str, label: str, field_name: str | None = None) -> pd.Timestamp:
    """``YYYY-MM-DD HH:MM`` / ISO8601 を naive な Timestamp にする（TZ 付きは拒否）。"""
    text = str(text).strip()
    for fmt in TIME_FORMATS:
        try:
            return pd.Timestamp(datetime.strptime(text, fmt))
        except ValueError:
            continue
    try:
        ts = pd.Timestamp(text)
    except Exception as exc:  # 多様な例外を投げうる外部入力の境界
        raise ParamError(f"{label} を解釈できません: {text!r}", field_name) from exc
    if pd.isna(ts):
        raise ParamError(f"{label} を解釈できません: {text!r}", field_name)
    if ts.tzinfo is not None:
        raise ParamError(
            f"{label} にタイムゾーンは付けられません（ログが naive のため）: {text!r}",
            field_name,
        )
    return ts


# ---------------------------------------------------------------------------
# 整形ヘルパ
# ---------------------------------------------------------------------------


def fmt_dt(value: object) -> str:
    if value is None:
        return "-"
    try:
        if pd.isna(value):
            return "-"
    except (TypeError, ValueError):
        pass
    return pd.Timestamp(value).strftime("%Y-%m-%d %H:%M:%S")


def _dt_or_none(value: object) -> str | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return pd.Timestamp(value).strftime("%Y-%m-%d %H:%M:%S")


def floor_sort_key(map_name: str) -> tuple[int, str]:
    """フロアの並び順。``（未割当）`` は必ず末尾に置く。"""
    return (1 if map_name == UNASSIGNED else 0, map_name)


def model_color(model: object) -> str:
    """モデル名 → 棒の色（RRGGBB）。**辞書に無いモデルでも落ちないこと**（灰色に落とす）。"""
    key = ("" if model is None else str(model)).strip().upper()
    return MODEL_COLORS.get(key, DEFAULT_MODEL_COLOR)


def _cell_argb(rgb: str) -> str:
    """セルの塗りつぶし用に ARGB（不透明）へ変換する。

    ``PatternFill`` に 6 桁の RGB をそのまま渡すと openpyxl が alpha を
    ``00``（完全透明）で補って塗りが見えなくなる。セルの塗りは常にこれを通すこと。
    グラフの ``DataPoint.solidFill`` は別のカラー型（alpha 無し）なので対象外。
    """
    return f"FF{rgb}"


def used_metrics_files(report: hangap_loader.LoadReport) -> int:
    """実際に ap_metrics として読み込んで分析に使ったファイル数。

    ``report.files_scanned`` はサイト・期間で絞る前の走査対象の総数（ログ
    ディレクトリ全体）で、``ap_metrics`` はそのうちのごく一部でしかない。
    ここでは ``file_stats`` から ap_metrics / ap_metrics_v1 として識別できた
    ファイル数だけを合計する（読み手が「入力ファイル数」を見て誤解しないため）。
    """
    return sum(
        report.file_stats[t].files
        for t in loader.METRICS_FILE_TYPES
        if t in report.file_stats
    )


def condition_text(
    params: AnalysisParams,
    n_files: int,
    site_label: str,
    peak: peak_mod.PeakResult,
    metrics_file_count: int,
) -> str:
    """分析条件のテキスト。

    **保存済み結果を後から見たときに「いつの・どこの・何の値か」が分かること。**
    グラフ単体で読み違えないよう、選定根拠（auto / manual）まで必ず残す。
    """
    if params.at is not None:
        window = f"時点指定 {fmt_dt(params.at)}（期間指定は無視）"
    else:
        left = fmt_dt(params.window_start) if params.window_start is not None else "(指定なし)"
        right = fmt_dt(params.window_end) if params.window_end is not None else "(指定なし)"
        window = f"期間 {left} 〜 {right}（半開区間 [start, end)）"
    how = "期間内で合計端末数が最大の時点" if peak.selected_by == peak_mod.SELECTED_AUTO else "指定時点に最も近い時点"
    return (
        f"分析条件: 対象サイト={site_label} / {window} / "
        f"ピーク時刻={fmt_dt(peak.peak_bucket)}（{how}・selected_by={peak.selected_by}） / "
        f"サイト合計端末数={peak.peak_total_clients} / "
        f"bucket_seconds={peak.bucket_seconds:g} / "
        f"走査ファイル数={n_files} / 使用ap_metricsファイル数={metrics_file_count}"
    )


# ---------------------------------------------------------------------------
# 分析
# ---------------------------------------------------------------------------


@dataclass
class AnalysisResult:
    """:func:`run_analysis` の戻り値。"""

    params: AnalysisParams
    n_files: int
    #: サイトの **全 AP** の行（トップ 20 の切り出しは表示側の責務）
    rows: pd.DataFrame
    meta: dict[str, Any]
    warnings: list[str]
    peak: peak_mod.PeakResult
    floors: floors.FloorResolution


def run_analysis(
    files: Sequence[Path],
    params: AnalysisParams,
    *,
    on_phase: Callable[[str], None] | None = None,
) -> AnalysisResult:
    """ログを読み込み、ピーク時点のフロア別 AP 端末数を出す（CLI と API の共通経路）。

    :raises loader.LoadError: 読み込みを続行できない（種別不明 / サイト無し / 0 行）
    """
    def phase(name: str) -> None:
        if on_phase is not None:
            on_phase(name)

    warnings: list[str] = []
    if params.at is not None and (params.window_start is not None or params.window_end is not None):
        warnings.append(
            "時点（at）を指定したため、期間の指定は無視しました"
            "（指定時点に最も近いバケットを全期間から選びます）"
        )

    phase(PHASE_LOADING)
    # at を指定したときは期間で絞らない。絞ると指定時点がその外にあったときに
    # 「最も近いバケット」が窓の端に張り付き、選定根拠を説明できなくなる。
    loaded = loader.load_metrics(
        files,
        site=params.site,
        window_start=None if params.at is not None else params.window_start,
        window_end=None if params.at is not None else params.window_end,
    )
    warnings.extend(loader.relevant_warnings(loaded.report.warnings))
    warnings.extend(loaded.warnings)

    phase(PHASE_PEAK)
    peak = peak_mod.find_peak(loaded.metrics, loaded.bucket_seconds, at=params.at)
    warnings.extend(peak.warnings)

    phase(PHASE_FLOORS)
    resolution = floors.resolve_floors(
        files, loaded.site_name, peak.peak_bucket, peak.ap_rows,
    )
    warnings.extend(resolution.warnings)

    rows = build_rows(peak.ap_rows, resolution)
    site_label = f"{loaded.site_name} [{loaded.site_id}]" if loaded.site_name else f"[{loaded.site_id}]"
    meta = build_meta(
        params=params,
        n_files=len(files),
        loaded=loaded,
        peak=peak,
        resolution=resolution,
        rows=rows,
        site_label=site_label,
        warnings=warnings,
    )
    return AnalysisResult(
        params=params,
        n_files=len(files),
        rows=rows,
        meta=meta,
        warnings=warnings,
        peak=peak,
        floors=resolution,
    )


def build_rows(ap_rows: pd.DataFrame, resolution: floors.FloorResolution) -> pd.DataFrame:
    """ピーク時点の AP 行を結果の形にする。

    - ``map_name`` は :meth:`floors.FloorResolution.floor_of`（= map_id 経由）で決める。
    - ``rank_in_floor`` はフロアごとの ``num_clients`` 降順順位（1 始まり）。
      同数は ``ap_name`` の昇順で決定論的に決める。
    - **トップ 20 への切り出しはここでは行わない**（サイトの全 AP を保存する）。
    """
    if ap_rows.empty:
        return pd.DataFrame(columns=list(RESULT_COLUMNS))

    df = pd.DataFrame({
        "ap_name": ap_rows.get("ap_name", "").astype("string").fillna(""),
        "mac": ap_rows.get("mac", "").astype("string").fillna(""),
        "model": ap_rows.get("model", "").astype("string").fillna(""),
        "num_clients": pd.to_numeric(ap_rows.get("num_clients"), errors="coerce").fillna(0).astype(int),
        "status": ap_rows.get("status", "").astype("string").fillna(""),
        "map_id": ap_rows.get("map_id", "").astype("string").fillna(""),
        "x_m": pd.to_numeric(ap_rows.get("x_m"), errors="coerce"),
        "y_m": pd.to_numeric(ap_rows.get("y_m"), errors="coerce"),
    })
    df["map_name"] = [
        resolution.floor_of(map_id, ap_name)
        for map_id, ap_name in zip(
            ap_rows.get("map_id", pd.Series(dtype=str)), ap_rows.get("ap_name", pd.Series(dtype=str))
        )
    ]

    df = df.sort_values(["map_name", "num_clients", "ap_name"], ascending=[True, False, True], kind="stable")
    df["rank_in_floor"] = df.groupby("map_name", sort=False).cumcount() + 1

    df["_floor_order"] = [floor_sort_key(v)[0] for v in df["map_name"]]
    df = df.sort_values(["_floor_order", "map_name", "rank_in_floor"], kind="stable")
    return df[list(RESULT_COLUMNS)].reset_index(drop=True)


def floor_summary(rows: pd.DataFrame) -> list[dict[str, Any]]:
    """フロアごとの AP 数・端末数（フロア選択の材料。``（未割当）`` は末尾）。"""
    if rows.empty:
        return []
    grouped = rows.groupby("map_name", sort=False).agg(
        ap_count=("ap_name", "size"), num_clients=("num_clients", "sum"),
    )
    out = [
        {"map_name": str(name), "ap_count": int(r.ap_count), "num_clients": int(r.num_clients)}
        for name, r in grouped.iterrows()
    ]
    out.sort(key=lambda d: floor_sort_key(d["map_name"]))
    return out


def default_floor(rows: pd.DataFrame) -> str | None:
    """既定で表示するフロア（端末数が最も多いフロア。同数はフロア名の昇順）。"""
    summary = floor_summary(rows)
    if not summary:
        return None
    return sorted(summary, key=lambda d: (-d["num_clients"], floor_sort_key(d["map_name"])))[0]["map_name"]


def top_rows(rows: pd.DataFrame, map_name: str, top_n: int = TOP_N) -> pd.DataFrame:
    """指定フロアの上位 ``top_n`` 行（``rank_in_floor`` 昇順）。"""
    if rows.empty:
        return rows
    hit = rows[rows["map_name"] == map_name]
    return hit.sort_values("rank_in_floor", kind="stable").head(top_n).reset_index(drop=True)


def build_meta(
    *,
    params: AnalysisParams,
    n_files: int,
    loaded: loader.MetricsLoad,
    peak: peak_mod.PeakResult,
    resolution: floors.FloorResolution,
    rows: pd.DataFrame,
    site_label: str,
    warnings: Sequence[str],
) -> dict[str, Any]:
    """結果に添えるメタ情報（json / API / xlsx で同じものを使う）。

    グラフ単体で「いつの・どこの・何の値か」が分かる必要があるので、
    **選定根拠と floormap のずれは必ず含める**。
    """
    return {
        "version": META_VERSION,
        "site_id": loaded.site_id,
        "site_name": loaded.site_name,
        "site_label": site_label,
        "requested_site": params.site,
        "window_start": _dt_or_none(params.window_start),
        "window_end": _dt_or_none(params.window_end),
        "requested_at": _dt_or_none(params.at),
        "selected_by": peak.selected_by,
        "peak_time": _dt_or_none(peak.peak_bucket),
        "peak_sample_first": _dt_or_none(peak.sample_timestamp_min),
        "peak_sample_last": _dt_or_none(peak.sample_timestamp_max),
        "peak_total_clients": int(peak.peak_total_clients),
        "manual_offset_seconds": peak.manual_offset_seconds,
        "bucket_seconds": float(peak.bucket_seconds),
        "bucket_seconds_estimated": bool(loaded.bucket_seconds_estimated),
        "bucket_count": int(peak.bucket_count),
        "floormap_file": resolution.source_file,
        "floormap_timestamp": _dt_or_none(resolution.source_timestamp),
        "floormap_offset_seconds": resolution.offset_seconds,
        #: 異常ではなく正常動作の補足（ap_name 経由のフォールバックで解決できた等）
        "floor_resolution_notes": list(resolution.notes),
        "ap_count": int(len(rows)),
        "floor_count": int(rows["map_name"].nunique()) if not rows.empty else 0,
        "floors": floor_summary(rows),
        "default_floor": default_floor(rows),
        "top_n": TOP_N,
        # 色はここが唯一の定義。フロントで色分けを定義し直さないこと
        "model_colors": dict(MODEL_COLORS),
        "default_model_color": DEFAULT_MODEL_COLOR,
        "files_scanned": int(loaded.report.files_scanned),
        "metrics_rows": int(loaded.report.metrics_rows),
        "rows_in_window": int(len(loaded.metrics)),
        "condition_text": condition_text(
            params, n_files, site_label, peak, used_metrics_files(loaded.report)
        ),
        "warning_count": len(warnings),
        "warnings": list(warnings),
    }


# ---------------------------------------------------------------------------
# 出力（CLI と API で同一のファイルを作るため、書き出しはここだけに置く）
# ---------------------------------------------------------------------------


def write_csv(path: Path, rows: pd.DataFrame) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def summary_text(meta: dict[str, Any]) -> str:
    """xlsx / txt の先頭に置く「前提」テキスト。グラフだけを見て誤読しないためのもの。"""
    lines = [meta["condition_text"]]
    lines.append(
        f"ピーク時点: {meta['peak_time']}"
        f"（実サンプル {meta['peak_sample_first']} 〜 {meta['peak_sample_last']}）"
        f" / 対象AP数 {meta['ap_count']} / フロア数 {meta['floor_count']}"
    )
    if meta.get("floormap_file"):
        offset = meta.get("floormap_offset_seconds")
        drift = "-" if offset is None else f"{offset / 60:.0f} 分"
        lines.append(
            f"フロア名の出典: {meta['floormap_file']}"
            f"（{meta['floormap_timestamp']} / ピーク時点とのずれ {drift}）"
        )
    else:
        lines.append(f"フロア名の出典: なし（すべての AP を「{UNASSIGNED}」として扱いました）")
    if meta.get("floor_resolution_notes"):
        lines.extend(f"  ℹ {n}" for n in meta["floor_resolution_notes"])
    if meta.get("warnings"):
        lines.append(f"警告 {len(meta['warnings'])} 件:")
        lines.extend(f"  ⚠ {w}" for w in meta["warnings"])
    else:
        lines.append("警告: なし")
    return "\n".join(lines)


def write_summary(path: Path, meta: dict[str, Any]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("フロア別ピーク時点分析\n\n" + summary_text(meta) + "\n", encoding="utf-8")
    return path


def _cell(value: object) -> object:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        try:
            return value.item()
        except (TypeError, ValueError):
            return value
    return value


def build_workbook(rows: pd.DataFrame, meta: dict[str, Any], floor: str | None = None) -> Workbook:
    """2 シート（chart / data）の Workbook を組み立てる。

    - ``chart``: 指定フロアの上位 :data:`TOP_N` を横棒グラフにする。棒はモデルで
      色分けする（単一系列 + DataPoint）。**Excel 標準の凡例は単一系列だと
      モデル名を出せない** ので、凡例はセルに塗りつぶし矩形とモデル名を並べて自作する。
    - ``data``: 全フロアの全行 + メタ情報。

    フロアごとにファイルを分けない。どのフロアのグラフかはシート内に明記する。
    """
    target = floor or meta.get("default_floor")
    wb = Workbook()
    _write_chart_sheet(wb.active, rows, meta, target)
    _write_data_sheet(wb.create_sheet("data"), rows, meta)
    return wb


def write_xlsx(
    path: Path,
    rows: pd.DataFrame,
    meta: dict[str, Any],
    floor: str | None = None,
) -> Path:
    """:func:`build_workbook` の結果をファイルへ書く（CLI / ジョブの出力）。"""
    wb = build_workbook(rows, meta, floor)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def xlsx_bytes(rows: pd.DataFrame, meta: dict[str, Any], floor: str | None = None) -> bytes:
    """指定フロアの xlsx をその場で組み立てて返す（一時ファイルを作らない）。"""
    buffer = io.BytesIO()
    build_workbook(rows, meta, floor).save(buffer)
    return buffer.getvalue()


#: 保存済み csv を読み戻すときの型。**書き出したときと同じ型に戻すこと**
#: （数値として書いた列を文字列で読み戻すと、順位や色分けが変わる）
_READ_DTYPES: dict[str, str] = {
    "ap_name": "string", "mac": "string", "model": "string",
    "status": "string", "map_id": "string", "map_name": "string",
}


def read_result_csv(path: Path) -> pd.DataFrame:
    """保存済みの結果 csv を読み戻す（**再分析はしない**）。"""
    df = pd.read_csv(path, dtype=_READ_DTYPES, encoding="utf-8-sig")
    missing = [c for c in RESULT_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"結果 csv の列が足りません: {', '.join(missing)}")
    for column in ("ap_name", "mac", "model", "status", "map_id", "map_name"):
        df[column] = df[column].fillna("")
    for column in ("num_clients", "rank_in_floor"):
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0).astype(int)
    for column in ("x_m", "y_m"):
        df[column] = pd.to_numeric(df[column], errors="coerce")
    return df[list(RESULT_COLUMNS)]


def _write_chart_sheet(ws, rows: pd.DataFrame, meta: dict[str, Any], target: str | None) -> None:
    ws.title = "chart"
    ws.cell(row=1, column=1, value="フロア別ピーク時点分析").font = _TITLE_FONT
    info = ws.cell(row=2, column=1, value=summary_text(meta))
    info.alignment = Alignment(wrap_text=True, vertical="top")

    top = top_rows(rows, target) if target else rows.head(0)
    label = target or UNASSIGNED
    ws.cell(
        row=4, column=1,
        value=f"グラフ対象フロア: {label}（接続端末数トップ {TOP_N} / このフロアの AP {int((rows['map_name'] == label).sum()) if not rows.empty else 0} 台）",
    ).font = _HEADER_FONT

    # 凡例（このグラフに出ているモデルだけ）
    models = list(dict.fromkeys(str(m) for m in top["model"])) if not top.empty else []
    ws.cell(row=6, column=1, value="凡例（モデル）").font = _HEADER_FONT
    for i, model in enumerate(models):
        color = _cell_argb(model_color(model))
        swatch = ws.cell(row=7 + i, column=1, value=None)
        swatch.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.cell(row=7 + i, column=2, value=model or "(不明)")

    data_start = 7 + len(models) + 1
    ws.cell(row=data_start, column=1, value="ap_name").font = _HEADER_FONT
    ws.cell(row=data_start, column=2, value="num_clients").font = _HEADER_FONT
    for i, row in enumerate(top.itertuples(index=False), start=1):
        ws.cell(row=data_start + i, column=1, value=str(row.ap_name))
        ws.cell(row=data_start + i, column=2, value=int(row.num_clients))

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14

    if top.empty:
        ws.cell(row=data_start + 1, column=1, value="（このフロアに AP がありません）")
        return

    chart = BarChart()
    chart.type = "bar"  # 横棒
    chart.style = 10
    chart.title = f"{label} / ピーク時点 {meta.get('peak_time')} の接続端末数トップ {TOP_N}"
    # 単一系列なので Excel 標準の凡例にはモデル名が出ない。凡例はセル側で自作する
    chart.legend = None
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showSerName = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showVal = True
    chart.dataLabels.showLegendKey = False

    values = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + len(top))
    cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(top))
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)

    # openpyxl は delete を明示しないと Excel 側で軸が描画されない
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    # 横棒（type="bar"）ではカテゴリ軸が x_axis、値軸が y_axis
    # 1 位を上に出す（既定は下から積むため）。値軸は左から右に増える向きに戻す
    chart.x_axis.scaling.orientation = "maxMin"
    chart.y_axis.scaling.orientation = "minMax"
    chart.x_axis.tickLblPos = "low"
    # 軸タイトルは冗長かつ横棒では表示位置が直感と合わないため付けない
    chart.x_axis.title = None
    chart.y_axis.title = None

    series = chart.series[0]
    for i, model in enumerate(top["model"]):
        point = DataPoint(idx=i)
        point.graphicalProperties.solidFill = model_color(model)
        point.graphicalProperties.line.solidFill = model_color(model)
        series.data_points.append(point)

    chart.gapWidth = 40
    chart.height = 4 + 0.55 * len(top)
    chart.width = 17.5
    ws.add_chart(chart, "D2")


def _write_data_sheet(ws, rows: pd.DataFrame, meta: dict[str, Any]) -> None:
    """全フロアの全行 + メタ情報を出す。

    分析条件は :func:`summary_text` の完全版（condition_text を含む）だけを
    1 セルに置く。以前は A2 に condition_text 単体、A3 に summary_text（同じ
    condition_text を先頭に含む完全版）と重複して入っていた
    """
    ws.cell(row=1, column=1, value="フロア別ピーク時点分析（全フロア）").font = _TITLE_FONT
    c2 = ws.cell(row=2, column=1, value=summary_text(meta))
    c2.alignment = Alignment(wrap_text=True, vertical="top")

    header_row = 4
    for col, name in enumerate(RESULT_COLUMNS, start=1):
        ws.cell(row=header_row, column=col, value=name).font = _HEADER_FONT
    for r, row in enumerate(rows.itertuples(index=False, name=None), start=header_row + 1):
        for col, value in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=_cell(value))
    for col in range(1, len(RESULT_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
