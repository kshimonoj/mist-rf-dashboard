"""hangap CLI — ローダ・検出エンジン・トポロジ診断を呼び出すだけの薄い配線。

``hangap.loader.load()`` / ``hangap.detector.detect()`` / ``hangap.topology.analyze()`` /
``hangap.neighbors.build_context()`` のロジックはここでは再実装しない。
ネットワークアクセス・LLM 呼び出しは行わない。
"""
from __future__ import annotations

import argparse
import glob as globlib
import sys
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import detector, loader, neighbors, topology

EXIT_OK = 0
EXIT_INPUT_ERROR = 1
EXIT_OUTPUT_ERROR = 2

_DATA_SUFFIXES = loader.CSV_SUFFIXES | loader.EXCEL_SUFFIXES

_TIME_FORMATS: tuple[str, ...] = ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S")

_RECOVERED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

_STATUS_ORDER: tuple[str, ...] = (
    detector.STATUS_RECOVERED,
    detector.STATUS_ONGOING,
    detector.STATUS_CUT_GAP,
    detector.STATUS_CUT_AP_DOWN,
)


class CliError(RuntimeError):
    """入力エラー（終了コード 1）。"""


class OutputError(RuntimeError):
    """出力エラー（終了コード 2）。"""


class _ArgumentParser(argparse.ArgumentParser):
    """argparse の既定終了コード(2)を使わせないための薄いラッパ。"""

    def error(self, message: str) -> None:  # noqa: D102 - argparse のオーバーライド
        self.print_usage(sys.stderr)
        raise CliError(f"{self.prog}: {message}")


# ---------------------------------------------------------------------------
# 引数パース
# ---------------------------------------------------------------------------


def build_parser() -> _ArgumentParser:
    parser = _ArgumentParser(
        prog="hangap",
        description="ハングAP候補（ゼロクライアント区間）を Mist ログから検出する",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("analyze", help="ログを読み込み、ゼロクライアント区間を検出する")
    p.add_argument("inputs", nargs="*", metavar="INPUT",
                    help="ファイル・ディレクトリ・glob パターン（複数可）")
    p.add_argument("--metrics", nargs="+", metavar="PATH", default=[],
                    help="メトリクスのみを明示指定（任意）")
    p.add_argument("--events", nargs="+", metavar="PATH", default=[],
                    help="イベントのみを明示指定（任意）")
    p.add_argument("--from", dest="window_from", metavar="TIME", default=None,
                    help="窓の開始（YYYY-MM-DD HH:MM または ISO8601、TZ 無し）")
    p.add_argument("--to", dest="window_to", metavar="TIME", default=None,
                    help="窓の終了（YYYY-MM-DD HH:MM または ISO8601、TZ 無し）")
    p.add_argument("--min-zero-samples", type=int,
                    default=detector.DEFAULT_MIN_ZERO_SAMPLES)
    p.add_argument("--min-zero-duration", metavar="DURATION", default=None,
                    help="例: 30m / 25min / 1h。指定時は --min-zero-samples より優先")
    p.add_argument("--event-window", metavar="DURATION", default="30m")
    p.add_argument("--exodus-threshold", type=float,
                    default=detector.DEFAULT_EXODUS_THRESHOLD)
    p.add_argument("--gap-factor", type=float, default=loader.DEFAULT_GAP_FACTOR)
    # 周辺AP判定（距離ベース）。既定値はいずれも暫定であり、実データを見ながら調整する前提。
    p.add_argument("--neighbor-count", type=int, default=neighbors.DEFAULT_NEIGHBOR_COUNT,
                    help=f"近傍として採用する最大台数（既定 {neighbors.DEFAULT_NEIGHBOR_COUNT}・暫定値）")
    p.add_argument("--max-distance-m", type=float, default=neighbors.DEFAULT_MAX_DISTANCE_M,
                    help=f"近傍として認める最大距離 m（既定 {neighbors.DEFAULT_MAX_DISTANCE_M:g}・暫定値）")
    p.add_argument("--neighbor-client-threshold", type=float,
                    default=neighbors.DEFAULT_NEIGHBOR_CLIENT_THRESHOLD,
                    help="周辺AP端末数合計がこれ以上なら「周辺に端末あり」"
                         f"（既定 {neighbors.DEFAULT_NEIGHBOR_CLIENT_THRESHOLD:g}・暫定値）")
    p.add_argument("--truncated-warn-ratio", type=float,
                    default=detector.DEFAULT_TRUNCATED_WARN_RATIO,
                    help="打ち切り(欠測)の比率がこれを超えたら警告する"
                         f"（既定 {detector.DEFAULT_TRUNCATED_WARN_RATIO:g}）")
    p.add_argument("--explain", metavar="AP_NAME", action="append", default=[],
                    help="指定した AP の各区間について判定根拠を表示する（複数指定可）")
    p.add_argument("--out", required=True, metavar="DIR", help="出力先ディレクトリ（必須）")
    p.add_argument("--format", choices=("xlsx", "csv", "both"), default="both")

    t = sub.add_parser(
        "topology-report",
        help="RF 隣接（rf_neighbors）と地図上の距離隣接のズレを実測する",
    )
    t.add_argument("inputs", nargs="*", metavar="INPUT",
                   help="ファイル・ディレクトリ・glob パターン（複数可）")
    t.add_argument("--out", required=True, metavar="DIR", help="出力先ディレクトリ（必須）")
    t.add_argument("--band", default=topology.DEFAULT_BAND,
                   help=f"対象バンド（24 / 5 / 6。既定 {topology.DEFAULT_BAND}）")
    t.add_argument("--top-n", dest="top_n", metavar="N[,N...]",
                   default=",".join(str(n) for n in topology.DEFAULT_TOP_N),
                   help="評価する上位N台（カンマ区切り。既定 "
                        f"{','.join(str(n) for n in topology.DEFAULT_TOP_N)}）")
    return parser


# ---------------------------------------------------------------------------
# 入力の解決
# ---------------------------------------------------------------------------


def _resolve_one(raw: str) -> list[Path]:
    p = Path(raw)
    if p.is_dir():
        found = [f for f in sorted(p.rglob("*")) if f.is_file() and f.suffix.lower() in _DATA_SUFFIXES]
        if not found:
            raise CliError(f"ディレクトリに CSV/XLSX が見つかりません: {raw}")
        return found
    if p.is_file():
        return [p]

    matches = sorted(globlib.glob(raw, recursive=True))
    if not matches:
        raise CliError(f"入力パスが見つかりません: {raw}")
    out: list[Path] = []
    for hit in matches:
        hp = Path(hit)
        if hp.is_dir():
            out.extend(f for f in sorted(hp.rglob("*")) if f.is_file() and f.suffix.lower() in _DATA_SUFFIXES)
        elif hp.is_file():
            out.append(hp)
    if not out:
        raise CliError(f"入力パスが見つかりません: {raw}")
    return out


def resolve_inputs(raw_paths: Sequence[str]) -> list[Path]:
    """ファイル・ディレクトリ・glob パターンを、重複を除いた具体的なファイル一覧へ展開する。

    ``loader.load()`` 自身も同じ探索を行うが、ここでは「見つからない入力を明示的に
    エラーにする」ための検証を目的とする（loader 側は判定できないだけなら黙って続行する）。
    """
    seen: set[Path] = set()
    files: list[Path] = []
    for raw in raw_paths:
        for f in _resolve_one(raw):
            rp = f.resolve()
            if rp not in seen:
                seen.add(rp)
                files.append(f)
    return files


def _check_output_not_input(out_dir: str, raw_paths: Sequence[str], files: Sequence[Path]) -> None:
    out_real = Path(out_dir).resolve()
    for raw in raw_paths:
        p = Path(raw)
        if p.is_dir() and p.resolve() == out_real:
            raise OutputError(f"--out が入力ディレクトリと同一です: {out_dir}")
    for f in files:
        if f.resolve().parent == out_real:
            raise OutputError(f"--out が入力ファイルの所在ディレクトリと同一です: {out_dir}")


# ---------------------------------------------------------------------------
# 時刻・時間指定のパース
# ---------------------------------------------------------------------------


def _parse_time(text: str, label: str) -> pd.Timestamp:
    text = text.strip()
    for fmt in _TIME_FORMATS:
        try:
            return pd.Timestamp(datetime.strptime(text, fmt))
        except ValueError:
            continue
    try:
        ts = pd.Timestamp(text)
    except Exception as exc:  # 多様な例外を投げうる外部入力の境界
        raise CliError(f"{label} を解釈できません: {text!r}") from exc
    if ts.tzinfo is not None:
        raise CliError(f"{label} にタイムゾーンは付けられません（ログが naive のため）: {text!r}")
    return ts


def _parse_duration(text: str, label: str) -> pd.Timedelta:
    try:
        td = pd.Timedelta(text)
    except Exception as exc:
        raise CliError(f"{label} を解釈できません: {text!r}") from exc
    if pd.isna(td):
        raise CliError(f"{label} を解釈できません: {text!r}")
    return td


# ---------------------------------------------------------------------------
# 整形ヘルパ
# ---------------------------------------------------------------------------


def _fmt_dt(dt: object) -> str:
    if dt is None or pd.isna(dt):
        return "-"
    return pd.Timestamp(dt).strftime("%Y-%m-%d %H:%M:%S")


def _fmt_period(period: tuple | None) -> str:
    if period is None:
        return "（なし）"
    return f"{_fmt_dt(period[0])} 〜 {_fmt_dt(period[1])}"


def _fmt_window(ws: pd.Timestamp | None, we: pd.Timestamp | None) -> str:
    left = _fmt_dt(ws) if ws is not None else "(指定なし)"
    right = _fmt_dt(we) if we is not None else "(指定なし)"
    return f"{left} 〜 {right}"


def _fmt_td(td: pd.Timedelta) -> str:
    total = td.total_seconds()
    if total % 3600 == 0 and total >= 3600:
        return f"{total / 3600:g}h"
    if total % 60 == 0:
        return f"{total / 60:g}m"
    return f"{total:g}s"


def _condition_text(
    args: argparse.Namespace,
    ws: pd.Timestamp | None,
    we: pd.Timestamp | None,
    min_duration: pd.Timedelta | None,
    event_window: pd.Timedelta,
    n_files: int,
) -> str:
    zero_desc = (
        f"min_zero_duration={_fmt_td(min_duration)}"
        if min_duration is not None
        else f"min_zero_samples={args.min_zero_samples}"
    )
    return (
        f"分析条件: 窓 {_fmt_window(ws, we)} / {zero_desc} / "
        f"event_window={_fmt_td(event_window)} / exodus_threshold={args.exodus_threshold} / "
        f"gap_factor={args.gap_factor} / 入力ファイル数={n_files} / "
        f"neighbor_count={args.neighbor_count} / max_distance_m={args.max_distance_m:g} / "
        f"neighbor_client_threshold={args.neighbor_client_threshold:g}（周辺AP判定の既定値は暫定）"
    )


def _coverage_and_warnings_text(
    report: loader.LoadReport,
    detector_warnings: list[str],
    quality_warnings: Sequence[str] = (),
) -> str:
    lines = [
        f"データ範囲: metrics {_fmt_period(report.metrics_period)} / "
        f"events {_fmt_period(report.events_period)}"
    ]
    all_warnings = list(report.warnings) + list(detector_warnings) + list(quality_warnings)
    if all_warnings:
        lines.append(f"警告 {len(all_warnings)} 件:")
        lines.extend(f"  ⚠ {w}" for w in all_warnings)
    else:
        lines.append("警告: なし")
    return "\n".join(lines)


def _format_result_summary(df: pd.DataFrame) -> str:
    total = len(df)
    lines = [f"検出区間数: {total}"]
    if total:
        counts = df["回復状況"].value_counts()
        for status in _STATUS_ORDER:
            lines.append(f"  {status}: {int(counts.get(status, 0))}")
        lines.append(f"退場疑い: {int(df['退場疑い'].sum())} 件")
        lines.append(f"イベントが該当した区間数: {int((df['AP Event（±30分）'] == 'あり').sum())} 件")
        # 周辺AP判定は判断材料であって絞り込み条件ではない。内訳を出すだけで行は落とさない。
        verdicts = df["周辺AP判定"].value_counts()
        lines.append("周辺AP判定:")
        for verdict in (neighbors.VERDICT_PRESENT, neighbors.VERDICT_ABSENT,
                        neighbors.VERDICT_UNKNOWN):
            lines.append(f"  {verdict}: {int(verdicts.get(verdict, 0))}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 出力
# ---------------------------------------------------------------------------


@dataclass
class _Meta:
    title: str
    condition_text: str
    coverage_and_warnings_text: str
    result_summary_text: str


def _cell_value(value: object) -> object:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        try:
            return value.item()
        except (TypeError, ValueError):
            return value
    return value


def _write_xlsx(path: Path, df: pd.DataFrame, meta: _Meta) -> None:
    columns = detector.RESULT_COLUMNS
    ncols = len(columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "ハングAP分析結果"

    ws.cell(row=1, column=1, value=meta.title).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=meta.condition_text)
    c3 = ws.cell(row=3, column=1, value=meta.coverage_and_warnings_text)
    c3.alignment = Alignment(wrap_text=True, vertical="top")

    header_row = 5
    for col, name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = Font(bold=True)

    status_col = columns.index("回復状況") + 1
    for r, row in enumerate(df.itertuples(index=False, name=None), start=header_row + 1):
        for col, value in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=_cell_value(value))
        if row[status_col - 1] == detector.STATUS_RECOVERED:
            for col in range(1, ncols + 1):
                ws.cell(row=r, column=col).fill = _RECOVERED_FILL

    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_summary(path: Path, meta: _Meta) -> None:
    text = "\n\n".join([
        meta.title,
        meta.condition_text,
        meta.coverage_and_warnings_text,
        meta.result_summary_text,
    ])
    path.write_text(text + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# 本体
# ---------------------------------------------------------------------------


def run(args: argparse.Namespace) -> int:
    """サブコマンドごとの処理へ振り分ける。"""
    if args.command == "topology-report":
        return run_topology_report(args)
    return run_analyze(args)


def _parse_top_n(text: str) -> tuple[int, ...]:
    values: list[int] = []
    for part in str(text).split(","):
        part = part.strip()
        if not part:
            continue
        try:
            n = int(part)
        except ValueError as exc:
            raise CliError(f"--top-n を解釈できません: {text!r}") from exc
        if n <= 0:
            raise CliError(f"--top-n は 1 以上で指定してください: {text!r}")
        values.append(n)
    if not values:
        raise CliError(f"--top-n を解釈できません: {text!r}")
    return tuple(sorted(set(values)))


def run_topology_report(args: argparse.Namespace) -> int:
    raw_paths = list(args.inputs)
    if not raw_paths:
        raise CliError("入力パスを指定してください")

    files = resolve_inputs(raw_paths)
    _check_output_not_input(args.out, raw_paths, files)
    top_n = _parse_top_n(args.top_n)

    load_result = loader.load(files)
    report = load_result.report

    if report.metrics_rows == 0 and report.rf_neighbors_rows == 0 and report.unclassified:
        sample = ", ".join(report.unclassified[:5])
        more = " ..." if len(report.unclassified) > 5 else ""
        raise CliError(
            "入力ファイルの種別を判定できませんでした"
            f"（ap_metrics / rf_neighbors のいずれにも一致しません）: {sample}{more}"
        )

    result = topology.analyze(
        load_result.metrics,
        load_result.rf_neighbors,
        band=args.band,
        top_n=top_n,
    )

    text = result.render()
    print(report.render())
    print()
    print(text)
    print()

    out_dir = Path(args.out)
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise OutputError(f"出力先ディレクトリを作成できません: {args.out} ({exc})") from exc

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    txt_path = out_dir / f"topology_report_{stamp}.txt"
    csv_path = out_dir / f"topology_report_{stamp}.csv"
    try:
        txt_path.write_text(text + "\n", encoding="utf-8")
        result.detail.to_csv(csv_path, index=False, encoding="utf-8-sig")
    except OSError as exc:
        raise OutputError(f"出力ファイルの書き込みに失敗しました: {exc}") from exc

    print("[ 出力ファイル ]")
    for p in (txt_path, csv_path):
        print(f"  {p}")
    return EXIT_OK


def run_analyze(args: argparse.Namespace) -> int:
    raw_paths = list(args.inputs) + list(args.metrics) + list(args.events)
    if not raw_paths:
        raise CliError("入力パスを指定してください（位置引数 / --metrics / --events のいずれか）")

    files = resolve_inputs(raw_paths)
    _check_output_not_input(args.out, raw_paths, files)

    ws = _parse_time(args.window_from, "--from") if args.window_from else None
    we = _parse_time(args.window_to, "--to") if args.window_to else None
    min_duration = (
        _parse_duration(args.min_zero_duration, "--min-zero-duration")
        if args.min_zero_duration
        else None
    )
    event_window = _parse_duration(args.event_window, "--event-window")

    load_result = loader.load(files, gap_factor=args.gap_factor)
    report = load_result.report

    if report.metrics_rows == 0 and report.events_rows == 0 and report.unclassified:
        sample = ", ".join(report.unclassified[:5])
        more = " ..." if len(report.unclassified) > 5 else ""
        raise CliError(
            "入力ファイルの種別を判定できませんでした"
            f"（ap_metrics / ap_events のいずれにも一致しません）: {sample}{more}"
        )

    # 近傍AP のインデックスは検出と explain で共有する（座標は AP の最新行から 1 度だけ取る）
    neighbor_context = neighbors.build_context(
        load_result.metrics,
        load_result.rf_neighbors,
        neighbor_count=args.neighbor_count,
        max_distance_m=args.max_distance_m,
    )

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        result_df = detector.detect(
            load_result.metrics,
            load_result.events,
            load_result.gaps,
            window_start=ws,
            window_end=we,
            min_zero_samples=args.min_zero_samples,
            min_zero_duration=min_duration,
            event_window=event_window,
            exodus_threshold=args.exodus_threshold,
            neighbor_context=neighbor_context,
            neighbor_client_threshold=args.neighbor_client_threshold,
        )
    detector_warnings = [str(w.message) for w in caught if issubclass(w.category, UserWarning)]

    truncated = detector.truncated_warning(result_df, args.truncated_warn_ratio)
    quality_warnings = [truncated] if truncated else []

    condition_text = _condition_text(args, ws, we, min_duration, event_window, len(files))
    coverage_text = _coverage_and_warnings_text(report, detector_warnings, quality_warnings)
    result_summary_text = _format_result_summary(result_df)
    meta = _Meta(
        title="ハングAP分析結果",
        condition_text=condition_text,
        coverage_and_warnings_text=coverage_text,
        result_summary_text=result_summary_text,
    )

    # 1. ローダのレポート
    print(report.render())
    print()
    # 2. 警告（検出時。データ範囲不足など）
    print(f"[ 警告（検出時） ] {len(detector_warnings)} 件")
    if not detector_warnings:
        print("  （なし）")
    for w in detector_warnings:
        print(f"  ! {w}")
    print()
    # 3. 分析条件
    print(condition_text)
    print()
    # 4. 結果サマリー
    print("[ 結果サマリー ]")
    print(result_summary_text)
    print()
    # 5. データ品質の警告（件数だけを見ていると気づけないため独立した節にする）
    if quality_warnings:
        print("[ データ品質の警告 ]")
        for w in quality_warnings:
            print(f"  ⚠ 警告: {w}")
        print()
    # 6. 判定根拠（--explain）
    if args.explain:
        print(neighbors.render_explain(result_df, args.explain, neighbor_context))
        print()

    out_dir = Path(args.out)
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        raise OutputError(f"出力先ディレクトリを作成できません: {args.out} ({exc})") from exc

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    written: list[Path] = []
    try:
        if args.format in ("xlsx", "both"):
            xlsx_path = out_dir / f"hangap_result_{stamp}.xlsx"
            _write_xlsx(xlsx_path, result_df, meta)
            written.append(xlsx_path)
        if args.format in ("csv", "both"):
            csv_path = out_dir / f"hangap_result_{stamp}.csv"
            result_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
            written.append(csv_path)
            # xlsx には条件・警告を埋め込めるが、CSV 単体は表形式のみのため、
            # --format both でも CSV を受け取った人が読めるよう summary は常に添える。
            summary_path = out_dir / f"hangap_result_{stamp}_summary.txt"
            _write_summary(summary_path, meta)
            written.append(summary_path)
    except OSError as exc:
        raise OutputError(f"出力ファイルの書き込みに失敗しました: {exc}") from exc

    # 7. 出力ファイルのパス
    print("[ 出力ファイル ]")
    for p in written:
        print(f"  {p}")

    return EXIT_OK


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    try:
        args = parser.parse_args(argv)
        return run(args)
    except CliError as e:
        print(f"error: {e}", file=sys.stderr)
        return EXIT_INPUT_ERROR
    except OutputError as e:
        print(f"error: {e}", file=sys.stderr)
        return EXIT_OUTPUT_ERROR


if __name__ == "__main__":
    raise SystemExit(main())
