"""複数ファイルに分かれた Mist ログを結合し、正規化済みデータへ変換するローダ。

設計方針:
- 種別判定は **ヘッダー行の列集合** で行う（``pseudonymizer.schemas`` の仕組みを再利用）。
  ファイル名は判定に一切使わない。
- 結合時の危険（重複行・欠測をまたいだ「連続」誤認・環境混在）を、
  **データを書き換えずに** レポートとして可視化することが本モジュールの責務。
- 検出ロジック（区間検出・イベント相関・近傍判定）はここには置かない。
- ネットワークアクセス・LLM 呼び出しは行わない。入力はローカルファイルのみ。
"""
from __future__ import annotations

import csv
import glob as globlib
import os
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable, Sequence

import pandas as pd

from pseudonymizer.schemas import (
    AP_EVENTS_COLUMNS,
    AP_METRICS_COLUMNS,
    FILE_TYPES_BY_KEY,
    RF_NEIGHBORS_COLUMNS,
    detect_file_type,
)

# ---------------------------------------------------------------------------
# 既定値
# ---------------------------------------------------------------------------

#: 既定で DataFrame まで読み込む種別（他の種別は件数のみ数える）
DEFAULT_FILE_TYPES: tuple[str, ...] = (
    "ap_metrics", "ap_metrics_v1", "ap_events", "rf_neighbors",
)

#: ap_metrics として結合する種別（座標列追加前後のバージョン違い）
METRICS_FILE_TYPES: frozenset[str] = frozenset({"ap_metrics", "ap_metrics_v1"})

#: RRM 隣接の種別（rf_neighbors は日次取得のため 1 種別のみ）
RF_NEIGHBORS_FILE_TYPE: str = "rf_neighbors"

#: rf_neighbors の重複判定キー（方向を潰さないため neighbor_mac までを含める）
RF_NEIGHBORS_KEY: tuple[str, ...] = (
    "site_id", "band", "ap_mac", "neighbor_mac", "timestamp",
)

#: ギャップ判定のしきい値係数（推定間隔 × この値を超えたら欠測とみなす）
DEFAULT_GAP_FACTOR: float = 1.5

#: 「イベントが存在しない区間」として報告する最小の長さ（秒）
DEFAULT_EVENT_GAP_SECONDS: float = 3600.0

CSV_SUFFIXES: frozenset[str] = frozenset({".csv"})
EXCEL_SUFFIXES: frozenset[str] = frozenset({".xlsx", ".xlsm"})

#: 走査から必ず外すディレクトリ名。
#: ``hangap_results`` は hangap 自身の**出力**（保存済みの分析結果）を置く場所で、
#: これを入力として拾うとヘッダー判定で弾かれ、既定（``on_unclassified='error'``）
#: では分析が止まる。ディレクトリ名で無条件に除外する。
#: ``floorpeak_results`` / ``rrm_results`` も同じ理由で外す
#: （``floorpeak.archive.RESULTS_DIR_NAME`` / ``rrm.archive.RESULTS_DIR_NAME``）。
EXCLUDED_DIR_NAMES: frozenset[str] = frozenset(
    {"hangap_results", "floorpeak_results", "rrm_results"}
)

#: ファイル名に現れうるタイムゾーントークン（変換には使わない。混在の警告のみ）
TZ_TOKENS: frozenset[str] = frozenset(
    {"JST", "UTC", "GMT", "KST", "CST", "PST", "PDT", "EST", "EDT", "CET", "IST", "AEST"}
)

#: MAC として正規化する列（コロン無し小文字。プロジェクト規約に合わせる）
MAC_COLUMNS: frozenset[str] = frozenset({"mac", "ap_mac", "bssid", "neighbor_mac"})

#: ap_metrics で必ず文字列として読む列（先頭ゼロ落ち・数値化を防ぐ）
_METRICS_STR_COLUMNS: tuple[str, ...] = (
    "site_id", "site_name", "ap_id", "ap_name", "model", "mac", "status", "map_id",
)

#: rf_neighbors で必ず文字列として読む列（band の "5" が 5.0 になるのを防ぐ）
_RF_NEIGHBORS_STR_COLUMNS: tuple[str, ...] = (
    "site_id", "site_name", "band", "ap_mac", "ap_name", "neighbor_mac", "neighbor_name",
)

#: ap_events は全列一致で重複判定するため、いったん全列を文字列で読む。
#: 読み込み後に数値へ戻す列だけをここに列挙する。
_EVENTS_NUMERIC_COLUMNS: tuple[str, ...] = (
    "channel", "pre_channel", "bandwidth", "pre_bandwidth",
)

GAP_COLUMNS: tuple[str, ...] = (
    "ap_id", "ap_name", "gap_start", "gap_end", "gap_seconds", "missing_samples",
)

#: 推定間隔をグループ化するときの相対許容幅（±25%）
_INTERVAL_TOLERANCE: float = 0.25


# ---------------------------------------------------------------------------
# レポートの構成要素
# ---------------------------------------------------------------------------


@dataclass
class FileTypeStat:
    """種別ごとの読み込み統計。"""

    file_type: str
    files: int = 0
    rows: int = 0
    duplicates_removed: int = 0
    loaded: bool = False  # DataFrame まで読み込んだ種別か（False は件数のみ）


@dataclass
class ApInterval:
    """AP ごとの推定サンプリング間隔。"""

    ap_id: str
    ap_name: str
    samples: int
    interval_seconds: float | None


@dataclass
class SitePeriod:
    """site_id ごとの出現期間。"""

    site_id: str
    site_name: str
    rows: int
    ap_count: int
    first: datetime
    last: datetime


@dataclass
class SiteFilter:
    """サイト指定の解決結果（指定が無いときは ``LoadReport.site_filter`` が None）。"""

    #: 利用者が指定した文字列（site_id または site_name）
    requested: tuple[str, ...] = ()
    #: 解決できた site_id（指定順・重複なし）
    site_ids: tuple[str, ...] = ()
    #: 表示用のラベル（``名前 [site_id]``）
    labels: tuple[str, ...] = ()
    #: ログに存在しなかった指定。**空でなければ分析を続けないこと**
    missing: tuple[str, ...] = ()
    #: 絞り込む前にログへ含まれていたサイトのラベル（指定ミスを説明するために持つ）
    available: tuple[str, ...] = ()
    #: 絞り込む前の ap_metrics 行数（「ログが無い」と「サイトが無い」を区別するため）
    rows_before: int = 0


@dataclass
class GapSummary:
    """ギャップ（欠測）の集計。"""

    count: int = 0
    total_seconds: float = 0.0
    max_seconds: float = 0.0
    max_ap_name: str = ""
    max_start: datetime | None = None
    max_end: datetime | None = None
    total_missing_samples: int = 0


@dataclass
class LoadReport:
    """load() の結果を人間が検証するためのレポート。"""

    files_scanned: int = 0
    gap_factor: float = DEFAULT_GAP_FACTOR
    file_stats: dict[str, FileTypeStat] = field(default_factory=dict)
    #: 種別を判定できなかったファイル／シート（**名前のみ**。中身の値は含めない）
    unclassified: list[str] = field(default_factory=list)
    ap_intervals: list[ApInterval] = field(default_factory=list)
    overall_interval_seconds: float | None = None
    #: (代表間隔[秒], その間隔と判定された AP 数) のリスト
    interval_groups: list[tuple[float, int]] = field(default_factory=list)
    gaps: GapSummary = field(default_factory=GapSummary)
    site_periods: list[SitePeriod] = field(default_factory=list)
    #: サイト指定の解決結果（指定が無ければ None）
    site_filter: SiteFilter | None = None
    metrics_period: tuple[datetime, datetime] | None = None
    events_period: tuple[datetime, datetime] | None = None
    #: メトリクス期間のうちイベントが存在しない区間
    event_blind_spots: list[tuple[datetime, datetime]] = field(default_factory=list)
    event_gap_seconds: float = DEFAULT_EVENT_GAP_SECONDS
    ap_count: int = 0
    metrics_rows: int = 0
    events_rows: int = 0
    #: rf_neighbors の全行数（複数日分を読み込んだ場合はその合計）
    rf_neighbors_rows: int = 0
    #: (取得時刻, 行数) の一覧。日次取得なので通常は 1 日 1 件
    rf_neighbors_snapshots: list[tuple[datetime, int]] = field(default_factory=list)
    #: 分析に使う取得時刻（最新のスナップショット）
    rf_neighbors_latest: datetime | None = None
    tz_tokens: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    # -- 整形 ---------------------------------------------------------------

    def render(self) -> str:
        """人が読めるテキストへ整形する（後段の CLI がそのまま出力に使う）。"""
        lines: list[str] = []
        add = lines.append

        add("=" * 68)
        add("Mist ログ結合ローダ レポート")
        add("=" * 68)
        add(f"走査ファイル数: {self.files_scanned}")
        add(f"gap_factor: {self.gap_factor}")

        add("")
        add("[ 種別ごとのファイル数・行数・重複除去 ]")
        if not self.file_stats:
            add("  （該当なし）")
        for key in sorted(self.file_stats):
            st = self.file_stats[key]
            mark = "読込" if st.loaded else "件数のみ"
            add(
                f"  {key:<20} files={st.files:>5}  rows={st.rows:>8}  "
                f"dedupe={st.duplicates_removed:>6}  ({mark})"
            )

        add("")
        add(f"[ 判定できなかったファイル／シート ] {len(self.unclassified)} 件")
        for name in self.unclassified[:20]:
            add(f"  - {name}")
        if len(self.unclassified) > 20:
            add(f"  ... 他 {len(self.unclassified) - 20} 件")

        add("")
        add("[ サンプリング間隔の推定 ]")
        add(f"  全体の代表値: {_fmt_secs(self.overall_interval_seconds)}")
        for rep, n in self.interval_groups:
            add(f"    {_fmt_secs(rep):>12} : {n} AP")
        deviating = [
            a for a in self.ap_intervals
            if a.interval_seconds is not None
            and self.overall_interval_seconds is not None
            and not _close(a.interval_seconds, self.overall_interval_seconds)
        ]
        unknown = [a for a in self.ap_intervals if a.interval_seconds is None]
        if len(self.ap_intervals) <= 20:
            add("  AP ごとの推定値:")
            for a in self.ap_intervals:
                add(
                    f"    {a.ap_name or a.ap_id:<24} "
                    f"{_fmt_secs(a.interval_seconds):>12}  (samples={a.samples})"
                )
        else:
            add(f"  AP 数: {len(self.ap_intervals)}（詳細は report.ap_intervals）")
            if deviating:
                add(f"  代表値と異なる AP: {len(deviating)}")
                for a in deviating[:20]:
                    add(
                        f"    {a.ap_name or a.ap_id:<24} "
                        f"{_fmt_secs(a.interval_seconds):>12}  (samples={a.samples})"
                    )
                if len(deviating) > 20:
                    add(f"    ... 他 {len(deviating) - 20} AP")
            if unknown:
                add(f"  推定不能（サンプル 1 件以下）の AP: {len(unknown)}")

        add("")
        add("[ ギャップ（欠測） ]")
        g = self.gaps
        add(f"  件数: {g.count}")
        add(f"  合計: {_fmt_secs(g.total_seconds)}  欠測サンプル合計: {g.total_missing_samples}")
        if g.count:
            add(
                f"  最大: {_fmt_secs(g.max_seconds)}  "
                f"{_fmt_dt(g.max_start)} → {_fmt_dt(g.max_end)}  ap={g.max_ap_name}"
            )

        add("")
        add("[ サイト指定 ]")
        sf = self.site_filter
        if sf is None:
            add("  すべてのサイト（指定なし）")
        else:
            add(f"  指定: {', '.join(sf.requested) or '（なし）'}")
            add(f"  対象: {', '.join(sf.labels) or '（該当なし）'}")
            add(f"  絞り込み前の ap_metrics 行数: {sf.rows_before}")
            if sf.missing:
                add(f"  ログに存在しない指定: {', '.join(sf.missing)}")

        add("")
        add("[ site_id ごとの出現期間 ]")
        if not self.site_periods:
            add("  （メトリクスなし）")
        for sp in self.site_periods:
            add(
                f"  {sp.site_id}  ({sp.site_name})  rows={sp.rows}  ap={sp.ap_count}"
            )
            add(f"      {_fmt_dt(sp.first)} → {_fmt_dt(sp.last)}")

        add("")
        add("[ 期間 ]")
        add(f"  メトリクス: {_fmt_period(self.metrics_period)}  rows={self.metrics_rows}")
        add(f"  イベント  : {_fmt_period(self.events_period)}  rows={self.events_rows}")
        add(f"  期間内に登場した AP 数: {self.ap_count}")

        add("")
        add(
            f"[ イベントが存在しない区間 ] "
            f"（{_fmt_secs(self.event_gap_seconds)} 超のみ）{len(self.event_blind_spots)} 件"
        )
        if self.events_rows == 0:
            add("  この期間のイベントログはありません（正常系）")
        for start, end in self.event_blind_spots[:10]:
            add(
                f"  - {_fmt_dt(start)} → {_fmt_dt(end)} "
                f"({_fmt_secs((end - start).total_seconds())})"
            )
        if len(self.event_blind_spots) > 10:
            add(f"  ... 他 {len(self.event_blind_spots) - 10} 件")

        add("")
        add("[ RF 隣接（rf_neighbors） ]")
        if self.rf_neighbors_rows == 0:
            add("  読み込みなし（既存ログには存在しないため、これが通常状態）")
        else:
            add(f"  全行数: {self.rf_neighbors_rows}  取得時刻: {len(self.rf_neighbors_snapshots)} 時点")
            for ts, n in self.rf_neighbors_snapshots:
                mark = " ← 分析に使用" if ts == self.rf_neighbors_latest else ""
                add(f"    {_fmt_dt(ts)}  rows={n}{mark}")

        if self.tz_tokens:
            add("")
            add("[ ファイル名の TZ トークン ]（変換には使っていない）")
            for tok, n in sorted(self.tz_tokens.items()):
                add(f"  {tok}: {n} ファイル")

        add("")
        add(f"[ 警告 ] {len(self.warnings)} 件")
        if not self.warnings:
            add("  （なし）")
        for w in self.warnings:
            add(f"  ! {w}")
        add("=" * 68)
        return "\n".join(lines)


def _empty_rf_neighbors() -> pd.DataFrame:
    empty = pd.DataFrame(columns=list(RF_NEIGHBORS_COLUMNS))
    empty["timestamp"] = pd.to_datetime(empty["timestamp"])
    return empty


@dataclass
class LoadResult:
    """load() の戻り値。"""

    metrics: pd.DataFrame
    events: pd.DataFrame
    gaps: pd.DataFrame
    report: LoadReport
    #: RRM 隣接。読み込んだ**全時点**を保持する（分析側で最新時点だけを使う）
    rf_neighbors: pd.DataFrame = field(default_factory=_empty_rf_neighbors)


# ---------------------------------------------------------------------------
# 整形ヘルパ
# ---------------------------------------------------------------------------


def _fmt_secs(sec: float | None) -> str:
    if sec is None:
        return "不明"
    sec = float(sec)
    if sec < 60:
        return f"{sec:g}s"
    if sec < 3600:
        return f"{sec / 60:g}m ({sec:g}s)"
    return f"{sec / 3600:.2f}h ({sec:g}s)"


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None or pd.isna(dt):
        return "-"
    return pd.Timestamp(dt).strftime("%Y-%m-%d %H:%M:%S")


def _fmt_period(period: tuple[datetime, datetime] | None) -> str:
    if period is None:
        return "（なし）"
    return f"{_fmt_dt(period[0])} → {_fmt_dt(period[1])}"


def _close(a: float, b: float, tol: float = _INTERVAL_TOLERANCE) -> bool:
    if a <= 0 or b <= 0:
        return a == b
    return abs(a - b) <= tol * max(a, b)


# ---------------------------------------------------------------------------
# ファイル探索・種別判定
# ---------------------------------------------------------------------------


def is_data_file(p: Path) -> bool:
    """走査対象のデータファイルか。:data:`EXCLUDED_DIR_NAMES` 配下は常に外す。"""
    if p.suffix.lower() not in (CSV_SUFFIXES | EXCEL_SUFFIXES):
        return False
    return not EXCLUDED_DIR_NAMES.intersection(p.parts)


def _iter_input_files(paths: str | os.PathLike | Iterable[str | os.PathLike]) -> list[Path]:
    """ファイル・ディレクトリ・glob パターンを展開して、対象ファイルの一覧を返す。"""
    if isinstance(paths, (str, os.PathLike)):
        paths = [paths]
    found: list[Path] = []
    seen: set[Path] = set()

    def _push(p: Path) -> None:
        if EXCLUDED_DIR_NAMES.intersection(p.parts):
            return
        rp = p.resolve()
        if rp not in seen:
            seen.add(rp)
            found.append(p)

    for raw in paths:
        p = Path(raw)
        if p.is_dir():
            for sub in sorted(p.rglob("*")):
                if sub.is_file() and sub.suffix.lower() in (CSV_SUFFIXES | EXCEL_SUFFIXES):
                    _push(sub)
        elif p.is_file():
            _push(p)
        else:
            for hit in sorted(globlib.glob(str(raw), recursive=True)):
                hp = Path(hit)
                if hp.is_dir():
                    for sub in sorted(hp.rglob("*")):
                        if sub.is_file() and sub.suffix.lower() in (CSV_SUFFIXES | EXCEL_SUFFIXES):
                            _push(sub)
                elif hp.is_file():
                    _push(hp)
    return found


def _read_csv_header(path: Path) -> list[str] | None:
    """CSV の 1 行目を列名リストとして返す。空ファイルなら None。"""
    try:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            for row in csv.reader(f):
                return [c.strip() for c in row]
    except OSError:
        return None
    return None


def _count_csv_rows(path: Path) -> int:
    """ヘッダーを除いたデータ行数を数える（読み込まない種別の行数用）。"""
    try:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            n = sum(1 for _ in csv.reader(f))
    except OSError:
        return 0
    return max(0, n - 1)


def _tz_token_of(path: Path) -> str | None:
    for part in re.split(r"[._\-]", path.stem.upper()):
        if part in TZ_TOKENS:
            return part
    return None


# ---------------------------------------------------------------------------
# 正規化
# ---------------------------------------------------------------------------


def _normalize_mac(s: pd.Series) -> pd.Series:
    return (
        s.astype("string")
        .str.replace(":", "", regex=False)
        .str.replace("-", "", regex=False)
        .str.lower()
    )


def _parse_timestamps(s: pd.Series) -> pd.Series:
    """naive のまま datetime へ変換する（TZ 変換・秒の丸めは行わない）。"""
    text = s.astype("string").str.strip()
    out = pd.to_datetime(text, format="%Y-%m-%d %H:%M:%S", errors="coerce")
    if out.isna().any():
        fallback = pd.to_datetime(text, format="mixed", errors="coerce")
        out = out.fillna(fallback)
    if isinstance(out.dtype, pd.DatetimeTZDtype):
        out = out.dt.tz_localize(None)
    return out


def _str_columns_for(file_type_key: str) -> tuple[str, ...]:
    """種別ごとに「必ず文字列として読む列」を返す。"""
    if file_type_key == RF_NEIGHBORS_FILE_TYPE:
        return _RF_NEIGHBORS_STR_COLUMNS
    return _METRICS_STR_COLUMNS


def _read_table(
    path: Path,
    sheet: str | None,
    columns: Sequence[str],
    as_str: bool,
    str_columns: Sequence[str] = _METRICS_STR_COLUMNS,
) -> pd.DataFrame:
    """1 ファイル（または 1 シート）を DataFrame として読む。"""
    dtype: object
    if as_str:
        dtype = "string"
    else:
        dtype = {c: "string" for c in str_columns if c in columns}
    if sheet is None:
        df = pd.read_csv(
            path,
            dtype=dtype,
            keep_default_na=not as_str,
            na_filter=not as_str,
            encoding="utf-8-sig",
        )
    else:
        df = pd.read_excel(path, sheet_name=sheet, dtype=dtype)
        if as_str:
            df = df.fillna("")
    df.columns = [str(c).strip() for c in df.columns]
    return df.reindex(columns=list(columns))


# ---------------------------------------------------------------------------
# 推定・検出
# ---------------------------------------------------------------------------


def _mode_seconds(diffs: Sequence[float]) -> float | None:
    """差分の最頻値（秒）。中央値は欠測に押し上げられるため使わない。"""
    if not len(diffs):
        return None
    counter = Counter(int(round(d)) for d in diffs if d > 0)
    if not counter:
        return None
    top = max(counter.values())
    # 同数なら小さい方（＝取りこぼしの少ない側）を採る
    return float(min(v for v, c in counter.items() if c == top))


def _estimate_intervals(metrics: pd.DataFrame) -> tuple[list[ApInterval], float | None]:
    """AP ごと・全体の推定サンプリング間隔を返す。"""
    intervals: list[ApInterval] = []
    if metrics.empty:
        return intervals, None
    for ap_id, grp in metrics.groupby("ap_id", sort=True):
        ts = grp["timestamp"].dropna().drop_duplicates().sort_values()
        names = grp["ap_name"].dropna()
        ap_name = str(names.iloc[-1]) if len(names) else ""
        diffs = ts.diff().dropna().dt.total_seconds().to_numpy()
        intervals.append(
            ApInterval(
                ap_id=str(ap_id),
                ap_name=ap_name,
                samples=int(len(ts)),
                interval_seconds=_mode_seconds(diffs),
            )
        )
    per_ap = [a.interval_seconds for a in intervals if a.interval_seconds]
    overall = _mode_seconds(per_ap) if per_ap else None
    return intervals, overall


def _group_intervals(intervals: Sequence[ApInterval]) -> list[tuple[float, int]]:
    """推定間隔を相対許容幅でまとめる（30 秒と 5 分の混在を検出するため）。"""
    counter = Counter(a.interval_seconds for a in intervals if a.interval_seconds)
    groups: list[tuple[float, int]] = []
    for value, n in sorted(counter.items()):
        for i, (rep, total) in enumerate(groups):
            if _close(value, rep):
                groups[i] = (rep, total + n)
                break
        else:
            groups.append((value, n))
    return groups


def _detect_gaps(
    metrics: pd.DataFrame,
    intervals: Sequence[ApInterval],
    overall: float | None,
    gap_factor: float,
) -> pd.DataFrame:
    """AP ごとに、推定間隔 × gap_factor を超える空きをギャップとして返す。

    データは書き換えない（センチネル挿入・補完は行わない）。
    """
    rows: list[dict] = []
    by_ap = {a.ap_id: a for a in intervals}
    if not metrics.empty:
        for ap_id, grp in metrics.groupby("ap_id", sort=True):
            info = by_ap.get(str(ap_id))
            interval = (info.interval_seconds if info else None) or overall
            if not interval:
                continue
            ts = grp["timestamp"].dropna().drop_duplicates().sort_values().reset_index(drop=True)
            if len(ts) < 2:
                continue
            names = grp["ap_name"].dropna()
            ap_name = str(names.iloc[-1]) if len(names) else ""
            deltas = ts.diff().dt.total_seconds()
            threshold = interval * gap_factor
            for i in range(1, len(ts)):
                d = float(deltas.iloc[i])
                if d > threshold:
                    rows.append(
                        {
                            "ap_id": str(ap_id),
                            "ap_name": ap_name,
                            "gap_start": ts.iloc[i - 1],
                            "gap_end": ts.iloc[i],
                            "gap_seconds": d,
                            "missing_samples": int(d // interval) - 1,
                        }
                    )
    gaps = pd.DataFrame(rows, columns=list(GAP_COLUMNS))
    if not gaps.empty:
        gaps = gaps.sort_values(["gap_start", "ap_id"], kind="stable").reset_index(drop=True)
    return gaps


def _site_periods(metrics: pd.DataFrame) -> list[SitePeriod]:
    out: list[SitePeriod] = []
    if metrics.empty or "site_id" not in metrics.columns:
        return out
    for site_id, grp in metrics.groupby(metrics["site_id"].fillna(""), sort=True):
        names = grp["site_name"].dropna()
        out.append(
            SitePeriod(
                site_id=str(site_id),
                site_name=str(names.iloc[0]) if len(names) else "",
                rows=int(len(grp)),
                ap_count=int(grp["ap_id"].nunique()),
                first=grp["timestamp"].min(),
                last=grp["timestamp"].max(),
            )
        )
    return sorted(out, key=lambda s: (s.first, s.site_id))


def _site_label(site_id: str, site_name: str) -> str:
    return f"{site_name} [{site_id}]" if site_name else f"[{site_id}]"


def _resolve_site_filter(metrics: pd.DataFrame, sites: Sequence[str]) -> SiteFilter:
    """指定された文字列（site_id または site_name）を site_id へ解決する。

    ログには site_id と site_name の両方が入っているので、どちらで指定しても通す
    （UI は site_id を送り、CLI では人が読める名前で指定できる）。同じ名前の
    サイトが複数の site_id で存在する場合は、その **すべて** を対象にする。
    """
    requested: list[str] = []
    for raw in sites:
        token = str(raw).strip()
        if token and token not in requested:
            requested.append(token)

    by_id: dict[str, str] = {}
    by_name: dict[str, list[str]] = {}
    if not metrics.empty and "site_id" in metrics.columns:
        pairs = (
            metrics[["site_id", "site_name"]]
            .fillna("")
            .astype(str)
            .drop_duplicates()
        )
        for site_id, site_name in pairs.itertuples(index=False, name=None):
            by_id.setdefault(site_id, site_name)
            if site_name:
                by_name.setdefault(site_name, []).append(site_id)

    selected: list[str] = []
    missing: list[str] = []
    for token in requested:
        hits = [token] if token in by_id else by_name.get(token, [])
        if not hits:
            missing.append(token)
            continue
        for site_id in hits:
            if site_id not in selected:
                selected.append(site_id)

    return SiteFilter(
        requested=tuple(requested),
        site_ids=tuple(selected),
        labels=tuple(_site_label(sid, by_id.get(sid, "")) for sid in selected),
        missing=tuple(missing),
        available=tuple(_site_label(sid, name) for sid, name in sorted(by_id.items())),
        rows_before=int(len(metrics)),
    )


def _filter_by_site(df: pd.DataFrame, site_ids: Sequence[str]) -> pd.DataFrame:
    """``site_id`` が指定に含まれる行だけを残す。"""
    if df.empty or "site_id" not in df.columns:
        return df
    keep = df["site_id"].fillna("").astype(str).isin(list(site_ids))
    return df[keep].reset_index(drop=True)


def _event_blind_spots(
    metrics_period: tuple[datetime, datetime] | None,
    event_times: Sequence[pd.Timestamp],
    threshold: float,
) -> list[tuple[datetime, datetime]]:
    """メトリクス期間のうち、イベントが存在しない区間を返す。"""
    if metrics_period is None:
        return []
    start, end = metrics_period
    spots: list[tuple[datetime, datetime]] = []
    prev = start
    for t in sorted(event_times):
        if t < start or t > end:
            continue
        if (t - prev).total_seconds() > threshold:
            spots.append((prev, t))
        prev = t
    if (end - prev).total_seconds() > threshold:
        spots.append((prev, end))
    return spots


# ---------------------------------------------------------------------------
# 本体
# ---------------------------------------------------------------------------


def load(
    paths: str | os.PathLike | Iterable[str | os.PathLike],
    *,
    file_types: Iterable[str] | None = None,
    gap_factor: float = DEFAULT_GAP_FACTOR,
    event_gap_seconds: float = DEFAULT_EVENT_GAP_SECONDS,
    sites: Iterable[str] | None = None,
) -> LoadResult:
    """ログを結合して正規化済みの DataFrame とレポートを返す。

    :param paths: ファイル／ディレクトリ／glob パターン（複数可）。ディレクトリは再帰探索。
    :param file_types: DataFrame まで読み込む種別（既定は ap_metrics と ap_events）。
        ここに無い種別も判定は行い、ファイル数・行数はレポートに残す。
    :param gap_factor: 推定間隔の何倍を超えたらギャップとみなすか。
    :param event_gap_seconds: 「イベントが存在しない区間」として報告する最小の長さ（秒）。
    :param sites: 対象サイト（site_id または site_name）。None ならすべてのサイト。
        絞り込みは **推定・検出より前** に行う（ギャップも間隔も、絞り込んだ後の
        データで数えないとレポートと結果が食い違う）。解決できなかった指定は
        ``report.site_filter.missing`` に残す（ここでは例外にしない）。
    """
    wanted = tuple(file_types) if file_types is not None else DEFAULT_FILE_TYPES
    for key in wanted:
        if key not in FILE_TYPES_BY_KEY:
            raise ValueError(f"unknown file type: {key}")

    report = LoadReport(gap_factor=gap_factor, event_gap_seconds=event_gap_seconds)
    files = _iter_input_files(paths)
    report.files_scanned = len(files)

    metrics_parts: list[pd.DataFrame] = []
    events_parts: list[pd.DataFrame] = []
    rf_parts: list[pd.DataFrame] = []

    def collect(key: str, df: pd.DataFrame) -> None:
        if key in METRICS_FILE_TYPES:
            metrics_parts.append(df)
        elif key == RF_NEIGHBORS_FILE_TYPE:
            rf_parts.append(df)
        else:
            events_parts.append(df)

    def stat(key: str) -> FileTypeStat:
        st = report.file_stats.get(key)
        if st is None:
            st = FileTypeStat(file_type=key, loaded=key in wanted)
            report.file_stats[key] = st
        return st

    for path in files:
        suffix = path.suffix.lower()
        token = _tz_token_of(path)
        if token:
            report.tz_tokens[token] = report.tz_tokens.get(token, 0) + 1

        if suffix in CSV_SUFFIXES:
            header = _read_csv_header(path)
            ft = detect_file_type(header) if header else None
            if ft is None:
                report.unclassified.append(path.name)
                continue
            st = stat(ft.key)
            st.files += 1
            if ft.key in wanted:
                try:
                    df = _read_table(
                        path, None, ft.columns,
                        as_str=(ft.key == "ap_events"),
                        str_columns=_str_columns_for(ft.key),
                    )
                except (ValueError, OSError, pd.errors.ParserError) as exc:
                    report.warnings.append(f"読み込みに失敗したファイル: {path.name} ({type(exc).__name__})")
                    continue
                st.rows += len(df)
                collect(ft.key, df)
            else:
                st.rows += _count_csv_rows(path)

        elif suffix in EXCEL_SUFFIXES:
            try:
                sheets = _excel_sheet_headers(path)
            except Exception as exc:  # openpyxl が投げる例外は多岐にわたる
                report.unclassified.append(path.name)
                report.warnings.append(f"XLSX を開けませんでした: {path.name} ({type(exc).__name__})")
                continue
            for sheet_name, header, nrows in sheets:
                ft = detect_file_type(header) if header else None
                if ft is None:
                    report.unclassified.append(f"{path.name}#{sheet_name}")
                    continue
                st = stat(ft.key)
                st.files += 1
                if ft.key in wanted:
                    try:
                        df = _read_table(
                            path, sheet_name, ft.columns,
                            as_str=(ft.key == "ap_events"),
                            str_columns=_str_columns_for(ft.key),
                        )
                    except (ValueError, OSError) as exc:
                        report.warnings.append(
                            f"読み込みに失敗したシート: {path.name}#{sheet_name} ({type(exc).__name__})"
                        )
                        continue
                    st.rows += len(df)
                    collect(ft.key, df)
                else:
                    st.rows += nrows
        else:
            report.unclassified.append(path.name)

    if len(report.tz_tokens) > 1:
        report.warnings.append(
            "ファイル名の TZ トークンが混在しています（"
            + ", ".join(f"{k}:{v}" for k, v in sorted(report.tz_tokens.items()))
            + "）。タイムスタンプは naive のまま変換していません"
        )

    metrics = _finalize_metrics(metrics_parts, report)

    # サイトの絞り込み。ap_events は site_id を持たず、イベントは AP 単位で
    # 突き合わせる（detector が ap_name で引く）ため、ここでは絞らない。
    # 絞ると「site_name が空のイベント」を落として結果が変わってしまう。
    if sites is not None:
        report.site_filter = _resolve_site_filter(metrics, list(sites))
        metrics = _filter_by_site(metrics, report.site_filter.site_ids)

    events = _finalize_events(events_parts, report)
    rf_neighbors = _finalize_rf_neighbors(
        rf_parts, report,
        site_ids=report.site_filter.site_ids if report.site_filter else None,
    )

    # --- 推定・検出 ---
    intervals, overall = _estimate_intervals(metrics)
    report.ap_intervals = intervals
    report.overall_interval_seconds = overall
    report.interval_groups = _group_intervals(intervals)
    if len(report.interval_groups) > 1:
        detail = ", ".join(f"{_fmt_secs(rep)}={n}AP" for rep, n in report.interval_groups)
        report.warnings.append(
            f"AP ごとの推定サンプリング間隔がばらついています（{detail}）。"
            "異なる環境のログを混ぜている可能性があります"
        )

    gaps = _detect_gaps(metrics, intervals, overall, gap_factor)
    report.gaps = _summarize_gaps(gaps)

    report.metrics_rows = int(len(metrics))
    report.events_rows = int(len(events))
    report.ap_count = int(metrics["ap_id"].nunique()) if not metrics.empty else 0
    if not metrics.empty:
        report.metrics_period = (metrics["timestamp"].min(), metrics["timestamp"].max())
    if not events.empty:
        report.events_period = (
            events["event_timestamp"].min(),
            events["event_timestamp"].max(),
        )
    report.site_periods = _site_periods(metrics)
    if len(report.site_periods) > 1:
        disjoint = sum(
            1
            for i, a in enumerate(report.site_periods)
            for b in report.site_periods[i + 1:]
            if a.last < b.first or b.last < a.first
        )
        if disjoint:
            report.warnings.append(
                f"出現期間が重ならない site_id の組み合わせが {disjoint} 件あります。"
                "異なる環境（org）のログが混在している可能性があります"
            )

    report.event_blind_spots = _event_blind_spots(
        report.metrics_period,
        events["event_timestamp"].dropna().tolist() if not events.empty else [],
        event_gap_seconds,
    )
    if report.events_rows == 0:
        report.warnings.append("ap_events が 1 件もありません（この期間のイベントログはありません）")

    return LoadResult(
        metrics=metrics, events=events, gaps=gaps, report=report,
        rf_neighbors=rf_neighbors,
    )


def _excel_sheet_headers(path: Path) -> list[tuple[str, list[str], int]]:
    """XLSX の各シートについて (シート名, ヘッダー列, データ行数) を返す。"""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out: list[tuple[str, list[str], int]] = []
        for ws in wb.worksheets:
            header: list[str] = []
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                header = [str(c).strip() for c in row if c is not None]
                break
            nrows = max(0, (ws.max_row or 0) - 1)
            out.append((ws.title, header, nrows))
        return out
    finally:
        wb.close()


def _finalize_metrics(parts: list[pd.DataFrame], report: LoadReport) -> pd.DataFrame:
    """ap_metrics を結合・正規化し、(ap_id, timestamp) で重複排除する。"""
    columns = list(AP_METRICS_COLUMNS)
    if not parts:
        empty = pd.DataFrame(columns=columns)
        empty["timestamp"] = pd.to_datetime(empty["timestamp"])
        return empty

    df = pd.concat(parts, ignore_index=True)
    # ap_metrics_v1（座標列なし）だけが入力の場合でも、列は常に AP_METRICS_COLUMNS 全体を保証する
    df = df.reindex(columns=columns)
    df["timestamp"] = _parse_timestamps(df["timestamp"])
    bad = int(df["timestamp"].isna().sum())
    if bad:
        report.warnings.append(f"timestamp を解釈できなかった ap_metrics の行 {bad} 件を除外しました")
        df = df[df["timestamp"].notna()]
    for col in MAC_COLUMNS & set(df.columns):
        df[col] = _normalize_mac(df[col])
    df["ap_id"] = df["ap_id"].astype("string").fillna("")

    before = len(df)
    df = df.sort_values(["ap_id", "timestamp"], kind="stable")
    df = df[~df.duplicated(subset=["ap_id", "timestamp"], keep="first")]
    removed = before - len(df)
    for key in METRICS_FILE_TYPES:
        if key in report.file_stats:
            report.file_stats[key].duplicates_removed = removed
    return df.reset_index(drop=True)


def _finalize_events(parts: list[pd.DataFrame], report: LoadReport) -> pd.DataFrame:
    """ap_events を結合・正規化し、**全列一致** の行だけを重複排除する。

    同一 AP・同一時刻に別種のイベントが並ぶことは実際にあるため、キーを絞らない。
    """
    columns = list(AP_EVENTS_COLUMNS)
    if not parts:
        empty = pd.DataFrame(columns=columns)
        empty["event_timestamp"] = pd.to_datetime(empty["event_timestamp"])
        return empty

    df = pd.concat(parts, ignore_index=True)
    for col in columns:
        df[col] = df[col].astype("string").fillna("").str.strip()
    for col in MAC_COLUMNS & set(df.columns):
        df[col] = _normalize_mac(df[col])

    before = len(df)
    df = df[~df.duplicated(keep="first")]  # 全列一致のみ除去
    removed = before - len(df)
    if "ap_events" in report.file_stats:
        report.file_stats["ap_events"].duplicates_removed = removed

    df["event_timestamp"] = _parse_timestamps(df["event_timestamp"])
    bad = int(df["event_timestamp"].isna().sum())
    if bad:
        report.warnings.append(f"event_timestamp を解釈できなかった ap_events の行 {bad} 件を除外しました")
        df = df[df["event_timestamp"].notna()]
    for col in _EVENTS_NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].replace("", None), errors="coerce")

    df = df.sort_values(["event_timestamp", "ap_name", "event_type"], kind="stable")
    return df.reset_index(drop=True)


def _finalize_rf_neighbors(
    parts: list[pd.DataFrame],
    report: LoadReport,
    site_ids: Sequence[str] | None = None,
) -> pd.DataFrame:
    """rf_neighbors を結合・正規化する。

    重複判定は ``(site_id, band, ap_mac, neighbor_mac, timestamp)``。
    対称化（max / mean / min）は行わない。方向ごとの行をそのまま保持する。
    日次取得のため複数日分が混ざりうるので、**全時点を保持したまま**
    最新の取得時刻をレポートに記録する（どの時点を使ったかを追えるようにするため）。
    """
    columns = list(RF_NEIGHBORS_COLUMNS)
    if not parts:
        return _empty_rf_neighbors()

    df = pd.concat(parts, ignore_index=True).reindex(columns=columns)
    df["timestamp"] = _parse_timestamps(df["timestamp"])
    bad = int(df["timestamp"].isna().sum())
    if bad:
        report.warnings.append(
            f"timestamp を解釈できなかった rf_neighbors の行 {bad} 件を除外しました"
        )
        df = df[df["timestamp"].notna()]
    for col in MAC_COLUMNS & set(df.columns):
        df[col] = _normalize_mac(df[col])
    for col in ("site_id", "site_name", "band", "ap_name", "neighbor_name"):
        df[col] = df[col].astype("string").fillna("")
    df["rssi"] = pd.to_numeric(df["rssi"], errors="coerce")

    before = len(df)
    df = df.sort_values(list(RF_NEIGHBORS_KEY), kind="stable")
    df = df[~df.duplicated(subset=list(RF_NEIGHBORS_KEY), keep="first")]
    removed = before - len(df)
    if RF_NEIGHBORS_FILE_TYPE in report.file_stats:
        report.file_stats[RF_NEIGHBORS_FILE_TYPE].duplicates_removed = removed

    # サイトを絞る場合はレポートの集計より前に落とす（対象外サイトの取得時刻を
    # 「分析に使用」と報告してしまわないため）
    if site_ids is not None:
        df = _filter_by_site(df, site_ids)

    df = df.reset_index(drop=True)
    report.rf_neighbors_rows = int(len(df))
    if not df.empty:
        counts = df.groupby("timestamp").size().sort_index()
        report.rf_neighbors_snapshots = [(ts, int(n)) for ts, n in counts.items()]
        report.rf_neighbors_latest = counts.index.max()
    return df


def latest_rf_neighbors(rf_neighbors: pd.DataFrame) -> pd.DataFrame:
    """最新の取得時刻の行だけを返す（日次取得なので通常は 1 日分）。"""
    if rf_neighbors.empty:
        return rf_neighbors
    latest = rf_neighbors["timestamp"].max()
    return rf_neighbors[rf_neighbors["timestamp"] == latest].reset_index(drop=True)


def _summarize_gaps(gaps: pd.DataFrame) -> GapSummary:
    if gaps.empty:
        return GapSummary()
    idx = gaps["gap_seconds"].idxmax()
    row = gaps.loc[idx]
    return GapSummary(
        count=int(len(gaps)),
        total_seconds=float(gaps["gap_seconds"].sum()),
        max_seconds=float(row["gap_seconds"]),
        max_ap_name=str(row["ap_name"]),
        max_start=row["gap_start"],
        max_end=row["gap_end"],
        total_missing_samples=int(gaps["missing_samples"].sum()),
    )
