"""CLI と API が共用する分析パイプラインと出力の書き出し。

``hangap/cli.py`` と ``backend/routers/hangap.py`` は、どちらもこのモジュールだけを
呼ぶ。``loader.load()`` / ``detector.detect()`` / ``neighbors.build_context()`` の
ロジックはここでも再実装しない（呼び出し順と整形だけを持つ）。CLI と UI で結果が
食い違わないことがこのモジュールの存在理由である。

ネットワークアクセス・LLM 呼び出しは行わない。
"""
from __future__ import annotations

import warnings as warnings_module
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import detector, loader, neighbors

#: 入力として受け付けるファイル拡張子
DATA_SUFFIXES = loader.CSV_SUFFIXES | loader.EXCEL_SUFFIXES

#: 受け付ける時刻表記（これで解釈できなければ pandas に委ねる）
TIME_FORMATS: tuple[str, ...] = ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S")

#: 進捗フェーズ（利用者に「まだ動いている」と伝えるためのもの。進捗率ではない）
PHASE_LOADING = "loading"
PHASE_NEIGHBORS = "neighbors"
PHASE_DETECTING = "detecting"
PHASE_WRITING = "writing"

_RECOVERED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

STATUS_ORDER: tuple[str, ...] = (
    detector.STATUS_RECOVERED,
    detector.STATUS_ONGOING,
    detector.STATUS_CUT_GAP,
    detector.STATUS_CUT_AP_DOWN,
)

VERDICT_ORDER: tuple[str, ...] = (
    neighbors.VERDICT_PRESENT,
    neighbors.VERDICT_ABSENT,
    neighbors.VERDICT_UNKNOWN,
)


# ---------------------------------------------------------------------------
# 例外
# ---------------------------------------------------------------------------


class AnalysisError(RuntimeError):
    """分析を続行できない状態（CLI は終了コード 1、API は 400 / failed）。"""


class ParamError(AnalysisError):
    """パラメータが不正。``field`` にどの項目が不正かを持つ。"""

    def __init__(self, message: str, field_name: str | None = None) -> None:
        super().__init__(message)
        self.field_name = field_name


class UnclassifiedInputError(AnalysisError):
    """入力ファイルの種別を判定できなかった。"""


class NoMetricsError(AnalysisError):
    """ap_metrics を1行も読み込めなかった。

    **検出0件とは別の状態である。** 「ハングが無かった」ではなく「そもそも分析対象が
    無かった」であり、検出0件と同じ扱いにすると結果を誤読する。
    """


# ---------------------------------------------------------------------------
# パラメータ
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class AnalysisParams:
    """分析条件。既定値は必ず hangap 側の定数を参照する（CLI と API で二重定義しない）。"""

    window_start: pd.Timestamp | None = None
    window_end: pd.Timestamp | None = None
    min_zero_samples: int = detector.DEFAULT_MIN_ZERO_SAMPLES
    min_zero_duration: pd.Timedelta | None = None
    event_window: pd.Timedelta = field(
        default_factory=lambda: pd.Timedelta(detector.DEFAULT_EVENT_WINDOW)
    )
    exodus_threshold: float = detector.DEFAULT_EXODUS_THRESHOLD
    gap_factor: float = loader.DEFAULT_GAP_FACTOR
    neighbor_count: int = neighbors.DEFAULT_NEIGHBOR_COUNT
    max_distance_m: float = neighbors.DEFAULT_MAX_DISTANCE_M
    neighbor_client_threshold: float = neighbors.DEFAULT_NEIGHBOR_CLIENT_THRESHOLD
    truncated_warn_ratio: float = detector.DEFAULT_TRUNCATED_WARN_RATIO


# ---------------------------------------------------------------------------
# 時刻・時間指定のパース
# ---------------------------------------------------------------------------


def parse_time(text: str, label: str, field_name: str | None = None) -> pd.Timestamp:
    """``YYYY-MM-DD HH:MM`` / ISO8601 を naive な Timestamp にする（TZ 付きは拒否）。"""
    text = str(text).strip()
    for fmt in TIME_FORMATS:
        try:
            return pd.Timestamp(datetime.strptime(text, fmt))
        except ValueError:
            continue
    try:
        ts = pd.Timestamp(text)
    except Exception as exc:  # 多様な例外を投げうる外部入力の境界
        raise ParamError(f"{label} を解釈できません: {text!r}", field_name) from exc
    if ts.tzinfo is not None:
        raise ParamError(
            f"{label} にタイムゾーンは付けられません（ログが naive のため）: {text!r}",
            field_name,
        )
    return ts


def parse_duration(text: str, label: str, field_name: str | None = None) -> pd.Timedelta:
    try:
        td = pd.Timedelta(text)
    except Exception as exc:
        raise ParamError(f"{label} を解釈できません: {text!r}", field_name) from exc
    if pd.isna(td):
        raise ParamError(f"{label} を解釈できません: {text!r}", field_name)
    return td


# ---------------------------------------------------------------------------
# 入力ファイルの収集
# ---------------------------------------------------------------------------


def collect_files(directory: str | Path) -> list[Path]:
    """ディレクトリ配下の CSV/XLSX を列挙する（**呼び出し時点で確定させる**）。

    分析中に ``save_hourly_logs`` がファイルを足しても結果が揺れないよう、
    ジョブ開始時にこの一覧を固定して使う。
    """
    root = Path(directory)
    if not root.is_dir():
        return []
    return [f for f in sorted(root.rglob("*")) if f.is_file() and f.suffix.lower() in DATA_SUFFIXES]


# ---------------------------------------------------------------------------
# 整形ヘルパ
# ---------------------------------------------------------------------------


def fmt_dt(dt: object) -> str:
    if dt is None or pd.isna(dt):
        return "-"
    return pd.Timestamp(dt).strftime("%Y-%m-%d %H:%M:%S")


def fmt_period(period: tuple | None) -> str:
    if period is None:
        return "（なし）"
    return f"{fmt_dt(period[0])} 〜 {fmt_dt(period[1])}"


def fmt_window(ws: pd.Timestamp | None, we: pd.Timestamp | None) -> str:
    left = fmt_dt(ws) if ws is not None else "(指定なし)"
    right = fmt_dt(we) if we is not None else "(指定なし)"
    return f"{left} 〜 {right}"


def fmt_td(td: pd.Timedelta) -> str:
    total = pd.Timedelta(td).total_seconds()
    if total % 3600 == 0 and total >= 3600:
        return f"{total / 3600:g}h"
    if total % 60 == 0:
        return f"{total / 60:g}m"
    return f"{total:g}s"


def condition_text(params: AnalysisParams, n_files: int) -> str:
    zero_desc = (
        f"min_zero_duration={fmt_td(params.min_zero_duration)}"
        if params.min_zero_duration is not None
        else f"min_zero_samples={params.min_zero_samples}"
    )
    return (
        f"分析条件: 窓 {fmt_window(params.window_start, params.window_end)} / {zero_desc} / "
        f"event_window={fmt_td(params.event_window)} / exodus_threshold={params.exodus_threshold} / "
        f"gap_factor={params.gap_factor} / 入力ファイル数={n_files} / "
        f"neighbor_count={params.neighbor_count} / max_distance_m={params.max_distance_m:g} / "
        f"neighbor_client_threshold={params.neighbor_client_threshold:g}（周辺AP判定の既定値は暫定）"
    )


def coverage_and_warnings_text(
    report: loader.LoadReport,
    detector_warnings: Sequence[str],
    quality_warnings: Sequence[str] = (),
) -> str:
    lines = [
        f"データ範囲: metrics {fmt_period(report.metrics_period)} / "
        f"events {fmt_period(report.events_period)}"
    ]
    all_warnings = list(report.warnings) + list(detector_warnings) + list(quality_warnings)
    if all_warnings:
        lines.append(f"警告 {len(all_warnings)} 件:")
        lines.extend(f"  ⚠ {w}" for w in all_warnings)
    else:
        lines.append("警告: なし")
    return "\n".join(lines)


def format_result_summary(df: pd.DataFrame) -> str:
    total = len(df)
    lines = [f"検出区間数: {total}"]
    if total:
        counts = df["回復状況"].value_counts()
        for status in STATUS_ORDER:
            lines.append(f"  {status}: {int(counts.get(status, 0))}")
        lines.append(f"退場疑い: {int(df['退場疑い'].sum())} 件")
        lines.append(f"イベントが該当した区間数: {int((df['AP Event（±30分）'] == 'あり').sum())} 件")
        # 周辺AP判定は判断材料であって絞り込み条件ではない。内訳を出すだけで行は落とさない。
        verdicts = df["周辺AP判定"].value_counts()
        lines.append("周辺AP判定:")
        for verdict in VERDICT_ORDER:
            lines.append(f"  {verdict}: {int(verdicts.get(verdict, 0))}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 分析
# ---------------------------------------------------------------------------


@dataclass
class Meta:
    """出力ファイルの先頭に埋め込むメタ情報。"""

    title: str
    condition_text: str
    coverage_and_warnings_text: str
    result_summary_text: str


@dataclass
class AnalysisResult:
    """:func:`run_analysis` の戻り値。"""

    params: AnalysisParams
    n_files: int
    report: loader.LoadReport
    result: pd.DataFrame
    detector_warnings: list[str]
    quality_warnings: list[str]
    neighbor_context: neighbors.NeighborContext

    @property
    def all_warnings(self) -> list[str]:
        return list(self.report.warnings) + self.detector_warnings + self.quality_warnings

    def meta(self, title: str = "ハングAP分析結果") -> Meta:
        return Meta(
            title=title,
            condition_text=condition_text(self.params, self.n_files),
            coverage_and_warnings_text=coverage_and_warnings_text(
                self.report, self.detector_warnings, self.quality_warnings
            ),
            result_summary_text=format_result_summary(self.result),
        )


def run_analysis(
    files: Sequence[Path],
    params: AnalysisParams | None = None,
    *,
    on_phase: Callable[[str], None] | None = None,
) -> AnalysisResult:
    """ログを読み込み、ゼロクライアント区間を検出する（CLI と API の共通経路）。

    :raises UnclassifiedInputError: 入力ファイルの種別を判定できなかった
    :raises NoMetricsError: ap_metrics を 1 行も読み込めなかった（検出0件とは別）
    """
    p = params or AnalysisParams()

    def phase(name: str) -> None:
        if on_phase is not None:
            on_phase(name)

    phase(PHASE_LOADING)
    load_result = loader.load(list(files), gap_factor=p.gap_factor)
    report = load_result.report

    if report.metrics_rows == 0:
        if report.unclassified:
            sample = ", ".join(report.unclassified[:5])
            more = " ..." if len(report.unclassified) > 5 else ""
            raise UnclassifiedInputError(
                "入力ファイルの種別を判定できませんでした"
                f"（ap_metrics / ap_events のいずれにも一致しません）: {sample}{more}"
            )
        raise NoMetricsError(
            "ap_metrics を 1 行も読み込めませんでした"
            f"（走査ファイル数={report.files_scanned}）。"
            "分析対象のログが存在しないか、保存期間の設定で削除された可能性があります。"
            "これは「ハングが検出されなかった（0 件）」とは別の状態です。"
        )

    # 近傍AP のインデックスは検出と explain で共有する（座標は AP の最新行から 1 度だけ取る）
    phase(PHASE_NEIGHBORS)
    neighbor_context = neighbors.build_context(
        load_result.metrics,
        load_result.rf_neighbors,
        neighbor_count=p.neighbor_count,
        max_distance_m=p.max_distance_m,
    )

    phase(PHASE_DETECTING)
    with warnings_module.catch_warnings(record=True) as caught:
        warnings_module.simplefilter("always")
        result_df = detector.detect(
            load_result.metrics,
            load_result.events,
            load_result.gaps,
            window_start=p.window_start,
            window_end=p.window_end,
            min_zero_samples=p.min_zero_samples,
            min_zero_duration=p.min_zero_duration,
            event_window=p.event_window,
            exodus_threshold=p.exodus_threshold,
            neighbor_context=neighbor_context,
            neighbor_client_threshold=p.neighbor_client_threshold,
        )
    detector_warnings = [
        str(w.message) for w in caught if issubclass(w.category, UserWarning)
    ]

    truncated = detector.truncated_warning(result_df, p.truncated_warn_ratio)

    return AnalysisResult(
        params=p,
        n_files=len(files),
        report=report,
        result=result_df,
        detector_warnings=detector_warnings,
        quality_warnings=[truncated] if truncated else [],
        neighbor_context=neighbor_context,
    )


# ---------------------------------------------------------------------------
# 出力（CLI と API で同一のファイルを作るため、書き出しはここだけに置く）
# ---------------------------------------------------------------------------


def _cell_value(value: object) -> object:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        try:
            return value.item()
        except (TypeError, ValueError):
            return value
    return value


def write_xlsx(path: Path, df: pd.DataFrame, meta: Meta) -> Path:
    columns = detector.RESULT_COLUMNS
    ncols = len(columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "ハングAP分析結果"

    ws.cell(row=1, column=1, value=meta.title).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=meta.condition_text)
    c3 = ws.cell(row=3, column=1, value=meta.coverage_and_warnings_text)
    c3.alignment = Alignment(wrap_text=True, vertical="top")

    header_row = 5
    for col, name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = Font(bold=True)

    status_col = columns.index("回復状況") + 1
    for r, row in enumerate(df.itertuples(index=False, name=None), start=header_row + 1):
        for col, value in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=_cell_value(value))
        if row[status_col - 1] == detector.STATUS_RECOVERED:
            for col in range(1, ncols + 1):
                ws.cell(row=r, column=col).fill = _RECOVERED_FILL

    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def write_csv(path: Path, df: pd.DataFrame) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def write_summary(path: Path, meta: Meta) -> Path:
    text = "\n\n".join([
        meta.title,
        meta.condition_text,
        meta.coverage_and_warnings_text,
        meta.result_summary_text,
    ])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text + "\n", encoding="utf-8")
    return path
