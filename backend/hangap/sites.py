"""``data/logs`` に含まれるサイトの一覧を作る（分析対象を選ぶための材料）。

設計方針:

- 選択肢は **ログから作る**。``/api/sites``（現在の監視対象）からは作らない。
  環境を切り替えると ``data/logs`` には現在監視していないサイトのログが残るため、
  監視対象だけを選択肢にすると、そのログを分析できなくなる。
- 全ファイルを pandas で読むと重い（実測 4,000 ファイル / 60 万行）。ここで欲しいのは
  「どのサイトが・何台の AP で・いつからいつまで含まれるか」だけなので、
  **ap_metrics の 4 列だけ**を行から切り出して数える。期間・行数の厳密さは要らない。
- 種別判定はヘッダーの列集合で行う（:mod:`pseudonymizer.schemas`）。ファイル名は使わない。
- ネットワークアクセス・LLM 呼び出しは行わない。入力はローカルファイルのみ。
"""
from __future__ import annotations

import csv
import re
import threading
from dataclasses import dataclass, field, replace
from datetime import datetime, timezone
from pathlib import Path
from typing import Sequence

from pseudonymizer.schemas import detect_file_type

from . import loader

#: 走査に必要な列。ap_metrics / ap_metrics_v1 のどちらにも存在する
SCAN_COLUMNS: tuple[str, ...] = ("timestamp", "site_id", "site_name", "ap_id")

#: 期間として採用する時刻表記。この形なら文字列のまま大小比較できる
_TS_RE = re.compile(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}")


@dataclass(frozen=True)
class LogSite:
    """ログに含まれる 1 サイト分の要約。"""

    site_id: str
    site_name: str
    ap_count: int
    rows: int
    files: int
    #: ログ中の時刻表記そのまま（naive）。解釈できない値しか無ければ None
    first: str | None
    last: str | None


@dataclass(frozen=True)
class SiteScan:
    """:func:`scan` の結果。"""

    sites: tuple[LogSite, ...]
    files_scanned: int
    metrics_files: int
    scanned_at: datetime
    #: キャッシュを返したか（利用者が「再取得」を判断するための情報）
    cached: bool = False


@dataclass
class _Acc:
    """サイトごとの集計中の状態。"""

    site_id: str
    site_name: str = ""
    rows: int = 0
    files: int = 0
    ap_ids: set[str] = field(default_factory=set)
    first: str | None = None
    last: str | None = None

    def add(self, site_name: str, ap_id: str, timestamp: str) -> None:
        self.rows += 1
        if site_name and not self.site_name:
            self.site_name = site_name
        if ap_id:
            self.ap_ids.add(ap_id)
        if _TS_RE.match(timestamp):
            if self.first is None or timestamp < self.first:
                self.first = timestamp
            if self.last is None or timestamp > self.last:
                self.last = timestamp

    def to_site(self) -> LogSite:
        return LogSite(
            site_id=self.site_id,
            site_name=self.site_name,
            ap_count=len(self.ap_ids),
            rows=self.rows,
            files=self.files,
            first=self.first,
            last=self.last,
        )


# ---------------------------------------------------------------------------
# 走査
# ---------------------------------------------------------------------------


def _metrics_indexes(header: Sequence[str]) -> dict[str, int] | None:
    """ap_metrics のヘッダーなら、必要な列の位置を返す。違えば None。"""
    columns = [str(c).strip() for c in header]
    ft = detect_file_type(columns)
    if ft is None or ft.key not in loader.METRICS_FILE_TYPES:
        return None
    try:
        return {c: columns.index(c) for c in SCAN_COLUMNS}
    except ValueError:  # 種別が一致していれば起きないが、壊れたヘッダーで落ちない
        return None


def _scan_csv(path: Path, acc: dict[str, _Acc]) -> bool:
    """CSV を 1 本走査する。ap_metrics でなければ False。

    1 行を丸ごと csv でパースすると重いので、クォートを含まない行は
    ``split`` で必要な列までだけ切り出す（site_name にカンマが入る行だけ csv に回す）。
    """
    try:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            first_line = f.readline()
            if not first_line:
                return False
            header = next(csv.reader([first_line]), None)
            if not header:
                return False
            idx = _metrics_indexes(header)
            if idx is None:
                return False
            ts_i, site_i, name_i, ap_i = (idx[c] for c in SCAN_COLUMNS)
            width = max(idx.values()) + 1
            seen: set[str] = set()
            for line in f:
                if not line.strip():
                    continue
                if '"' in line:
                    fields = next(csv.reader([line]), None)
                    if fields is None:
                        continue
                else:
                    fields = line.rstrip("\r\n").split(",", width)
                if len(fields) < width:
                    continue
                site_id = fields[site_i].strip()
                seen.add(site_id)
                _get(acc, site_id).add(
                    fields[name_i].strip(), fields[ap_i].strip(), fields[ts_i].strip()
                )
            for site_id in seen:
                acc[site_id].files += 1
    except OSError:
        return False
    return True


def _scan_excel(path: Path, acc: dict[str, _Acc]) -> bool:
    """XLSX を 1 本走査する（シートごとに判定する）。"""
    from openpyxl import load_workbook

    found = False
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # openpyxl が投げる例外は多岐にわたる。一覧作成は止めない
        return False
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            idx = _metrics_indexes([c for c in header if c is not None])
            if idx is None:
                continue
            found = True
            ts_i, site_i, name_i, ap_i = (idx[c] for c in SCAN_COLUMNS)
            width = max(idx.values()) + 1
            seen: set[str] = set()
            for row in rows:
                if row is None or len(row) < width:
                    continue
                site_id = _text(row[site_i])
                seen.add(site_id)
                _get(acc, site_id).add(_text(row[name_i]), _text(row[ap_i]), _text(row[ts_i]))
            for site_id in seen:
                acc[site_id].files += 1
    finally:
        wb.close()
    return found


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _get(acc: dict[str, _Acc], site_id: str) -> _Acc:
    a = acc.get(site_id)
    if a is None:
        a = _Acc(site_id=site_id)
        acc[site_id] = a
    return a


def _scan(files: Sequence[Path]) -> SiteScan:
    acc: dict[str, _Acc] = {}
    metrics_files = 0
    for path in files:
        suffix = path.suffix.lower()
        if suffix in loader.CSV_SUFFIXES:
            hit = _scan_csv(path, acc)
        elif suffix in loader.EXCEL_SUFFIXES:
            hit = _scan_excel(path, acc)
        else:
            hit = False
        if hit:
            metrics_files += 1
    # 名前順。名前が空のサイト（ログに site_name が入っていない）は末尾へ回す
    sites = sorted(
        (a.to_site() for a in acc.values()),
        key=lambda s: (s.site_name == "", s.site_name, s.site_id),
    )
    return SiteScan(
        sites=tuple(sites),
        files_scanned=len(files),
        metrics_files=metrics_files,
        scanned_at=datetime.now(timezone.utc),
    )


# ---------------------------------------------------------------------------
# キャッシュ（プロセス内。入力ファイルが変われば自動で作り直す）
# ---------------------------------------------------------------------------

_LOCK = threading.Lock()
_CACHE: tuple[tuple, SiteScan] | None = None


def _signature(files: Sequence[Path]) -> tuple:
    """入力ファイルの同一性。1 本でも増減・更新されればキャッシュを捨てる。"""
    sig: list[tuple[str, int, int]] = []
    for path in files:
        try:
            st = path.stat()
        except OSError:
            continue
        sig.append((str(path), st.st_mtime_ns, st.st_size))
    return tuple(sig)


def scan(files: Sequence[Path], *, refresh: bool = False) -> SiteScan:
    """ログを走査してサイト一覧を返す。

    :param refresh: True ならキャッシュを使わずに読み直す（明示的な再取得）。
    """
    global _CACHE
    files = list(files)
    signature = _signature(files)
    if not refresh:
        with _LOCK:
            cached = _CACHE
        if cached is not None and cached[0] == signature:
            return replace(cached[1], cached=True)

    result = _scan(files)
    with _LOCK:
        _CACHE = (signature, result)
    return result


def clear_cache() -> None:
    """キャッシュを捨てる（テストと明示的な再取得のため）。"""
    global _CACHE
    with _LOCK:
        _CACHE = None
