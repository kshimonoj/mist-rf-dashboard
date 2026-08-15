"""analyze CLI の周辺AP判定まわり（--explain / 打ち切り警告 / 周辺AP列の出力）。

合成データのみを使う（実データは使わない）。
"""
from __future__ import annotations

import csv
from datetime import datetime, timedelta
from pathlib import Path

import _synth as S
from openpyxl import load_workbook

from hangap import cli, neighbors
from hangap.detector import RESULT_COLUMNS

START = datetime(2026, 1, 1, 10, 0, 0)
INTERVAL = 60

TARGET = "TARGET-AP"
NEIGHBOR = "NEAR-AP"

#: 一目で偽物と分かる MAC（コロンなし小文字）
_MACS: dict[str, str] = {TARGET: "aabbccddee01", NEIGHBOR: "aabbccddee02"}

#: index 3〜12 の 10 サンプルがゼロ区間
TARGET_VALUES: list[int] = [1, 1, 1] + [0] * 10 + [1] * 7


def _rows(ap_name: str, x_m: float, values: list[int], *, skip: set[int] = frozenset()) -> list[dict]:
    return [
        S.metrics_row(
            START + timedelta(seconds=INTERVAL * i),
            ap_id=f"test-ap-{ap_name.lower()}",
            ap_name=ap_name,
            mac=_MACS[ap_name],
            num_clients=v,
            map_id="map-a",
            x_m=x_m,
            y_m=0.0,
        )
        for i, v in enumerate(values)
        if i not in skip
    ]


def _write_site(tmp_path: Path, *, neighbor_clients: int = 8) -> Path:
    rows = _rows(TARGET, 0.0, TARGET_VALUES)
    rows += _rows(NEIGHBOR, 10.0, [neighbor_clients] * len(TARGET_VALUES))
    S.write_metrics(tmp_path / "ap_metrics.csv", rows)
    return tmp_path


# ---------------------------------------------------------------------------
# 1. 周辺AP列が出力に載ること
# ---------------------------------------------------------------------------


def test_neighbor_columns_are_written_to_csv_and_xlsx(tmp_path):
    _write_site(tmp_path)
    out = tmp_path / "out"
    assert cli.main(["analyze", str(tmp_path), "--out", str(out)]) == cli.EXIT_OK

    csv_path = next(out.glob("*.csv"))
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    target = [r for r in rows if r["ap_name"] == TARGET]
    assert len(target) == 1
    assert target[0]["周辺AP名"] == NEIGHBOR
    assert target[0]["周辺AP距離"] == "10.0"
    assert target[0]["周辺AP端末数"] == "8.0"
    assert target[0]["周辺AP判定"] == neighbors.VERDICT_PRESENT
    assert target[0]["周辺AP RF隣接数"] == ""  # rf_neighbors 無し

    ws = load_workbook(next(out.glob("*.xlsx"))).active
    header = [ws.cell(row=5, column=c).value for c in range(1, len(RESULT_COLUMNS) + 1)]
    assert header == list(RESULT_COLUMNS)


def test_neighbor_options_are_recorded_in_the_condition_line(tmp_path):
    _write_site(tmp_path)
    out = tmp_path / "out"
    assert cli.main([
        "analyze", str(tmp_path), "--out", str(out),
        "--neighbor-count", "6", "--max-distance-m", "30", "--neighbor-client-threshold", "2",
    ]) == cli.EXIT_OK

    summary = next(out.glob("*_summary.txt")).read_text(encoding="utf-8")
    assert "neighbor_count=6" in summary
    assert "max_distance_m=30" in summary
    assert "neighbor_client_threshold=2" in summary
    assert "暫定" in summary  # 既定値が確定値でないことを結果にも残す


# ---------------------------------------------------------------------------
# 2. --explain
# ---------------------------------------------------------------------------


def test_explain_prints_the_reasoning_for_the_requested_ap(tmp_path, capsys):
    _write_site(tmp_path)
    out = tmp_path / "out"
    assert cli.main([
        "analyze", str(tmp_path), "--out", str(out), "--explain", TARGET,
    ]) == cli.EXIT_OK

    stdout = capsys.readouterr().out
    assert f"[ 判定根拠: {TARGET} ]" in stdout
    assert "区間 #1" in stdout
    assert "近傍AP（距離 <= 25.0m / 上位 4 台）" in stdout
    assert NEIGHBOR in stdout
    assert f"→ {neighbors.VERDICT_PRESENT}" in stdout


def test_explain_accepts_multiple_aps(tmp_path, capsys):
    _write_site(tmp_path)
    out = tmp_path / "out"
    assert cli.main([
        "analyze", str(tmp_path), "--out", str(out),
        "--explain", TARGET, "--explain", NEIGHBOR,
    ]) == cli.EXIT_OK

    stdout = capsys.readouterr().out
    assert f"[ 判定根拠: {TARGET} ]" in stdout
    assert f"[ 判定根拠: {NEIGHBOR} ]" in stdout


def test_explain_for_unknown_ap_is_not_an_error(tmp_path, capsys):
    _write_site(tmp_path)
    out = tmp_path / "out"
    assert cli.main([
        "analyze", str(tmp_path), "--out", str(out), "--explain", "NO-SUCH-AP",
    ]) == cli.EXIT_OK

    assert "該当する区間がありません" in capsys.readouterr().out


# ---------------------------------------------------------------------------
# 3. 打ち切り(欠測)比率の警告
# ---------------------------------------------------------------------------


def _truncation_heavy_rows() -> list[dict]:
    """ゼロ区間の途中に欠測を挟み、2 区間のうち 1 区間を「打ち切り(欠測)」にする。"""
    values = [1, 1, 1] + [0] * 8 + [0] * 6 + [0] * 8 + [1, 1, 1]
    skip = set(range(11, 17))  # 6 サンプル欠測（サンプリング間隔の 1.5 倍を大きく超える）
    rows = _rows(TARGET, 0.0, values, skip=skip)
    rows += _rows(NEIGHBOR, 10.0, [8] * len(values), skip=skip)
    return rows


def test_truncated_ratio_warning_appears_in_stdout_and_output_files(tmp_path, capsys):
    S.write_metrics(tmp_path / "ap_metrics.csv", _truncation_heavy_rows())
    out = tmp_path / "out"
    assert cli.main([
        "analyze", str(tmp_path), "--out", str(out), "--format", "both",
    ]) == cli.EXIT_OK

    stdout = capsys.readouterr().out
    assert "[ データ品質の警告 ]" in stdout
    assert "データ欠測により打ち切られています" in stdout
    assert "%" in stdout

    summary = next(out.glob("*_summary.txt")).read_text(encoding="utf-8")
    assert "データ欠測により打ち切られています" in summary

    ws = load_workbook(next(out.glob("*.xlsx"))).active
    assert "データ欠測により打ち切られています" in ws.cell(row=3, column=1).value


def test_truncated_ratio_warning_is_suppressed_below_the_threshold(tmp_path, capsys):
    S.write_metrics(tmp_path / "ap_metrics.csv", _truncation_heavy_rows())
    out = tmp_path / "out"
    assert cli.main([
        "analyze", str(tmp_path), "--out", str(out), "--truncated-warn-ratio", "0.9",
    ]) == cli.EXIT_OK

    assert "データ品質の警告" not in capsys.readouterr().out


def test_no_truncation_means_no_warning(tmp_path, capsys):
    _write_site(tmp_path)
    out = tmp_path / "out"
    assert cli.main(["analyze", str(tmp_path), "--out", str(out)]) == cli.EXIT_OK

    assert "データ品質の警告" not in capsys.readouterr().out
