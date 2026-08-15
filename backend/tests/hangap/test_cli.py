"""hangap analyze CLI の配線テスト。合成データのみを使う（実データは使わない）。"""
from __future__ import annotations

import csv
from datetime import datetime, timedelta
from pathlib import Path

import _synth as S
import pytest
from openpyxl import load_workbook

from hangap import cli
from hangap.detector import RESULT_COLUMNS

START = datetime(2026, 1, 1, 10, 0, 0)
INTERVAL = 60


def _series(ap_id: str, ap_name: str, start: datetime, values: list[int]) -> list[dict]:
    return [
        S.metrics_row(start + timedelta(seconds=INTERVAL * i), ap_id=ap_id, ap_name=ap_name, num_clients=v)
        for i, v in enumerate(values)
    ]


def _two_ap_rows() -> list[dict]:
    """AAA-AP は 1 区間、BBB-AP は 2 区間。並び順テスト用に名前を意図的に離す。"""
    aaa = _series("test-ap-aaa", "AAA-AP", START, [1, 1, 1] + [0] * 7 + [1, 1, 1])
    bbb = _series(
        "test-ap-bbb", "BBB-AP", START,
        [1, 1, 1] + [0] * 6 + [1, 1, 1] + [0] * 8 + [1, 1, 1],
    )
    return aaa + bbb


def _write_two_ap(tmp_path: Path) -> Path:
    S.write_metrics(tmp_path / "ap_metrics.csv", _two_ap_rows())
    return tmp_path


def _run(argv: list[str]) -> int:
    return cli.main(argv)


# ---------------------------------------------------------------------------
# 1. 時刻の任意指定
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("use_from,use_to", [(False, False), (True, False), (False, True), (True, True)])
def test_time_window_is_optional(tmp_path, capsys, use_from, use_to):
    _write_two_ap(tmp_path)
    out = tmp_path / "out"
    argv = ["analyze", str(tmp_path), "--out", str(out)]
    if use_from:
        argv += ["--from", "2026-01-01 09:00"]
    if use_to:
        argv += ["--to", "2026-01-01 12:00"]
    assert _run(argv) == cli.EXIT_OK
    assert list(out.glob("*.xlsx"))


# ---------------------------------------------------------------------------
# 2. 時刻フォーマット
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("text", ["2026-01-01 09:00", "2026-01-01T09:00:00"])
def test_time_formats_accepted(tmp_path, text):
    _write_two_ap(tmp_path)
    out = tmp_path / "out"
    assert _run(["analyze", str(tmp_path), "--out", str(out), "--from", text]) == cli.EXIT_OK


def test_timezone_aware_time_is_rejected(tmp_path):
    _write_two_ap(tmp_path)
    out = tmp_path / "out"
    rc = _run(["analyze", str(tmp_path), "--out", str(out), "--from", "2026-01-01T09:00:00+09:00"])
    assert rc == cli.EXIT_INPUT_ERROR


# ---------------------------------------------------------------------------
# 3. --min-zero-duration の表記と優先順位
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("duration", ["30m", "25min", "1h"])
def test_min_zero_duration_formats_are_parsed(tmp_path, duration):
    _write_two_ap(tmp_path)
    out = tmp_path / "out"
    assert _run(["analyze", str(tmp_path), "--out", str(out), "--min-zero-duration", duration]) == cli.EXIT_OK


def test_min_zero_duration_overrides_min_zero_samples(tmp_path):
    _write_two_ap(tmp_path)

    out_samples = tmp_path / "out_samples"
    _run(["analyze", str(tmp_path), "--out", str(out_samples), "--min-zero-samples", "1000"])
    csv_samples = next(out_samples.glob("*.csv"))
    with open(csv_samples, newline="", encoding="utf-8-sig") as f:
        rows_samples = list(csv.reader(f))
    assert len(rows_samples) == 1  # ヘッダーのみ（しきい値が高すぎて 0 件）

    out_duration = tmp_path / "out_duration"
    _run([
        "analyze", str(tmp_path), "--out", str(out_duration),
        "--min-zero-samples", "1000", "--min-zero-duration", "4m",
    ])
    csv_duration = next(out_duration.glob("*.csv"))
    with open(csv_duration, newline="", encoding="utf-8-sig") as f:
        rows_duration = list(csv.reader(f))
    assert len(rows_duration) > 1  # duration 指定が優先され、区間が検出される


# ---------------------------------------------------------------------------
# 4. 検出 0 件
# ---------------------------------------------------------------------------


def test_zero_detections_exit_ok_with_empty_result_files(tmp_path):
    rows = _series("test-ap-aaa", "AAA-AP", START, [1] * 10)
    S.write_metrics(tmp_path / "ap_metrics.csv", rows)
    out = tmp_path / "out"

    rc = _run(["analyze", str(tmp_path), "--out", str(out), "--format", "both"])
    assert rc == cli.EXIT_OK

    csv_path = next(out.glob("*.csv"))
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        data_rows = list(csv.reader(f))
    assert len(data_rows) == 1  # ヘッダーのみ

    xlsx_path = next(out.glob("*.xlsx"))
    ws = load_workbook(xlsx_path).active
    assert ws.cell(row=6, column=1).value is None


# ---------------------------------------------------------------------------
# 5. --out 必須
# ---------------------------------------------------------------------------


def test_out_is_required(tmp_path):
    _write_two_ap(tmp_path)
    rc = _run(["analyze", str(tmp_path)])
    assert rc == cli.EXIT_INPUT_ERROR


# ---------------------------------------------------------------------------
# 6. 入出力同一
# ---------------------------------------------------------------------------


def test_out_same_as_input_directory_is_error(tmp_path):
    _write_two_ap(tmp_path)
    rc = _run(["analyze", str(tmp_path), "--out", str(tmp_path)])
    assert rc == cli.EXIT_OUTPUT_ERROR


# ---------------------------------------------------------------------------
# 7. XLSX の体裁
# ---------------------------------------------------------------------------


def test_xlsx_layout(tmp_path):
    _write_two_ap(tmp_path)
    out = tmp_path / "out"
    assert _run(["analyze", str(tmp_path), "--out", str(out)]) == cli.EXIT_OK

    ws = load_workbook(next(out.glob("*.xlsx"))).active
    assert ws.cell(row=1, column=1).value  # タイトル
    assert "分析条件" in ws.cell(row=2, column=1).value
    assert "データ範囲" in ws.cell(row=3, column=1).value
    assert all(ws.cell(row=4, column=c).value is None for c in range(1, len(RESULT_COLUMNS) + 1))
    header = [ws.cell(row=5, column=c).value for c in range(1, len(RESULT_COLUMNS) + 1)]
    assert header == list(RESULT_COLUMNS)
    # 回復状況が「回復」の行に薄緑の背景色
    status_col = list(RESULT_COLUMNS).index("回復状況") + 1
    found_recovered_row = False
    for r in range(6, ws.max_row + 1):
        if ws.cell(row=r, column=status_col).value == "回復":
            found_recovered_row = True
            assert ws.cell(row=r, column=1).fill.start_color.rgb == "00C6EFCE"
    assert found_recovered_row


# ---------------------------------------------------------------------------
# 8. 警告の伝播
# ---------------------------------------------------------------------------


def test_insufficient_coverage_warning_propagates_to_stdout_and_files(tmp_path, capsys):
    _write_two_ap(tmp_path)
    out = tmp_path / "out"
    # データ終端よりずっと先を window_end にして「届いていません」警告を誘発する
    rc = _run([
        "analyze", str(tmp_path), "--out", str(out),
        "--to", "2026-01-02 00:00", "--format", "both",
    ])
    assert rc == cli.EXIT_OK

    stdout = capsys.readouterr().out
    assert "届いていません" in stdout

    summary_path = next(out.glob("*_summary.txt"))
    assert "届いていません" in summary_path.read_text(encoding="utf-8")

    xlsx_path = next(out.glob("*.xlsx"))
    ws = load_workbook(xlsx_path).active
    assert "届いていません" in ws.cell(row=3, column=1).value


# ---------------------------------------------------------------------------
# 9. 並び順
# ---------------------------------------------------------------------------


def test_result_order_is_ap_name_then_interval_number(tmp_path):
    _write_two_ap(tmp_path)
    out = tmp_path / "out"
    assert _run(["analyze", str(tmp_path), "--out", str(out)]) == cli.EXIT_OK

    csv_path = next(out.glob("*.csv"))
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    keys = [(r["ap_name"], int(r["区間番号"])) for r in rows]
    assert keys == [("AAA-AP", 1), ("BBB-AP", 1), ("BBB-AP", 2)]


# ---------------------------------------------------------------------------
# 10. glob とディレクトリ
# ---------------------------------------------------------------------------


def test_directory_input_merges_multiple_files(tmp_path):
    rows = _two_ap_rows()
    half = len(rows) // 2
    S.write_metrics(tmp_path / "part1.csv", rows[:half])
    S.write_metrics(tmp_path / "part2.csv", rows[half:])
    out = tmp_path / "out"

    assert _run(["analyze", str(tmp_path), "--out", str(out)]) == cli.EXIT_OK
    csv_path = next(out.glob("*.csv"))
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        data_rows = list(csv.reader(f))[1:]
    assert len(data_rows) == 3  # AAA-AP 1区間 + BBB-AP 2区間


def test_glob_input_merges_multiple_files(tmp_path):
    rows = _two_ap_rows()
    half = len(rows) // 2
    S.write_metrics(tmp_path / "part1.csv", rows[:half])
    S.write_metrics(tmp_path / "part2.csv", rows[half:])
    out = tmp_path / "out"

    assert _run(["analyze", str(tmp_path / "*.csv"), "--out", str(out)]) == cli.EXIT_OK
    csv_path = next(out.glob("*.csv"))
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        data_rows = list(csv.reader(f))[1:]
    assert len(data_rows) == 3
