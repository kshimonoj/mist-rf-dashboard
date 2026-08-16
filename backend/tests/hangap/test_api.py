"""ハングAP分析 API（/api/hangap）のテスト。

合成データのみを使う（実データ・実データ由来の値は fixtures に置かない）。

このテストは「API が CLI と同じ結果を返すこと」を要にしている。ロジックを
API 側で書き直すと真っ先にここが落ちる。
"""
from __future__ import annotations

import gc
import tempfile
import threading
import time
import types
from datetime import datetime, timedelta
from pathlib import Path

import _synth as S
import numpy as np
import pandas as pd
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from hangap import analysis, cli
from hangap.detector import RESULT_COLUMNS
from routers import hangap as api

START = datetime(2026, 1, 1, 10, 0, 0)
INTERVAL = 60

#: 「1〜3 サンプル → ゼロ 7 サンプル → 回復」。既定の min_zero_samples=5 を満たす。
HANG_PATTERN = [1, 1, 1] + [0] * 7 + [1, 1, 1]


# ---------------------------------------------------------------------------
# fixtures / helpers
# ---------------------------------------------------------------------------


@pytest.fixture
def api_client(tmp_path, monkeypatch):
    """LOGS_DIR / RESULTS_DIR を隔離したディレクトリに向けた TestClient を返す。"""
    logs_dir = tmp_path / "logs"
    logs_dir.mkdir()
    monkeypatch.setattr(api, "LOGS_DIR", str(logs_dir))
    # 保存先は logs_dir の外に置く（配下に置くと次の分析が自分の出力を読む）
    monkeypatch.setattr(api, "RESULTS_DIR", str(tmp_path / "hangap_results"))

    _clear_jobs()
    app = FastAPI()
    app.include_router(api.router)
    with TestClient(app) as client:
        yield client, logs_dir
    _clear_jobs()


def _clear_jobs() -> None:
    """プロセス内に残ったジョブを片付ける（_JOBS はモジュールグローバル）。"""
    with api._LOCK:
        for job in list(api._JOBS.values()):
            api._discard(job)
        api._JOBS.clear()


def _series(ap_id: str, ap_name: str, values: list[int], start: datetime = START) -> list[dict]:
    return [
        S.metrics_row(
            start + timedelta(seconds=INTERVAL * i),
            ap_id=ap_id, ap_name=ap_name, num_clients=v,
        )
        for i, v in enumerate(values)
    ]


def _write_hang_logs(logs_dir: Path, ap_count: int = 1) -> Path:
    rows: list[dict] = []
    for i in range(ap_count):
        rows += _series(f"test-ap-{i:04d}", f"TEST-AP-{i:02d}", HANG_PATTERN)
    return S.write_metrics(logs_dir / "ap_metrics_20260101_1000_TST.csv", rows)


def _wait_done(client: TestClient, job_id: str, timeout: float = 60.0) -> dict:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        state = client.get(f"/api/hangap/jobs/{job_id}").json()
        if state["status"] != api.STATUS_RUNNING:
            return state
        time.sleep(0.05)
    raise AssertionError(f"ジョブが {timeout}s 以内に終わりませんでした: {job_id}")


def _analyze(client: TestClient, body: dict | None = None) -> dict:
    """POST → 完了待ち。完了後の状態を返す。"""
    r = client.post("/api/hangap/analyze", json=body if body is not None else {})
    assert r.status_code == 202, r.text
    return _wait_done(client, r.json()["job_id"])


# ---------------------------------------------------------------------------
# 1. ジョブのライフサイクル
# ---------------------------------------------------------------------------


def test_job_lifecycle(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)

    r = client.post("/api/hangap/analyze", json={})
    assert r.status_code == 202
    started = r.json()
    assert started["status"] == api.STATUS_RUNNING
    assert started["job_id"]
    assert started["started_at"]

    state = _wait_done(client, started["job_id"])
    assert state["status"] == api.STATUS_DONE
    assert state["error"] is None
    assert state["finished_at"]
    assert state["phase"] in (
        analysis.PHASE_LOADING, analysis.PHASE_NEIGHBORS,
        analysis.PHASE_DETECTING, analysis.PHASE_WRITING,
    )
    assert state["summary"]["detected_intervals"] == 1

    result = client.get(f"/api/hangap/jobs/{started['job_id']}/result")
    assert result.status_code == 200
    assert result.json()["total"] == 1


def test_unknown_job_is_404(api_client):
    client, _ = api_client
    assert client.get("/api/hangap/jobs/does-not-exist").status_code == 404


# ---------------------------------------------------------------------------
# 2. 同時実行の拒否
# ---------------------------------------------------------------------------


def test_second_job_while_running_is_409(api_client, monkeypatch):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)

    gate = threading.Event()
    real_run = analysis.run_analysis

    def gated(*args, **kwargs):
        gate.wait(30)
        return real_run(*args, **kwargs)

    monkeypatch.setattr(api.analysis, "run_analysis", gated)

    first = client.post("/api/hangap/analyze", json={})
    assert first.status_code == 202
    job_id = first.json()["job_id"]
    try:
        second = client.post("/api/hangap/analyze", json={})
        assert second.status_code == 409
        detail = second.json()["detail"]
        assert detail["job_id"] == job_id
    finally:
        gate.set()
    assert _wait_done(client, job_id)["status"] == api.STATUS_DONE


def test_simultaneous_posts_start_only_one_job(api_client, monkeypatch):
    """同時に飛んできた POST でも 1 本しか走らないこと。

    実行中判定と登録が別のロックだと、どちらも「実行中なし」を見て 2 本走る。
    """
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)

    gate = threading.Event()
    real_run = analysis.run_analysis

    def gated(*args, **kwargs):
        gate.wait(30)
        return real_run(*args, **kwargs)

    monkeypatch.setattr(api.analysis, "run_analysis", gated)

    codes: list[int] = []
    codes_lock = threading.Lock()

    def post() -> None:
        r = client.post("/api/hangap/analyze", json={})
        with codes_lock:
            codes.append(r.status_code)

    threads = [threading.Thread(target=post) for _ in range(8)]
    for t in threads:
        t.start()
    for t in threads:
        t.join(30)

    try:
        assert codes.count(202) == 1, f"複数のジョブが開始された: {codes}"
        assert codes.count(409) == 7
        with api._LOCK:
            assert len(api._JOBS) == 1
            job_id = next(iter(api._JOBS))
    finally:
        gate.set()
    _wait_done(client, job_id)


# ---------------------------------------------------------------------------
# 3. 不正なパラメータ
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("body,field_name", [
    ({"from": "2026/01/01 10:00 JST"}, "from"),
    ({"from": "2026-01-01T10:00:00+09:00"}, "from"),
    ({"to": "not-a-time"}, "to"),
    ({"min_zero_samples": -5}, "min_zero_samples"),
    ({"min_zero_samples": 1.5}, "min_zero_samples"),
    ({"gap_factor": -1}, "gap_factor"),
    ({"max_distance_m": -25}, "max_distance_m"),
    ({"neighbor_count": -1}, "neighbor_count"),
    ({"neighbor_client_threshold": -1}, "neighbor_client_threshold"),
    ({"truncated_warn_ratio": 1.5}, "truncated_warn_ratio"),
    ({"min_zero_duration": "しばらく"}, "min_zero_duration"),
    ({"event_window_minutes": "abc"}, "event_window_minutes"),
    ({"unknown_field": 1}, "unknown_field"),
])
def test_invalid_parameters_are_400_and_name_the_field(api_client, body, field_name):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)

    r = client.post("/api/hangap/analyze", json=body)
    assert r.status_code == 400, r.text
    assert field_name in r.json()["detail"]


def test_from_after_to_is_400(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)
    r = client.post(
        "/api/hangap/analyze",
        json={"from": "2026-01-01 12:00", "to": "2026-01-01 10:00"},
    )
    assert r.status_code == 400
    assert "to" in r.json()["detail"]


# ---------------------------------------------------------------------------
# 4. 既定値の一致（ボディ空 = CLI の既定値）
# ---------------------------------------------------------------------------


def test_empty_body_uses_cli_defaults(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)

    cli_args = cli.build_parser().parse_args(["analyze", "dummy", "--out", "dummy_out"])
    defaults = analysis.AnalysisParams()

    assert defaults.min_zero_samples == cli_args.min_zero_samples
    assert defaults.min_zero_duration is None and cli_args.min_zero_duration is None
    assert defaults.event_window == analysis.parse_duration(cli_args.event_window, "--event-window")
    assert defaults.exodus_threshold == cli_args.exodus_threshold
    assert defaults.gap_factor == cli_args.gap_factor
    assert defaults.neighbor_count == cli_args.neighbor_count
    assert defaults.max_distance_m == cli_args.max_distance_m
    assert defaults.neighbor_client_threshold == cli_args.neighbor_client_threshold
    assert defaults.truncated_warn_ratio == cli_args.truncated_warn_ratio

    state = _analyze(client)
    assert state["status"] == api.STATUS_DONE
    # 実際にジョブが使った条件が、CLI の既定値そのものであること
    assert state["summary"]["condition_text"] == analysis.condition_text(defaults, n_files=1)


# ---------------------------------------------------------------------------
# 5. メトリクス 0 件は failed（6 の「検出 0 件」と区別する）
# ---------------------------------------------------------------------------


def test_no_ap_metrics_is_failed_with_reason(api_client):
    client, logs_dir = api_client
    # ap_events はあるが ap_metrics が 1 行も無い
    S.write_events(logs_dir / "ap_events_20260101_1000_TST.csv", [S.event_row(START)])

    state = _analyze(client)
    assert state["status"] == api.STATUS_FAILED
    assert state["error"]
    assert "ap_metrics" in state["error"]
    assert state["summary"] is None


def test_empty_logs_dir_is_failed(api_client):
    client, _ = api_client
    state = _analyze(client)
    assert state["status"] == api.STATUS_FAILED
    assert "ap_metrics" in state["error"]


def test_failed_job_result_is_409(api_client):
    client, logs_dir = api_client
    S.write_events(logs_dir / "ap_events_20260101_1000_TST.csv", [S.event_row(START)])
    r = client.post("/api/hangap/analyze", json={})
    job_id = r.json()["job_id"]
    _wait_done(client, job_id)
    assert client.get(f"/api/hangap/jobs/{job_id}/result").status_code == 409


# ---------------------------------------------------------------------------
# 6. 検出 0 件は done（5 との区別）
# ---------------------------------------------------------------------------


def test_zero_detections_is_done_not_failed(api_client):
    client, logs_dir = api_client
    S.write_metrics(
        logs_dir / "ap_metrics_20260101_1000_TST.csv",
        _series("test-ap-0000", "TEST-AP-00", [1] * 20),
    )

    state = _analyze(client)
    assert state["status"] == api.STATUS_DONE
    assert state["error"] is None
    assert state["summary"]["detected_intervals"] == 0
    assert state["summary"]["loader"]["metrics_rows"] == 20

    body = client.get(f"/api/hangap/jobs/{_job_id(state)}/result").json()
    assert body["total"] == 0
    assert body["rows"] == []
    assert body["columns"] == list(RESULT_COLUMNS)


def _job_id(state: dict) -> str:
    return state["job_id"]


# ---------------------------------------------------------------------------
# 7. ページング
# ---------------------------------------------------------------------------


def test_result_paging(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir, ap_count=5)
    state = _analyze(client)
    job_id = _job_id(state)
    assert state["summary"]["detected_intervals"] == 5

    first = client.get(f"/api/hangap/jobs/{job_id}/result?offset=0&limit=2").json()
    assert first["total"] == 5
    assert first["offset"] == 0 and first["limit"] == 2
    assert len(first["rows"]) == 2

    middle = client.get(f"/api/hangap/jobs/{job_id}/result?offset=2&limit=2").json()
    assert [r["ap_name"] for r in middle["rows"]] == ["TEST-AP-02", "TEST-AP-03"]

    tail = client.get(f"/api/hangap/jobs/{job_id}/result?offset=4&limit=100").json()
    assert len(tail["rows"]) == 1
    assert tail["total"] == 5

    past_end = client.get(f"/api/hangap/jobs/{job_id}/result?offset=99&limit=10").json()
    assert past_end["rows"] == [] and past_end["total"] == 5


def test_limit_bounds(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)
    job_id = _job_id(_analyze(client))

    assert client.get(f"/api/hangap/jobs/{job_id}/result").json()["limit"] == api.DEFAULT_RESULT_LIMIT
    ok = client.get(f"/api/hangap/jobs/{job_id}/result?limit={api.MAX_RESULT_LIMIT}")
    assert ok.status_code == 200
    too_big = client.get(f"/api/hangap/jobs/{job_id}/result?limit={api.MAX_RESULT_LIMIT + 1}")
    assert too_big.status_code == 422  # FastAPI の Query 制約
    assert client.get(f"/api/hangap/jobs/{job_id}/result?offset=-1").status_code == 422


# ---------------------------------------------------------------------------
# 8. フィルタとソート
# ---------------------------------------------------------------------------


def test_status_filter_and_sort(api_client):
    client, logs_dir = api_client
    rows = []
    # 3 台は回復、1 台はゼロのまま（継続中）
    for i in range(3):
        rows += _series(f"test-ap-{i:04d}", f"TEST-AP-{i:02d}", HANG_PATTERN)
    rows += _series("test-ap-0009", "TEST-AP-09", [1, 1, 1] + [0] * 10)
    S.write_metrics(logs_dir / "ap_metrics_20260101_1000_TST.csv", rows)

    job_id = _job_id(_analyze(client))
    all_rows = client.get(f"/api/hangap/jobs/{job_id}/result?limit=1000").json()
    assert all_rows["total"] == 4

    recovered = client.get(f"/api/hangap/jobs/{job_id}/result?status=回復").json()
    assert recovered["total"] == 3
    assert {r["回復状況"] for r in recovered["rows"]} == {"回復"}

    ongoing = client.get(f"/api/hangap/jobs/{job_id}/result?status=継続中").json()
    assert ongoing["total"] == 1
    assert ongoing["rows"][0]["ap_name"] == "TEST-AP-09"

    desc = client.get(f"/api/hangap/jobs/{job_id}/result?sort=ap_name&order=desc").json()
    assert [r["ap_name"] for r in desc["rows"]] == [
        "TEST-AP-09", "TEST-AP-02", "TEST-AP-01", "TEST-AP-00",
    ]
    asc = client.get(f"/api/hangap/jobs/{job_id}/result?sort=ゼロ開始&order=asc").json()
    starts = [r["ゼロ開始"] for r in asc["rows"]]
    assert starts == sorted(starts)

    # フィルタとソートの併用
    both = client.get(
        f"/api/hangap/jobs/{job_id}/result?status=回復&sort=ap_name&order=desc"
    ).json()
    assert [r["ap_name"] for r in both["rows"]] == ["TEST-AP-02", "TEST-AP-01", "TEST-AP-00"]


def test_invalid_filter_and_sort_are_400(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)
    job_id = _job_id(_analyze(client))

    bad_status = client.get(f"/api/hangap/jobs/{job_id}/result?status=そんな状態")
    assert bad_status.status_code == 400 and "status" in bad_status.json()["detail"]

    bad_sort = client.get(f"/api/hangap/jobs/{job_id}/result?sort=no_such_column")
    assert bad_sort.status_code == 400 and "sort" in bad_sort.json()["detail"]

    bad_order = client.get(f"/api/hangap/jobs/{job_id}/result?sort=ap_name&order=sideways")
    assert bad_order.status_code == 400 and "order" in bad_order.json()["detail"]


# ---------------------------------------------------------------------------
# 9. 列の一致
# ---------------------------------------------------------------------------


def test_result_columns_match_detector_exactly(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)
    body = client.get(f"/api/hangap/jobs/{_job_id(_analyze(client))}/result").json()

    assert body["columns"] == list(RESULT_COLUMNS)
    assert list(body["rows"][0].keys()) == list(RESULT_COLUMNS)


# ---------------------------------------------------------------------------
# 10. ダウンロード（CLI の出力と同一内容）
# ---------------------------------------------------------------------------


def _run_cli(logs_dir: Path, out_dir: Path) -> None:
    assert cli.main(["analyze", str(logs_dir), "--out", str(out_dir), "--format", "both"]) == cli.EXIT_OK


def test_download_csv_matches_cli_output(api_client, tmp_path):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir, ap_count=3)
    job_id = _job_id(_analyze(client))

    out_dir = tmp_path / "cli_out"
    _run_cli(logs_dir, out_dir)
    cli_csv = next(out_dir.glob("*.csv")).read_bytes()

    r = client.get(f"/api/hangap/jobs/{job_id}/download?format=csv")
    assert r.status_code == 200
    assert r.content == cli_csv


def test_download_xlsx_matches_cli_output(api_client, tmp_path):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir, ap_count=3)
    job_id = _job_id(_analyze(client))

    out_dir = tmp_path / "cli_out"
    _run_cli(logs_dir, out_dir)

    r = client.get(f"/api/hangap/jobs/{job_id}/download?format=xlsx")
    assert r.status_code == 200
    api_xlsx = tmp_path / "api_result.xlsx"
    api_xlsx.write_bytes(r.content)

    expected = load_workbook(next(out_dir.glob("*.xlsx"))).active
    actual = load_workbook(api_xlsx).active

    assert actual.title == expected.title
    assert (actual.max_row, actual.max_column) == (expected.max_row, expected.max_column)
    for row in range(1, expected.max_row + 1):
        for col in range(1, expected.max_column + 1):
            e = expected.cell(row=row, column=col)
            a = actual.cell(row=row, column=col)
            assert a.value == e.value, f"cell ({row},{col}) が一致しません"
            # 条件行・警告行・回復行の背景色も CLI と同じであること
            assert a.fill.start_color.rgb == e.fill.start_color.rgb, f"fill ({row},{col})"
            assert a.font.bold == e.font.bold, f"font ({row},{col})"


def test_download_bad_format_is_400(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)
    job_id = _job_id(_analyze(client))
    r = client.get(f"/api/hangap/jobs/{job_id}/download?format=pdf")
    assert r.status_code == 400 and "format" in r.json()["detail"]


# ---------------------------------------------------------------------------
# 11. ジョブの破棄
# ---------------------------------------------------------------------------


def test_delete_removes_result_and_temp_files(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)
    job_id = _job_id(_analyze(client))

    tmp_root = Path(api._JOBS[job_id].tmpdir.name)
    assert tmp_root.is_dir()
    assert list(tmp_root.iterdir())

    r = client.delete(f"/api/hangap/jobs/{job_id}")
    assert r.status_code == 200 and r.json()["deleted"] is True

    assert not tmp_root.exists()
    assert client.get(f"/api/hangap/jobs/{job_id}").status_code == 404
    assert client.get(f"/api/hangap/jobs/{job_id}/result").status_code == 404
    assert client.delete(f"/api/hangap/jobs/{job_id}").status_code == 404


def test_delete_while_running_keeps_the_slot_until_the_thread_finishes(api_client, monkeypatch):
    """実行中の DELETE で枠が空くと、読み込みが走ったまま次のジョブが始まってしまう。"""
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)

    gate = threading.Event()
    real_run = analysis.run_analysis

    def gated(*args, **kwargs):
        gate.wait(30)
        return real_run(*args, **kwargs)

    monkeypatch.setattr(api.analysis, "run_analysis", gated)

    job_id = client.post("/api/hangap/analyze", json={}).json()["job_id"]
    try:
        assert client.delete(f"/api/hangap/jobs/{job_id}").status_code == 200
        # 利用者からは消えているが、枠はまだ空かない
        assert client.get(f"/api/hangap/jobs/{job_id}").status_code == 404
        assert client.post("/api/hangap/analyze", json={}).status_code == 409
    finally:
        gate.set()

    deadline = time.monotonic() + 30
    while time.monotonic() < deadline:
        with api._LOCK:
            if job_id not in api._JOBS:
                break
        time.sleep(0.05)
    with api._LOCK:
        assert job_id not in api._JOBS, "スレッド終了後もレジストリに残っている"

    # スレッドが終われば次のジョブを開始できる
    assert _analyze(client)["status"] == api.STATUS_DONE


def test_oldest_job_is_discarded_beyond_max_jobs(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)

    job_ids = []
    for _ in range(api.MAX_JOBS + 1):
        job_ids.append(_job_id(_analyze(client)))

    assert client.get(f"/api/hangap/jobs/{job_ids[0]}").status_code == 404
    for job_id in job_ids[1:]:
        assert client.get(f"/api/hangap/jobs/{job_id}").status_code == 200


def test_finished_job_is_purged_after_ttl(api_client, monkeypatch):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)
    job_id = _job_id(_analyze(client))

    tmp_root = Path(api._JOBS[job_id].tmpdir.name)
    monkeypatch.setattr(api, "JOB_TTL_SECONDS", -1)  # 完了直後を「1時間経過」とみなす

    assert client.get(f"/api/hangap/jobs/{job_id}").status_code == 404
    assert not tmp_root.exists()


# ---------------------------------------------------------------------------
# 12. 一時ファイルの場所（data/ 配下に書かないこと）
# ---------------------------------------------------------------------------


def test_outputs_are_written_outside_the_data_directory(api_client, tmp_path):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)
    job_id = _job_id(_analyze(client))

    outputs = api._JOBS[job_id].outputs
    assert set(outputs) == {"xlsx", "csv", "summary"}
    system_tmp = Path(tempfile.gettempdir()).resolve()
    for path in outputs.values():
        resolved = Path(path).resolve()
        assert resolved.is_file()
        assert resolved.is_relative_to(system_tmp), f"一時ディレクトリの外に出力された: {resolved}"
        assert not resolved.is_relative_to(tmp_path.resolve())

    # data/logs に出力が混ざっていないこと（次回の分析が自分の出力を読んでしまう）
    assert sorted(p.name for p in logs_dir.iterdir()) == ["ap_metrics_20260101_1000_TST.csv"]
    # data 配下に出るのは保存済み結果（hangap_results）だけ。logs_dir 配下には出ない
    written = [p for p in tmp_path.rglob("hangap_result_*") if p.is_file()]
    assert written and {p.parent for p in written} == {Path(api.RESULTS_DIR)}


# ---------------------------------------------------------------------------
# 追加: 分析条件が実際に効いていること（既定値を素通りさせていない）
# ---------------------------------------------------------------------------


def test_parameters_are_actually_applied(api_client):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)

    strict = _analyze(client, {"min_zero_samples": 1000})
    assert strict["summary"]["detected_intervals"] == 0

    by_duration = _analyze(client, {"min_zero_samples": 1000, "min_zero_duration": "4m"})
    assert by_duration["summary"]["detected_intervals"] == 1
    assert "min_zero_duration=4m" in by_duration["summary"]["condition_text"]

    windowed = _analyze(client, {"from": "2026-01-01 09:00", "to": "2026-01-01 10:02"})
    assert windowed["summary"]["detected_intervals"] == 0


# ---------------------------------------------------------------------------
# 13. 最大実行時間（ワーカーがハングしても枠を解放する）
# ---------------------------------------------------------------------------


def test_timed_out_job_fails_and_frees_the_slot(api_client, monkeypatch):
    """最大実行時間を超えたジョブは failed になり、次のジョブを開始できること。

    枠が空かないと「ボタンを押しても永久に何も起きない」になり、復旧手段が
    コンテナ再起動しか無くなる。
    """
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)

    gate = threading.Event()
    real_run = analysis.run_analysis

    def gated(*args, **kwargs):
        gate.wait(30)
        return real_run(*args, **kwargs)

    monkeypatch.setattr(api.analysis, "run_analysis", gated)

    hung_id = client.post("/api/hangap/analyze", json={}).json()["job_id"]
    try:
        monkeypatch.setattr(api, "MAX_RUN_SECONDS", -1)  # 開始直後を「超過」とみなす
        state = client.get(f"/api/hangap/jobs/{hung_id}").json()
        assert state["status"] == api.STATUS_FAILED
        assert state["finished_at"]
        # 打ち切ったジョブは結果を返さない
        assert client.get(f"/api/hangap/jobs/{hung_id}/result").status_code == 409

        monkeypatch.setattr(api, "MAX_RUN_SECONDS", 600)  # 以降は通常どおり
        second = client.post("/api/hangap/analyze", json={})
        assert second.status_code == 202, "枠が解放されていない"
        second_id = second.json()["job_id"]
    finally:
        gate.set()

    assert _wait_done(client, second_id)["status"] == api.STATUS_DONE

    # 打ち切ったジョブのワーカーが後から結果を書き戻さないこと
    deadline = time.monotonic() + 30
    while time.monotonic() < deadline:
        with api._LOCK:
            if api._JOBS[hung_id].files == []:  # finally まで到達した
                break
        time.sleep(0.05)
    after = client.get(f"/api/hangap/jobs/{hung_id}").json()
    assert after["status"] == api.STATUS_FAILED
    assert after["summary"] is None
    with api._LOCK:
        assert api._JOBS[hung_id].result is None
        assert api._JOBS[hung_id].tmpdir is None


def test_timed_out_job_error_explains_the_reason(api_client, monkeypatch):
    client, logs_dir = api_client
    _write_hang_logs(logs_dir)

    gate = threading.Event()
    real_run = analysis.run_analysis

    def gated(*args, **kwargs):
        gate.wait(30)
        return real_run(*args, **kwargs)

    monkeypatch.setattr(api.analysis, "run_analysis", gated)

    job_id = client.post("/api/hangap/analyze", json={}).json()["job_id"]
    try:
        monkeypatch.setattr(api, "MAX_RUN_SECONDS", -1)
        state = client.get(f"/api/hangap/jobs/{job_id}").json()
        assert state["status"] == api.STATUS_FAILED
        assert "最大実行時間" in state["error"]
        # 「検出0件」と読み違えないこと
        assert "0 件" in state["error"]
    finally:
        gate.set()


# ---------------------------------------------------------------------------
# 14. 完了時に入力データを解放する
# ---------------------------------------------------------------------------


#: 入力が結果よりはるかに大きいログ。1 AP あたり 313 行の入力から 1 行の結果が出る。
_PADDED_HANG_PATTERN = [1] * 300 + HANG_PATTERN


def _write_padded_hang_logs(logs_dir: Path, ap_count: int) -> Path:
    rows: list[dict] = []
    for i in range(ap_count):
        rows += _series(f"test-ap-{i:04d}", f"TEST-AP-{i:02d}", _PADDED_HANG_PATTERN)
    return S.write_metrics(logs_dir / "ap_metrics_20260101_1000_TST.csv", rows)


def _reachable_data(root: object) -> tuple[list[pd.DataFrame], int]:
    """``root`` から辿れる DataFrame と、データの合計バイト数を返す。

    型・モジュール・関数は辿らない（辿るとプロセス全体に届いてしまい、何も測れない）。
    DataFrame は memory_usage で数え、中身の配列は二重に数えない。
    """
    skip = (
        type, types.ModuleType, types.FunctionType, types.MethodType,
        types.BuiltinFunctionType, types.FrameType,
    )
    seen: set[int] = set()
    frames: list[pd.DataFrame] = []
    total = 0
    stack: list[object] = [root]
    while stack:
        obj = stack.pop()
        if id(obj) in seen or isinstance(obj, skip):
            continue
        seen.add(id(obj))
        if isinstance(obj, pd.DataFrame):
            frames.append(obj)
            total += int(obj.memory_usage(deep=True).sum())
            continue
        if isinstance(obj, pd.Series):
            total += int(obj.memory_usage(deep=True))
            continue
        if isinstance(obj, np.ndarray):
            total += int(obj.nbytes)
            continue
        stack.extend(gc.get_referents(obj))
    return frames, total


def test_finished_jobs_do_not_retain_the_input(api_client):
    """完了ジョブが入力（実測で 516,859 行規模）を掴んだままにしないこと。

    結果は数百行しかないので、保持ジョブ 3 件分のメモリは結果行数に比例する程度で
    収まっていなければならない。入力を持ち続けると、保持ジョブ 3 件で入力 3 本分の
    メモリを抱えることになる。
    """
    client, logs_dir = api_client
    ap_count = 20
    files = [_write_padded_hang_logs(logs_dir, ap_count)]

    for _ in range(api.MAX_JOBS):
        assert _analyze(client)["status"] == api.STATUS_DONE

    gc.collect()
    with api._LOCK:
        assert len(api._JOBS) == api.MAX_JOBS
        frames, retained_bytes = _reachable_data(api._JOBS)

    # 到達できる DataFrame は各ジョブの結果表だけ（入力は 1 AP あたり 313 行ある）
    assert len(frames) == api.MAX_JOBS
    assert [len(df) for df in frames] == [ap_count] * api.MAX_JOBS

    # 保持 3 件の合計が、入力 1 本分の DataFrame よりも小さいこと
    loaded = analysis.loader.load(files)
    input_bytes = int(loaded.metrics.memory_usage(deep=True).sum())
    assert len(loaded.metrics) == ap_count * len(_PADDED_HANG_PATTERN)
    assert retained_bytes < input_bytes, (
        f"保持ジョブ {api.MAX_JOBS} 件が {retained_bytes} バイトを保持している"
        f"（入力 1 本 = {input_bytes} バイト）"
    )


def test_result_and_download_work_after_the_input_is_released(api_client):
    client, logs_dir = api_client
    _write_padded_hang_logs(logs_dir, ap_count=3)
    job_id = _job_id(_analyze(client))

    gc.collect()

    result = client.get(f"/api/hangap/jobs/{job_id}/result")
    assert result.status_code == 200
    body = result.json()
    assert body["total"] == 3
    assert len(body["rows"]) == 3
    assert list(body["columns"]) == list(RESULT_COLUMNS)

    for fmt in ("xlsx", "csv"):
        r = client.get(f"/api/hangap/jobs/{job_id}/download?format={fmt}")
        assert r.status_code == 200, fmt
        assert r.content
