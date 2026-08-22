"""rrm analyze CLI の配線テスト。合成データのみを使う。"""
from __future__ import annotations

import csv
from datetime import datetime, timedelta
from pathlib import Path

import _rrmsynth as S
import pytest
from openpyxl import load_workbook

from rrm import analysis, cli

START = datetime(2026, 1, 1, 10, 0, 0)


def _samples(n: int) -> list[dict[str, object]]:
    return [{"num_clients": 5, "util_24": 10, "util_5": 20, "util_6": 30} for _ in range(n)]


def write_logs(logs_dir: Path) -> Path:
    logs_dir.mkdir(parents=True, exist_ok=True)
    rows = (
        S.series(START, _samples(24), ap=S.AP1)
        + S.series(START, _samples(24), ap=S.AP2)
        + S.series(START, _samples(24), ap=S.AP3,
                   site_id=S.OTHER_SITE_ID, site_name=S.OTHER_SITE_NAME)
    )
    S.write_metrics(logs_dir / "ap_metrics_20260101_1000_TZT.csv", rows)
    S.write_events(logs_dir / "ap_events_20260101_1000_TZT.csv", [
        S.rrm_action(START + timedelta(minutes=7), pre_channel=36, channel=44, ap=S.AP1),
        S.rrm_action(START + timedelta(minutes=17), pre_channel=44, channel=44, ap=S.AP1),
        S.radar_detected(START + timedelta(minutes=27), pre_channel=64, channel=36, ap=S.AP2),
        S.rrm_action(START + timedelta(minutes=27, seconds=2), reason="radar-detected",
                     pre_channel=64, channel=36, ap=S.AP2),
        S.rrm_action(START + timedelta(minutes=37), reason="post-radar",
                     pre_channel=36, channel=40, ap=S.AP2),
        S.rrm_action(START + timedelta(minutes=47), pre_channel=36, channel=44, ap=S.AP3,
                     site_name=S.OTHER_SITE_NAME),
        S.config_changed_by_rrm(START + timedelta(minutes=7), ap=S.AP1),
    ])
    return logs_dir


def _run(argv: list[str]) -> int:
    return cli.main(argv)


def test_analyze_writes_all_three_files(tmp_path, capsys):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run(["analyze", "--logs", str(logs), "--out", str(out)]) == cli.EXIT_OK

    assert len(list(out.glob("*.xlsx"))) == 1
    assert len(list(out.glob("*.csv"))) == 1
    assert len(list(out.glob("*_summary.txt"))) == 1

    printed = capsys.readouterr().out
    assert "AP_RADAR_DETECTED" in printed
    assert "AP_CONFIG_CHANGED_BY_RRM" in printed
    assert "POST_RADAR" in printed


def test_result_csv_columns_and_rows(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run(["analyze", "--logs", str(logs), "--out", str(out)]) == cli.EXIT_OK

    with open(next(out.glob("*.csv")), newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    assert list(rows[0].keys()) == list(analysis.RESULT_COLUMNS)
    assert len(rows) == 5  # AP_RRM_ACTION のみ（RADAR / CONFIG_CHANGED は明細に入れない）
    # 値がすべてゼロでも列は落とさない
    assert all("clients_before" in r for r in rows)


def test_site_option_can_be_repeated(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run([
        "analyze", "--logs", str(logs), "--out", str(out),
        "--site", S.SITE_ID, "--site", S.OTHER_SITE_ID,
    ]) == cli.EXIT_OK
    with open(next(out.glob("*.csv")), newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert {r["site_name"] for r in rows} == {S.SITE_NAME, S.OTHER_SITE_NAME}


def test_single_site_filters_events_by_site_name(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run([
        "analyze", "--logs", str(logs), "--out", str(out), "--site", S.OTHER_SITE_NAME,
    ]) == cli.EXIT_OK
    with open(next(out.glob("*.csv")), newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert {r["site_name"] for r in rows} == {S.OTHER_SITE_NAME}


def test_unknown_site_is_an_input_error(tmp_path, capsys):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run([
        "analyze", "--logs", str(logs), "--out", str(out), "--site", "no-such-site",
    ]) == cli.EXIT_INPUT_ERROR
    assert "見つかりません" in capsys.readouterr().err


def test_to_before_from_is_an_input_error(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run([
        "analyze", "--logs", str(logs), "--out", str(out),
        "--from", "2026-01-01 11:00", "--to", "2026-01-01 10:00",
    ]) == cli.EXIT_INPUT_ERROR


def test_out_equal_to_logs_is_an_output_error(tmp_path):
    logs = write_logs(tmp_path / "logs")
    assert _run(["analyze", "--logs", str(logs), "--out", str(logs)]) == cli.EXIT_OUTPUT_ERROR


def test_xlsx_has_the_three_sheets(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run(["analyze", "--logs", str(logs), "--out", str(out)]) == cli.EXIT_OK
    wb = load_workbook(next(out.glob("*.xlsx")))
    assert wb.sheetnames == ["chart", "data", "summary"]
    assert len(wb["chart"]._charts) >= 1


def test_missing_logs_dir_is_an_input_error(tmp_path):
    assert _run([
        "analyze", "--logs", str(tmp_path / "nope"), "--out", str(tmp_path / "out"),
    ]) == cli.EXIT_INPUT_ERROR
