"""xlsx グラフ（chart シート）の描画属性テスト。合成データのみを使う。

27 番で確定した設定（軸を消さない / カテゴリ軸 maxMin・値軸 minMax / 軸タイトル
なし / データラベルは値のみ / gapWidth=40 / 高さ可変・幅 17.5cm / セルの塗りは
8 桁 ARGB）をそのまま固定する。

openpyxl は保存時、``chart.height`` / ``chart.width`` (cm) を描画アンカー
（``drawing*.xml`` の ``<xdr:ext>``）の EMU 値へ正しく書き出す。しかし
``load_workbook()`` で読み戻した ``chart.height`` / ``chart.width`` は
チャート本体の XML に物理サイズの情報が無いため常にクラス既定値を返す ――
実際に保存された値では **ない**。そのため本数依存の検証だけは
``chart.anchor.ext``（drawing 側の EMU）を直接読む。
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import _rrmsynth as S
from openpyxl import load_workbook

from rrm import cli
from rrm.events import CLASSIFICATIONS, CLASS_COLORS

START = datetime(2026, 1, 1, 0, 0, 0)
EMU_PER_CM = 360000


def _write_logs(logs_dir: Path, hours: int) -> Path:
    """``hours`` 時間ぶん、毎時 1 件のチャネル変更があるログを作る。"""
    logs_dir.mkdir(parents=True, exist_ok=True)
    samples = [{"num_clients": 5, "util_24": 10, "util_5": 20, "util_6": 30}] * (hours * 12 + 2)
    S.write_metrics(
        logs_dir / "ap_metrics_20260101_0000_TZT.csv",
        S.series(START, samples, ap=S.AP1),
    )
    reasons = ("scheduled-site-rrm", "radar-detected", "post-radar")
    S.write_events(logs_dir / "ap_events_20260101_0000_TZT.csv", [
        S.rrm_action(
            START + timedelta(hours=h, minutes=7),
            reason=reasons[h % len(reasons)],
            pre_channel=36, channel=44, ap=S.AP1,
        )
        for h in range(hours)
    ])
    return logs_dir


def _build(tmp_path: Path, hours: int):
    logs = _write_logs(tmp_path / f"logs{hours}", hours)
    out = tmp_path / f"out{hours}"
    assert cli.main(["analyze", "--logs", str(logs), "--out", str(out)]) == cli.EXIT_OK
    wb = load_workbook(next(out.glob("*.xlsx")))
    return wb


def _charts(tmp_path: Path, hours: int = 4):
    return _build(tmp_path, hours)["chart"]._charts


def test_chart_sheet_exists_with_two_charts(tmp_path):
    wb = _build(tmp_path, 4)
    assert wb.sheetnames == ["chart", "data", "summary"]
    assert len(wb["chart"]._charts) == 2


def test_axes_are_not_deleted(tmp_path):
    for chart in _charts(tmp_path):
        assert chart.x_axis.delete is False
        assert chart.y_axis.delete is False


def test_orientation_puts_the_first_category_on_top(tmp_path):
    for chart in _charts(tmp_path):
        # 横棒（type="bar"）ではカテゴリ軸が x_axis、値軸が y_axis
        assert chart.type == "bar"
        assert chart.x_axis.scaling.orientation == "maxMin"
        assert chart.y_axis.scaling.orientation == "minMax"


def test_axis_titles_are_removed(tmp_path):
    for chart in _charts(tmp_path):
        assert chart.x_axis.title is None
        assert chart.y_axis.title is None


def test_data_labels_show_value_only(tmp_path):
    for chart in _charts(tmp_path):
        assert chart.dLbls.showSerName is False
        assert chart.dLbls.showCatName is False
        assert chart.dLbls.showLegendKey is False
        assert chart.dLbls.showVal is True


def test_gap_width_is_40(tmp_path):
    for chart in _charts(tmp_path):
        assert chart.gapWidth == 40


def test_hourly_chart_is_stacked_with_one_series_per_classification(tmp_path):
    chart = _charts(tmp_path)[0]
    assert chart.grouping == "stacked"
    assert chart.overlap == 100
    assert len(chart.series) == len(CLASSIFICATIONS)
    # 読み戻すと ColorChoice になるので srgbClr を見る
    colors = [s.graphicalProperties.solidFill.srgbClr for s in chart.series]
    assert colors == [CLASS_COLORS[name] for name in CLASSIFICATIONS]


def test_chart_width_is_17_5cm_and_height_scales_with_bucket_count(tmp_path):
    chart4 = _build(tmp_path, 4)["chart"]._charts[0]
    chart12 = _build(tmp_path, 12)["chart"]._charts[0]

    assert round(chart4.anchor.ext.cx / EMU_PER_CM, 1) == 17.5
    height4 = chart4.anchor.ext.cy / EMU_PER_CM
    height12 = chart12.anchor.ext.cy / EMU_PER_CM
    assert height12 > height4


def test_legend_swatches_use_8_digit_argb(tmp_path):
    """6 桁 RGB を渡すと alpha が 00（透明）になって塗りが見えなくなる。"""
    ws = _build(tmp_path, 4)["chart"]
    for i, name in enumerate(CLASSIFICATIONS):
        cell = ws.cell(row=6 + i, column=1)
        assert ws.cell(row=6 + i, column=2).value == name
        assert cell.fill.start_color.rgb == f"FF{CLASS_COLORS[name]}"


def test_chart_is_capped_and_the_cap_is_stated(tmp_path):
    """バケットが多いときはグラフを直近に絞り、**絞ったことを必ず書く**。"""
    from rrm.analysis import CHART_MAX_BUCKETS

    ws = _build(tmp_path, CHART_MAX_BUCKETS + 6)["chart"]
    note = ws.cell(row=4, column=1).value
    assert f"直近 {CHART_MAX_BUCKETS} バケット" in note
    assert f"全 {CHART_MAX_BUCKETS + 6} バケット" in note
