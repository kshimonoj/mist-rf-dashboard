"""復元（再識別）のテスト。指示 24 の要件 1〜11 を固定する。合成データのみ。

要件 1（往復してバイト単位で一致）が核。ここが崩れたら「戻ったつもりで
戻っていない」ので、他がいくら通っても意味がない。

なお仮名化は入力をいくつか正規形に寄せる（csv の改行は CRLF、MAC はコロンなし
小文字）。往復のバイト一致はその正規形どうしの比較として見る必要があるため、
入力フィクスチャは csv モジュールで書き直してから使う。
"""
from __future__ import annotations

import csv
import os
import shutil
from pathlib import Path

import pytest

from conftest import ALL_FIXTURES, FIXTURES_DIR
from pseudonymizer import cli
from pseudonymizer import restore as restore_mod
from pseudonymizer.restore import RestoreError, UnsupportedFormatError, load_engine
from pseudonymizer.restore_cli import main as restore_main
from pseudonymizer.salt import DEFAULT_MAP_FILENAME, DEFAULT_SALT_FILENAME

# vlan_id は仮名が裸の整数（str(idx)）でテキスト置換では戻せない。
# 往復のテストは --keep-vlan を使う（この制約は README にも書いてある）。
PSEUDONYMIZE_ARGS = ("--keep-vlan",)


def canonical_copy(dst_dir: Path, names=ALL_FIXTURES) -> Path:
    """フィクスチャを csv モジュールで書き直して置く（改行を CRLF に揃える）。"""
    dst_dir.mkdir(parents=True, exist_ok=True)
    for name in names:
        with open(Path(FIXTURES_DIR) / name, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        with open(dst_dir / name, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)
    return dst_dir


@pytest.fixture
def roundtrip(tmp_path):
    """(元ディレクトリ, 仮名化後ディレクトリ, ソルトパス) を返す。"""
    src = canonical_copy(tmp_path / "src")
    pseudo = tmp_path / "pseudo"
    assert cli.main([str(src), "--out", str(pseudo), *PSEUDONYMIZE_ARGS]) == 0
    return src, pseudo, str(pseudo / DEFAULT_SALT_FILENAME)


def run_restore(inputs, out_dir, salt_path, *extra) -> int:
    return restore_main([*(str(i) for i in inputs), "--out", str(out_dir),
                         "--salt-file", str(salt_path), *extra])


def restored_for(out_dir: Path, source_name: str) -> Path:
    """入力名に対応する復元後ファイル（ファイル名の日付は戻るので前方一致で探す）。"""
    prefix = source_name.split("_2024")[0]
    matches = sorted(p for p in out_dir.glob("*") if p.name.startswith(prefix))
    assert len(matches) == 1, f"{source_name}: {[m.name for m in matches]}"
    return matches[0]


# ---------------------------------------------------------------------------
# 要件 1 / 2: 往復（識別子・時刻とも元に戻る）
# ---------------------------------------------------------------------------


def test_roundtrip_is_byte_identical(roundtrip, tmp_path):
    """仮名化 → 復元 で、元のファイルとバイト単位で一致する。"""
    src, pseudo, salt = roundtrip
    out = tmp_path / "restored"
    assert run_restore([pseudo], out, salt) == 0

    for name in ALL_FIXTURES:
        assert restored_for(out, name).read_bytes() == (src / name).read_bytes(), name


def test_roundtrip_restores_timestamps(roundtrip, tmp_path):
    """時刻も元の値に戻る（仮名化でずれていたことも併せて確かめる）。"""
    src, pseudo, salt = roundtrip
    name = "ap_metrics_20240101_0900_TZT.csv"

    def stamps(path: Path) -> list[str]:
        with open(path, newline="", encoding="utf-8") as f:
            return [r["timestamp"] for r in csv.DictReader(f)]

    assert stamps(pseudo / name) != stamps(src / name)  # 仮名化でずれている

    out = tmp_path / "restored"
    assert run_restore([pseudo], out, salt) == 0
    assert stamps(restored_for(out, name)) == stamps(src / name)


# ---------------------------------------------------------------------------
# 要件 3: --no-time
# ---------------------------------------------------------------------------


def test_no_time_restores_identifiers_only(roundtrip, tmp_path):
    src, pseudo, salt = roundtrip
    name = "ap_metrics_20240101_0900_TZT.csv"
    out = tmp_path / "restored"
    assert run_restore([pseudo], out, salt, "--no-time") == 0

    restored = restored_for(out, name)
    with open(restored, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    with open(pseudo / name, newline="", encoding="utf-8") as f:
        pseudo_rows = list(csv.DictReader(f))
    with open(src / name, newline="", encoding="utf-8") as f:
        src_rows = list(csv.DictReader(f))

    # 識別子は戻る
    assert [r["ap_name"] for r in rows] == [r["ap_name"] for r in src_rows]
    # 時刻は仮名化されたまま
    assert [r["timestamp"] for r in rows] == [r["timestamp"] for r in pseudo_rows]
    assert [r["timestamp"] for r in rows] != [r["timestamp"] for r in src_rows]
    # ファイル名の日付も動かさない（CLI の仮名化は名前を変えないので入力と同じ日付）
    assert restored.name == "ap_metrics_20240101_0900_TZT_restored.csv"


# ---------------------------------------------------------------------------
# 要件 4: 加工後のファイル（列の増減・改名・複数ファイルの結合）
# ---------------------------------------------------------------------------


def test_processed_and_merged_file_is_restored(roundtrip, tmp_path):
    """列を足し・削り・改名し、2 種別を結合したファイルでも識別子が戻る。"""
    src, pseudo, salt = roundtrip

    def read(name):
        with open(pseudo / name, newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))

    metrics = read("ap_metrics_20240101_0900_TZT.csv")
    events = read("ap_events_20240101_0900_TZT.csv")

    merged_dir = tmp_path / "merged"
    merged_dir.mkdir()
    merged = merged_dir / "analysis.csv"
    header = ["観測時刻", "装置", "拠点", "MAC", "件数"]  # 改名 + 集計列の追加
    with open(merged, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for r in metrics:
            w.writerow([r["timestamp"], r["ap_name"], r["site_name"], r["mac"], "12"])
        for r in events:
            w.writerow([r["event_timestamp"], r["ap_name"], r["site_name"], r["ap_mac"], "3"])

    out = tmp_path / "restored"
    assert run_restore([merged], out, salt) == 0
    with open(out / "analysis_restored.csv", newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    assert {r["装置"] for r in rows} == {"TEST-AP-01", "TEST-AP-02"}
    assert {r["拠点"] for r in rows} == {"TestSite Alpha"}
    assert {r["MAC"] for r in rows} == {"aabbccddee01", "aabbccddee02"}
    assert rows[0]["観測時刻"] == "2024-01-01 09:00:00"
    # 加工で生まれた列（要件 7）は素通し
    assert {r["件数"] for r in rows} == {"12", "3"}


# ---------------------------------------------------------------------------
# 要件 5: 長い仮名の優先
# ---------------------------------------------------------------------------


def test_longer_pseudonyms_win_over_shorter_prefixes(tmp_path):
    """SITE_001 と SITE_0012 が両方あっても、誤った部分一致が起きない。"""
    from pseudonymizer.restore import RestoreEngine
    from pseudonymizer.salt import generate_salt_material
    from pseudonymizer.schemas import TransformType as T
    from pseudonymizer.transforms import MappingStore

    material = generate_salt_material()
    mapping = MappingStore(salt_fingerprint=material.fingerprint)
    # SITE_NAME は %03d なので 1 -> SITE_001、12 -> SITE_012。
    # 4 桁の SITE_0012 は AP_NAME 側（%04d）と桁が違うので、ここは手で入れて確かめる。
    mapping.assignments[T.SITE_NAME] = {"Short": 1, "Longer": 12}
    mapping.assignments[T.AP_NAME] = {"ApShort": 1, "ApLonger": 12}
    engine = RestoreEngine(material, mapping, time_restore=False)

    text, counts = engine.restore_text("SITE_001,SITE_012,AP_0001,AP_0012")
    assert text == "Short,Longer,ApShort,ApLonger"
    assert counts[T.SITE_NAME.value] == 2 and counts[T.AP_NAME.value] == 2

    # マッピングに無い、より長い仮名は触らない（SITE_001 として食い込まない）
    text, _ = engine.restore_text("SITE_0019 AP_00012")
    assert text == "SITE_0019 AP_00012"


# ---------------------------------------------------------------------------
# 要件 6: 区切りの中
# ---------------------------------------------------------------------------


def test_elements_inside_separators_are_restored(roundtrip, tmp_path):
    """カンマ区切り（ap_list）と " | " 区切りの各要素が戻る。"""
    src, pseudo, salt = roundtrip
    name = "floormap_20240101_0900_TZT_summary.csv"
    out = tmp_path / "restored"
    assert run_restore([pseudo / name], out, salt) == 0

    with open(restored_for(out, name), newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["ap_list"] == "TEST-AP-01,TEST-AP-02"

    # " | " 区切り（Hang AP 分析結果のイベント列と同じ形）
    with open(pseudo / name, newline="", encoding="utf-8") as f:
        pseudo_rows = list(csv.DictReader(f))
    joined_dir = tmp_path / "joined"
    joined_dir.mkdir()
    joined = joined_dir / "events.csv"
    joined.write_text(
        "ap,発生時刻\n"
        f"{pseudo_rows[0]['ap_list'].replace(',', ' | ')},"
        f"{pseudo_rows[0]['timestamp']} | {pseudo_rows[1]['timestamp']}\n",
        encoding="utf-8",
    )
    out2 = tmp_path / "restored2"
    assert run_restore([joined], out2, salt) == 0
    text = (out2 / "events_restored.csv").read_text(encoding="utf-8")
    assert "TEST-AP-01 | TEST-AP-02" in text
    assert "2024-01-01 09:00:00 | 2024-01-01 09:00:00" in text


# ---------------------------------------------------------------------------
# 要件 7: マッピングに無い値はそのまま通す
# ---------------------------------------------------------------------------


def test_values_absent_from_the_mapping_pass_through(roundtrip, tmp_path):
    src, pseudo, salt = roundtrip
    work = tmp_path / "work"
    work.mkdir()
    target = work / "summary.csv"
    target.write_text(
        "label,mean,note\n"
        "利用率の平均,12.5,加工で生まれたラベル\n"
        "件数,3,unmapped-token-42\n",
        encoding="utf-8",
    )
    out = tmp_path / "restored"
    assert run_restore([target], out, salt) == 0
    restored = (out / "summary_restored.csv").read_text(encoding="utf-8")
    assert restored == target.read_text(encoding="utf-8")


# ---------------------------------------------------------------------------
# 要件 8: 残存の検出（値そのものは出さない）
# ---------------------------------------------------------------------------


def test_unmapped_pseudonyms_are_reported_without_values(roundtrip, tmp_path, capsys):
    """別環境のソルトで仮名化されたファイルを渡すと警告が出る。"""
    src, pseudo, salt = roundtrip
    work = tmp_path / "work"
    work.mkdir()
    foreign = work / "foreign.csv"
    # このマッピングには存在しない番号の仮名（別環境で仮名化されたファイル相当）
    foreign.write_text(
        "ap_name,site_name,mac\n"
        "AP_9999,SITE_998,029999999999\n"
        "AP_9998,SITE_998,029999999998\n",
        encoding="utf-8",
    )
    out = tmp_path / "restored"
    assert run_restore([foreign], out, salt) == 0

    err = capsys.readouterr().err
    assert "マッピングに無い仮名らしき文字列" in err
    assert "AP_NAME: 2 件" in err
    assert "列 ap_name" in err
    assert "行 2, 3" in err
    # 値そのものは出さない
    for value in ("AP_9999", "AP_9998", "SITE_998", "029999999999"):
        assert value not in err

    engine = load_engine(salt, os.path.join(os.path.dirname(salt), DEFAULT_MAP_FILENAME))
    report = engine.restore_file(foreign, tmp_path / "restored-again")
    assert report.residual_total == 6
    assert {g.kind for g in report.residuals} == {"AP_NAME", "SITE_NAME", "MAC"}
    for group in report.residuals:
        assert "9999" not in group.describe()


def test_a_clean_restore_reports_no_residuals(roundtrip, tmp_path):
    src, pseudo, salt = roundtrip
    engine = load_engine(salt, os.path.join(os.path.dirname(salt), DEFAULT_MAP_FILENAME))
    out = tmp_path / "restored"
    for name in ALL_FIXTURES:
        report = engine.restore_file(pseudo / name, out)
        assert report.residuals == [], name
        assert report.total_replacements > 0, name


# ---------------------------------------------------------------------------
# 要件 9: XLSX
# ---------------------------------------------------------------------------


def test_xlsx_cells_are_restored(roundtrip, tmp_path):
    from openpyxl import Workbook, load_workbook

    src, pseudo, salt = roundtrip
    with open(pseudo / "ap_metrics_20240101_0900_TZT.csv", newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    work = tmp_path / "work"
    work.mkdir()
    book = work / "report_20240101_0900.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["ap_name", "site_name", "timestamp", "備考"])
    for r in rows[:2]:
        ws.append([r["ap_name"], r["site_name"], r["timestamp"], "加工後のメモ"])
    wb.save(book)

    out = tmp_path / "restored"
    assert run_restore([book], out, salt) == 0
    restored = restored_for(out, "report_20240101_0900.xlsx")
    ws2 = load_workbook(restored).active
    values = [[c.value for c in row] for row in ws2.iter_rows()]
    assert {row[0] for row in values[1:]} == {"TEST-AP-01", "TEST-AP-02"}
    assert {row[1] for row in values[1:]} == {"TestSite Alpha"}
    assert values[1][2] == "2024-01-01 09:00:00"
    assert values[1][3] == "加工後のメモ"


def test_json_and_tsv_are_restored(roundtrip, tmp_path):
    """CSV 以外のテキスト形式（json / tsv）でも識別子と時刻が戻る。"""
    import json

    src, pseudo, salt = roundtrip
    with open(pseudo / "ap_metrics_20240101_0900_TZT.csv", newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    work = tmp_path / "work"
    work.mkdir()
    (work / "summary.json").write_text(
        json.dumps(
            {"ap": rows[0]["ap_name"], "at": rows[0]["timestamp"], "mac": rows[0]["mac"]},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (work / "summary.tsv").write_text(
        "ap_name\tsite_name\n" + f"{rows[0]['ap_name']}\t{rows[0]['site_name']}\n",
        encoding="utf-8",
    )

    out = tmp_path / "restored"
    assert run_restore([work], out, salt) == 0

    payload = json.loads((out / "summary_restored.json").read_text(encoding="utf-8"))
    assert payload == {
        "ap": "TEST-AP-01",
        "at": "2024-01-01 09:00:00",
        "mac": "aabbccddee01",
    }
    tsv = (out / "summary_restored.tsv").read_text(encoding="utf-8")
    assert tsv.splitlines()[1] == "TEST-AP-01\tTestSite Alpha"


# ---------------------------------------------------------------------------
# 要件 10 / 11: 非対応形式・入出力が同一
# ---------------------------------------------------------------------------


def test_unsupported_extension_is_an_error(roundtrip, tmp_path, capsys):
    src, pseudo, salt = roundtrip
    work = tmp_path / "work"
    work.mkdir()
    target = work / "notes.pdf"
    target.write_bytes(b"%PDF-1.4")
    assert run_restore([target], tmp_path / "restored", salt) == 1
    assert "対応していない形式です" in capsys.readouterr().err


def test_unsupported_extension_raises_from_the_engine(roundtrip, tmp_path):
    src, pseudo, salt = roundtrip
    engine = load_engine(salt, os.path.join(os.path.dirname(salt), DEFAULT_MAP_FILENAME))
    target = tmp_path / "work" / "notes.docx"
    target.parent.mkdir()
    target.write_bytes(b"PK\x03\x04")
    with pytest.raises(UnsupportedFormatError):
        engine.restore_file(target, tmp_path / "out")


def test_output_dir_equal_to_input_dir_is_an_error(roundtrip, tmp_path, capsys):
    src, pseudo, salt = roundtrip
    assert run_restore([pseudo], pseudo, salt) == 1
    assert "same as an input directory" in capsys.readouterr().err


def test_the_input_file_is_not_overwritten(roundtrip, tmp_path):
    src, pseudo, salt = roundtrip
    name = "ap_metrics_20240101_0900_TZT.csv"
    before = (pseudo / name).read_bytes()
    out = tmp_path / "restored"
    assert run_restore([pseudo], out, salt) == 0
    assert (pseudo / name).read_bytes() == before


# ---------------------------------------------------------------------------
# 出力の印・エラー処理
# ---------------------------------------------------------------------------


def test_output_name_marks_the_file_as_restored():
    from pseudonymizer.restore import RestoreEngine
    from pseudonymizer.salt import SaltMaterial
    from pseudonymizer.transforms import MappingStore

    material = SaltMaterial(salt=b"\x01" * 32, time_offset_seconds=-86400 * 100,
                            created_at="2024-01-01T00:00:00+00:00")
    engine = RestoreEngine(material, MappingStore(salt_fingerprint=material.fingerprint))

    name, shifted = engine.restore_name("ap_metrics_20250101_0900_JST_pseudonymized.csv")
    assert name == "ap_metrics_20250411_0900_JST_restored.csv"
    assert shifted == 1
    # 日付の無い名前でも印は付く（エラーにしない）
    assert engine.restore_name("merged.csv")[0] == "merged_restored.csv"
    # すでに印が付いていれば重ねない
    assert engine.restore_name("merged_restored.csv")[0] == "merged_restored.csv"


def test_missing_salt_or_mapping_is_a_clear_error(tmp_path):
    with pytest.raises(RestoreError, match="ソルトファイルが見つかりません"):
        load_engine(str(tmp_path / "nope.json"), str(tmp_path / DEFAULT_MAP_FILENAME))


def test_mapping_from_a_different_salt_is_rejected(roundtrip, tmp_path):
    """別環境のソルトとマッピングの取り違えは、静かに壊れず落ちる。"""
    from pseudonymizer.transforms import PseudonymizeError

    src, pseudo, salt = roundtrip
    other = canonical_copy(tmp_path / "src2", names=ALL_FIXTURES[:1])
    other_out = tmp_path / "pseudo2"
    assert cli.main([str(other), "--out", str(other_out), *PSEUDONYMIZE_ARGS]) == 0

    with pytest.raises(PseudonymizeError, match="different salt"):
        load_engine(salt, str(other_out / DEFAULT_MAP_FILENAME))
