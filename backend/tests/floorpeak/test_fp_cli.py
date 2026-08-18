"""floorpeak analyze CLI の配線テスト。合成データのみを使う。"""
from __future__ import annotations

import csv
from datetime import datetime, timedelta
from pathlib import Path

import _fpsynth as S
import pytest
from openpyxl import load_workbook

from floorpeak import analysis, cli

START = datetime(2026, 1, 1, 10, 0, 0)
INTERVAL = 300


def _ap(i: int) -> dict[str, str]:
    return {"ap_id": f"test-ap-{i:04d}", "ap_name": f"TEST-AP-{i:02d}", "mac": f"aabbccddee{i:02d}"}


def write_logs(logs_dir: Path) -> Path:
    """1F に 2 台・2F に 1 台。ピークは 10:05。"""
    logs_dir.mkdir(parents=True, exist_ok=True)
    rows = []
    rows += S.series(START, INTERVAL, [3, 9, 2], map_id=S.MAP_1F, model="AP45", **_ap(1))
    rows += S.series(START, INTERVAL, [1, 5, 1], map_id=S.MAP_1F, model="AP63E", **_ap(2))
    rows += S.series(START, INTERVAL, [0, 7, 0], map_id=S.MAP_2F, model="AP32", **_ap(3))
    S.write_metrics(logs_dir / "ap_metrics_20260101_1000_TZT.csv", rows)
    S.default_floormap(logs_dir, START + timedelta(minutes=5))
    return logs_dir


def _run(argv: list[str]) -> int:
    return cli.main(argv)


def test_analyze_writes_all_three_files(tmp_path, capsys):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run(["analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out)]) == cli.EXIT_OK

    assert len(list(out.glob("*.xlsx"))) == 1
    assert len(list(out.glob("*.csv"))) == 1
    assert len(list(out.glob("*_summary.txt"))) == 1

    printed = capsys.readouterr().out
    assert "ピーク時点" in printed
    assert S.FLOOR_1F in printed


def test_result_csv_columns_and_rows(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run(["analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out)]) == cli.EXIT_OK

    with open(next(out.glob("*.csv")), newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    assert list(rows[0].keys()) == list(analysis.RESULT_COLUMNS)
    # サイトの全 AP を保存する（トップ 20 の切り出しは表示側の責務）
    assert len(rows) == 3
    by_name = {r["ap_name"]: r for r in rows}
    assert by_name["TEST-AP-01"]["num_clients"] == "9"
    assert by_name["TEST-AP-01"]["map_name"] == S.FLOOR_1F
    assert by_name["TEST-AP-01"]["rank_in_floor"] == "1"
    assert by_name["TEST-AP-03"]["map_name"] == S.FLOOR_2F
    assert by_name["TEST-AP-03"]["rank_in_floor"] == "1"


def test_site_is_required(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run(["analyze", "--logs", str(logs), "--out", str(out)]) == cli.EXIT_INPUT_ERROR


def test_unknown_site_is_input_error(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    rc = _run(["analyze", "--logs", str(logs), "--site", "no-such-site", "--out", str(out)])
    assert rc == cli.EXIT_INPUT_ERROR


@pytest.mark.parametrize("text", ["2026-01-01 09:00", "2026-01-01T09:00:00"])
def test_time_formats_accepted(tmp_path, text):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    rc = _run(["analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out), "--from", text])
    assert rc == cli.EXIT_OK


def test_timezone_aware_time_is_rejected(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    rc = _run([
        "analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out),
        "--from", "2026-01-01T09:00:00+09:00",
    ])
    assert rc == cli.EXIT_INPUT_ERROR


def test_to_before_from_is_rejected(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    rc = _run([
        "analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out),
        "--from", "2026-01-01 11:00", "--to", "2026-01-01 10:00",
    ])
    assert rc == cli.EXIT_INPUT_ERROR


def test_out_must_not_be_the_logs_dir(tmp_path):
    logs = write_logs(tmp_path / "logs")
    assert _run(["analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(logs)]) == cli.EXIT_OUTPUT_ERROR


def test_at_overrides_the_window(tmp_path, capsys):
    """--at を指定したら --from / --to は無視する（README に書いた挙動）。"""
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    rc = _run([
        "analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out),
        "--from", "2026-01-01 10:00", "--to", "2026-01-01 10:05",
        "--at", "2026-01-01 10:10",
    ])
    assert rc == cli.EXIT_OK
    printed = capsys.readouterr().out
    assert "selected_by=manual" in printed
    assert "ピーク時刻=2026-01-01 10:10:00" in printed
    assert "期間の指定は無視" in printed


def test_unknown_floor_is_rejected(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    rc = _run([
        "analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out),
        "--floor", "No Such Floor",
    ])
    assert rc == cli.EXIT_INPUT_ERROR


# ---------------------------------------------------------------------------
# 12. xlsx の構造
# ---------------------------------------------------------------------------


def test_xlsx_has_chart_and_data_sheets(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    assert _run(["analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out)]) == cli.EXIT_OK

    wb = load_workbook(next(out.glob("*.xlsx")))
    assert wb.sheetnames == ["chart", "data"]

    charts = wb["chart"]._charts
    assert len(charts) == 1
    assert charts[0].type == "bar"          # 横棒
    assert charts[0].legend is None          # 凡例はセル側で自作する
    assert len(charts[0].series) == 1        # 単一系列 + DataPoint で色分け
    # このフロアの 2 台ぶんの色が付いている
    assert len(charts[0].series[0].data_points) == 2
    assert charts[0].series[0].data_points[0].graphicalProperties.solidFill is not None

    data = wb["data"]
    # 1〜3 行目がメタ、5 行目がヘッダー、6 行目以降が全フロアの全行
    assert [c.value for c in data[5]] == list(analysis.RESULT_COLUMNS)
    assert data.max_row == 5 + 3


def test_xlsx_chart_floor_can_be_chosen(tmp_path):
    logs = write_logs(tmp_path / "logs")
    out = tmp_path / "out"
    rc = _run([
        "analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out),
        "--floor", S.FLOOR_2F,
    ])
    assert rc == cli.EXIT_OK

    ws = load_workbook(next(out.glob("*.xlsx")))["chart"]
    assert S.FLOOR_2F in str(ws.cell(row=4, column=1).value)
    assert len(ws._charts[0].series[0].data_points) == 1


def test_unknown_model_falls_back_to_grey(tmp_path):
    """辞書に無いモデルでも落ちないこと（実サイトのモデル構成は変わりうる）。"""
    logs = tmp_path / "logs"
    logs.mkdir()
    rows = S.series(START, INTERVAL, [3, 9], map_id=S.MAP_1F, model="AP-NOT-IN-DICT", **_ap(1))
    S.write_metrics(logs / "ap_metrics_20260101_1000_TZT.csv", rows)
    S.default_floormap(logs, START)

    out = tmp_path / "out"
    assert _run(["analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out)]) == cli.EXIT_OK
    ws = load_workbook(next(out.glob("*.xlsx")))["chart"]
    fill = ws._charts[0].series[0].data_points[0].graphicalProperties.solidFill
    assert fill.srgbClr == analysis.DEFAULT_MODEL_COLOR
