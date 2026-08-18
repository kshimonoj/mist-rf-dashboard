"""フロア別ピーク時点分析 API（/api/floorpeak）のテスト。

合成データのみを使う。このテストは「API が CLI と同じ結果を返すこと」を要に
している。ロジックを API 側で書き直すと真っ先にここが落ちる。
"""
from __future__ import annotations

import time
from datetime import datetime, timedelta
from pathlib import Path

import _fpsynth as S
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from floorpeak import analysis, cli
from floorpeak.analysis import RESULT_COLUMNS
from routers import floorpeak as api

START = datetime(2026, 1, 1, 10, 0, 0)
INTERVAL = 300


def _ap(i: int) -> dict[str, str]:
    return {"ap_id": f"test-ap-{i:04d}", "ap_name": f"TEST-AP-{i:02d}", "mac": f"aabbccddee{i:02d}"}


def write_logs(logs_dir: Path) -> None:
    """1F に 2 台・2F に 1 台。ピークは 10:05。"""
    rows = []
    rows += S.series(START, INTERVAL, [3, 9, 2], map_id=S.MAP_1F, model="AP45", **_ap(1))
    rows += S.series(START, INTERVAL, [1, 5, 1], map_id=S.MAP_1F, model="AP63E", **_ap(2))
    rows += S.series(START, INTERVAL, [0, 7, 0], map_id=S.MAP_2F, model="AP32", **_ap(3))
    S.write_metrics(logs_dir / "ap_metrics_20260101_1000_TZT.csv", rows)
    S.default_floormap(logs_dir, START + timedelta(minutes=5))


@pytest.fixture
def api_client(tmp_path, monkeypatch):
    """LOGS_DIR / RESULTS_DIR を隔離したディレクトリに向けた TestClient を返す。"""
    logs_dir = tmp_path / "logs"
    logs_dir.mkdir()
    monkeypatch.setattr(api, "LOGS_DIR", str(logs_dir))
    # 保存先は logs_dir の外に置く（配下に置くと次の分析が自分の出力を読む）
    monkeypatch.setattr(api, "RESULTS_DIR", str(tmp_path / "floorpeak_results"))

    _clear_jobs()
    app = FastAPI()
    app.include_router(api.router)
    with TestClient(app) as client:
        yield client, logs_dir
    _clear_jobs()


def _clear_jobs() -> None:
    with api._LOCK:
        for job in list(api._JOBS.values()):
            api._discard(job)
        api._JOBS.clear()


def _analyze(client, body: dict | None = None) -> dict:
    r = client.post("/api/floorpeak/analyze", json=body or {"site": S.SITE_ID})
    assert r.status_code == 202, r.text
    job_id = r.json()["job_id"]
    for _ in range(200):
        state = client.get(f"/api/floorpeak/jobs/{job_id}").json()
        if state["status"] != "running":
            return state
        time.sleep(0.05)
    raise AssertionError("ジョブが終わりません")


def _done(client, body: dict | None = None) -> dict:
    state = _analyze(client, body)
    assert state["status"] == "done", state.get("error")
    return state


# ---------------------------------------------------------------------------
# 基本
# ---------------------------------------------------------------------------


def test_analyze_returns_rows_meta_and_warnings(api_client):
    client, logs_dir = api_client
    write_logs(logs_dir)
    state = _done(client)

    body = client.get(f"/api/floorpeak/jobs/{state['job_id']}/result").json()
    assert body["columns"] == list(RESULT_COLUMNS)
    assert list(body["rows"][0].keys()) == list(RESULT_COLUMNS)
    assert len(body["rows"]) == 3

    meta = body["meta"]
    assert meta["site_id"] == S.SITE_ID
    assert meta["site_name"] == S.SITE_NAME
    assert meta["selected_by"] == "auto"
    assert meta["peak_time"] == "2026-01-01 10:05:00"
    assert meta["peak_total_clients"] == 21
    assert meta["bucket_seconds"] == 300
    assert meta["floor_count"] == 2
    assert meta["default_floor"] == S.FLOOR_1F
    assert meta["floormap_file"].endswith("_summary.csv")
    assert meta["floormap_offset_seconds"] == 0
    assert meta["top_n"] == analysis.TOP_N
    # 色分けの定義はバックエンドが返す（フロントで定義し直さない）
    assert meta["model_colors"]["AP63E"] == analysis.MODEL_COLORS["AP63E"]
    assert meta["default_model_color"] == analysis.DEFAULT_MODEL_COLOR
    assert isinstance(body["warnings"], list)


def test_site_is_required(api_client):
    client, logs_dir = api_client
    write_logs(logs_dir)
    r = client.post("/api/floorpeak/analyze", json={})
    assert r.status_code == 400
    assert "site" in r.json()["detail"]


def test_unknown_body_field_is_rejected(api_client):
    client, logs_dir = api_client
    write_logs(logs_dir)
    r = client.post("/api/floorpeak/analyze", json={"site": S.SITE_ID, "sites": ["x"]})
    assert r.status_code == 400


def test_to_before_from_is_rejected(api_client):
    client, logs_dir = api_client
    write_logs(logs_dir)
    r = client.post("/api/floorpeak/analyze", json={
        "site": S.SITE_ID, "from": "2026-01-01 11:00", "to": "2026-01-01 10:00",
    })
    assert r.status_code == 400


def test_unknown_site_fails_the_job(api_client):
    """「対象が無かった」は done ではなく failed（0 件と取り違えない）。"""
    client, logs_dir = api_client
    write_logs(logs_dir)
    state = _analyze(client, {"site": "no-such-site"})
    assert state["status"] == "failed"
    assert "見つかりません" in state["error"]


def test_window_is_half_open_through_the_api(api_client):
    client, logs_dir = api_client
    write_logs(logs_dir)
    state = _done(client, {"site": S.SITE_ID, "to": "2026-01-01 10:05:00"})
    assert state["meta"]["peak_time"] == "2026-01-01 10:00:00"


def test_manual_at_through_the_api(api_client):
    client, logs_dir = api_client
    write_logs(logs_dir)
    state = _done(client, {"site": S.SITE_ID, "at": "2026-01-01 10:10:00"})
    assert state["meta"]["selected_by"] == "manual"
    assert state["meta"]["peak_time"] == "2026-01-01 10:10:00"


def test_sites_endpoint_lists_log_sites(api_client):
    client, logs_dir = api_client
    write_logs(logs_dir)
    body = client.get("/api/floorpeak/sites?refresh=true").json()
    assert [s["site_id"] for s in body["sites"]] == [S.SITE_ID]
    assert body["sites"][0]["ap_count"] == 3


def test_second_job_conflicts(api_client, monkeypatch):
    client, logs_dir = api_client
    write_logs(logs_dir)
    # 走り続けるジョブを 1 本仕込む
    started = api._Job(job_id="stuck", params=analysis.AnalysisParams(site=S.SITE_ID),
                       started_at=api._now())
    with api._LOCK:
        api._JOBS[started.job_id] = started
    r = client.post("/api/floorpeak/analyze", json={"site": S.SITE_ID})
    assert r.status_code == 409
    assert r.json()["detail"]["job_id"] == "stuck"


# ---------------------------------------------------------------------------
# 10. ダウンロード（CLI の出力と同一）
# ---------------------------------------------------------------------------


def test_download_csv_matches_cli_output_byte_for_byte(api_client, tmp_path):
    client, logs_dir = api_client
    write_logs(logs_dir)
    job_id = _done(client)["job_id"]

    out_dir = tmp_path / "cli_out"
    assert cli.main([
        "analyze", "--logs", str(logs_dir), "--site", S.SITE_ID,
        "--out", str(out_dir), "--format", "both",
    ]) == cli.EXIT_OK
    cli_csv = next(out_dir.glob("*.csv")).read_bytes()

    r = client.get(f"/api/floorpeak/jobs/{job_id}/download?format=csv")
    assert r.status_code == 200
    assert r.content == cli_csv


def test_download_xlsx_matches_cli_output(api_client, tmp_path):
    client, logs_dir = api_client
    write_logs(logs_dir)
    job_id = _done(client)["job_id"]

    out_dir = tmp_path / "cli_out"
    assert cli.main([
        "analyze", "--logs", str(logs_dir), "--site", S.SITE_ID, "--out", str(out_dir),
    ]) == cli.EXIT_OK

    r = client.get(f"/api/floorpeak/jobs/{job_id}/download?format=xlsx")
    assert r.status_code == 200
    api_xlsx = tmp_path / "api_result.xlsx"
    api_xlsx.write_bytes(r.content)

    for sheet in ("chart", "data"):
        expected = load_workbook(next(out_dir.glob("*.xlsx")))[sheet]
        actual = load_workbook(api_xlsx)[sheet]
        assert (actual.max_row, actual.max_column) == (expected.max_row, expected.max_column)
        for row in range(1, expected.max_row + 1):
            for col in range(1, expected.max_column + 1):
                assert actual.cell(row=row, column=col).value == expected.cell(row=row, column=col).value


def test_download_xlsx_for_a_chosen_floor(api_client, tmp_path):
    client, logs_dir = api_client
    write_logs(logs_dir)
    job_id = _done(client)["job_id"]

    r = client.get(f"/api/floorpeak/jobs/{job_id}/download?format=xlsx&floor={S.FLOOR_2F}")
    assert r.status_code == 200
    path = tmp_path / "floor2.xlsx"
    path.write_bytes(r.content)

    wb = load_workbook(path)
    assert S.FLOOR_2F in str(wb["chart"].cell(row=4, column=1).value)
    # data シートには全フロアが入る（グラフだけがフロアごと）
    assert wb["data"].max_row == 4 + 3


def test_download_xlsx_for_unknown_floor_is_400(api_client):
    client, logs_dir = api_client
    write_logs(logs_dir)
    job_id = _done(client)["job_id"]
    r = client.get(f"/api/floorpeak/jobs/{job_id}/download?format=xlsx&floor=No%20Such%20Floor")
    assert r.status_code == 400


# ---------------------------------------------------------------------------
# 保存済み結果（再分析はしない）
# ---------------------------------------------------------------------------


def test_done_job_is_archived_and_can_be_read_back(api_client):
    client, logs_dir = api_client
    write_logs(logs_dir)
    live = client.get(f"/api/floorpeak/jobs/{_done(client)['job_id']}/result").json()

    results = client.get("/api/floorpeak/results").json()["results"]
    assert len(results) == 1
    name = results[0]["name"]
    assert results[0]["peak_total_clients"] == 21
    assert results[0]["site_name"] == S.SITE_NAME

    saved = client.get(f"/api/floorpeak/results/{name}/rows").json()
    assert saved["job_id"] is None and saved["name"] == name
    assert saved["columns"] == live["columns"]
    assert saved["rows"] == live["rows"]
    assert saved["meta"]["peak_time"] == live["meta"]["peak_time"]


def test_saved_result_download_and_delete(api_client):
    client, logs_dir = api_client
    write_logs(logs_dir)
    _done(client)
    name = client.get("/api/floorpeak/results").json()["results"][0]["name"]

    for fmt in ("xlsx", "csv"):
        r = client.get(f"/api/floorpeak/results/{name}/download?format={fmt}")
        assert r.status_code == 200 and len(r.content) > 0

    r = client.get(f"/api/floorpeak/results/{name}/download?format=xlsx&floor={S.FLOOR_2F}")
    assert r.status_code == 200

    assert client.delete(f"/api/floorpeak/results/{name}").status_code == 200
    assert client.get("/api/floorpeak/results").json()["results"] == []


def test_invalid_saved_name_is_rejected(api_client):
    client, _ = api_client
    r = client.get("/api/floorpeak/results/..%2F..%2Fetc/rows")
    assert r.status_code in (400, 404)


def test_results_dir_is_not_read_as_input(api_client):
    """保存済みの結果を置いても、次の分析がそれを入力として拾わないこと。"""
    client, logs_dir = api_client
    write_logs(logs_dir)
    _done(client)
    first = client.get("/api/floorpeak/results").json()["results"][0]

    state = _done(client)
    assert state["meta"]["peak_total_clients"] == first["peak_total_clients"]
    assert state["meta"]["ap_count"] == first["ap_count"]
