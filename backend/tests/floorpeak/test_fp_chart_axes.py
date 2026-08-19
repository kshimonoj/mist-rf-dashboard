"""xlsx グラフ（chart シート）の描画属性テスト。合成データのみを使う。

openpyxl は保存時、``chart.height`` / ``chart.width`` (cm) を描画アンカー
（``drawing*.xml`` の ``<xdr:ext>``）の EMU 値へ正しく書き出す。しかし
``load_workbook()`` で読み戻した ``chart.height`` / ``chart.width`` は
チャート本体の XML（``chart*.xml``）に物理サイズの情報が無いため、常に
クラス既定値（7.5cm / 15cm）を返す ―― 実際に保存された値では **ない**。
そのため本数依存の検証だけは ``chart.anchor.ext``（drawing 側の EMU）を
直接読む。他の属性（delete / orientation / title / dLbls / gapWidth）は
chart XML 側の情報なので ``load_workbook()`` の値をそのまま検証できる。
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import _fpsynth as S
from openpyxl import load_workbook

from floorpeak import cli

START = datetime(2026, 1, 1, 10, 0, 0)
INTERVAL = 300
EMU_PER_CM = 360000


def _write_n_ap_logs(logs_dir: Path, n: int) -> Path:
    """1 フロアに AP を n 台配置する（本数依存の描画検証専用）。"""
    logs_dir.mkdir(parents=True, exist_ok=True)
    rows = []
    ap_names = []
    for i in range(1, n + 1):
        ap_name = f"TEST-AP-{i:03d}"
        ap_names.append(ap_name)
        rows += S.series(
            START, INTERVAL, [i],
            ap_id=f"test-ap-{i:04d}", ap_name=ap_name, mac=f"aabbccdd{i:04x}",
            map_id=S.MAP_1F, model="AP45",
        )
    S.write_metrics(logs_dir / "ap_metrics_20260101_1000_TZT.csv", rows)
    S.write_floormap(
        logs_dir, START,
        [S.floormap_row(START, map_name=S.FLOOR_1F, band="5", channel=36, ap_list=ap_names)],
    )
    return logs_dir


def _build_chart(tmp_path: Path, n: int):
    logs = _write_n_ap_logs(tmp_path / f"logs{n}", n)
    out = tmp_path / f"out{n}"
    rc = cli.main(["analyze", "--logs", str(logs), "--site", S.SITE_ID, "--out", str(out)])
    assert rc == cli.EXIT_OK
    wb = load_workbook(next(out.glob("*.xlsx")))
    return wb["chart"]._charts[0]


def test_axes_are_not_deleted(tmp_path):
    chart = _build_chart(tmp_path, 4)
    assert chart.x_axis.delete is False
    assert chart.y_axis.delete is False


def test_orientation_puts_rank_1_on_top_and_value_axis_left_to_right(tmp_path):
    chart = _build_chart(tmp_path, 4)
    # 横棒（type="bar"）ではカテゴリ軸が x_axis、値軸が y_axis
    assert chart.x_axis.scaling.orientation == "maxMin"
    assert chart.y_axis.scaling.orientation == "minMax"


def test_axis_titles_are_removed(tmp_path):
    chart = _build_chart(tmp_path, 4)
    assert chart.x_axis.title is None
    assert chart.y_axis.title is None


def test_data_labels_show_value_only(tmp_path):
    chart = _build_chart(tmp_path, 4)
    assert chart.dLbls.showSerName is False
    assert chart.dLbls.showCatName is False
    assert chart.dLbls.showLegendKey is False
    assert chart.dLbls.showVal is True


def test_gap_width_is_40(tmp_path):
    chart = _build_chart(tmp_path, 4)
    assert chart.gapWidth == 40


def test_chart_height_scales_with_bar_count(tmp_path):
    """本数に応じて高さが変わることを、保存された描画アンカー(EMU)で検証する。

    固定値への回帰を防ぐため、値そのものではなく「4 本より 20 本の方が高い」
    ことだけを見る。
    """
    chart4 = _build_chart(tmp_path, 4)
    chart20 = _build_chart(tmp_path, 20)

    height4_cm = chart4.anchor.ext.cy / EMU_PER_CM
    height20_cm = chart20.anchor.ext.cy / EMU_PER_CM

    assert height4_cm != height20_cm
    assert height20_cm > height4_cm
